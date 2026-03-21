from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from contacts.capabilities import ContactCapabilityType
from contacts.models import ContactType

from .canonical_dataset import BeContactDataset

CANONICAL_BE_SHEET_NAMES = ("2024", "2025", "2026")
CANONICAL_DESTINATION_CORRESPONDENT_OVERRIDES = {
    "BEY": "Tony MDAWAR",
    "BGF": "Christian LIMBIO",
    "NDJ": "Geovanie Kamtar NDANGMBAYE",
}
CANONICAL_FALLBACK_CORRESPONDENT_NAME = "Correspondant non renseigne"
PERSON_PREFIXES = {
    "dr",
    "madame",
    "mme",
    "monsieur",
    "mr",
    "pere",
    "père",
    "soeur",
    "sr",
}
EXPEDITEUR_TRANSPORTER_KEY = "expediteur"


def _normalize_text(value) -> str:
    return str(value or "").strip()


def _normalize_key(value) -> str:
    text = unicodedata.normalize("NFKD", _normalize_text(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _normalize_display_name(value) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()


def _normalize_iata(value) -> str:
    return _normalize_text(value).upper()


def _selected_sheet_names(workbook) -> list[str]:
    names = [name for name in CANONICAL_BE_SHEET_NAMES if name in workbook.sheetnames]
    if names:
        return names
    return [workbook.active.title]


def _iter_sheet_rows(sheet):
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, ())
    header_map: dict[str, list[int]] = {}
    for index, header in enumerate(headers or ()):
        key = _normalize_text(header)
        if not key:
            continue
        header_map.setdefault(key, []).append(index)
    for row in rows:
        if not any(_normalize_text(value) for value in row):
            continue
        yield row, header_map


def _get_cell(row, header_map, name: str) -> str:
    for index in header_map.get(name, []):
        value = _normalize_text(row[index])
        if value:
            return value
    return ""


def _note_conflict(
    *,
    review_items: list[dict],
    seen_conflicts: set[tuple[str, str, str]],
    entity_type: str,
    key: str,
    display_name: str,
    field_name: str,
    existing,
    incoming,
):
    marker = (entity_type, key, field_name)
    if marker in seen_conflicts:
        return
    seen_conflicts.add(marker)
    review_items.append(
        {
            "entity_type": entity_type,
            "key": key,
            "display_name": display_name or key,
            "reason": f"conflicting {entity_type} {field_name}",
            "existing": existing,
            "incoming": incoming,
        }
    )


def _note_missing(
    *,
    review_items: list[dict],
    seen_conflicts: set[tuple[str, str, str]],
    entity_type: str,
    key: str,
    display_name: str,
    field_name: str,
):
    marker = (entity_type, key, field_name)
    if marker in seen_conflicts:
        return
    seen_conflicts.add(marker)
    review_items.append(
        {
            "entity_type": entity_type,
            "key": key,
            "display_name": display_name or key,
            "reason": f"missing {entity_type} {field_name}",
            "existing": "",
            "incoming": "",
        }
    )


def _merge_scalar(
    *,
    record: dict,
    field_name: str,
    value: str,
    source_priority: int,
    review_items: list[dict],
    seen_conflicts: set[tuple[str, str, str]],
    entity_type: str,
    key: str,
):
    if value in ("", None):
        return
    existing = record.get(field_name, "")
    field_priorities = record.setdefault("_field_priorities", {})
    existing_priority = field_priorities.get(field_name, -1)
    if not existing:
        record[field_name] = value
        field_priorities[field_name] = source_priority
        return
    if source_priority > existing_priority:
        record[field_name] = value
        field_priorities[field_name] = source_priority
        return
    if source_priority < existing_priority:
        return
    if _normalize_key(existing) == _normalize_key(value):
        if len(value) > len(existing):
            record[field_name] = value
        return
    _note_conflict(
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        entity_type=entity_type,
        key=key,
        display_name=(record.get("name") or record.get("city") or key),
        field_name=field_name,
        existing=existing,
        incoming=value,
    )


def _get_or_create_entity(store: dict[str, dict], *, key: str) -> dict:
    if key not in store:
        store[key] = {
            "key": key,
            "_field_priorities": {},
        }
    return store[key]


def _clean_record(record: dict) -> dict:
    return {key: value for key, value in record.items() if not key.startswith("_")}


def _org_contact_key(name: str) -> str:
    return f"org:{_normalize_key(name)}"


def _person_contact_key(name: str, *, organization_contact_key: str = "") -> str:
    base = _normalize_key(name)
    if organization_contact_key:
        return f"person:{organization_contact_key}:{base}"
    return f"person:{base}"


def _looks_like_person(name: str) -> bool:
    first_token = _normalize_key(name).split(" ", 1)[0]
    return first_token in PERSON_PREFIXES


def _person_payload(
    *,
    title: str = "",
    first_name: str = "",
    last_name: str = "",
    fallback_name: str,
) -> dict:
    normalized_title = _normalize_display_name(title)
    normalized_first_name = _normalize_display_name(first_name)
    normalized_last_name = _normalize_display_name(last_name)
    email = ""
    if "@" in normalized_last_name and not normalized_first_name:
        email = normalized_last_name
        normalized_last_name = ""
    elif "@" in normalized_first_name and not normalized_last_name:
        email = normalized_first_name
        normalized_first_name = ""
    full_name = " ".join(
        part for part in (normalized_first_name, normalized_last_name) if part
    ).strip()
    if not full_name:
        full_name = fallback_name
    return {
        "title": normalized_title,
        "first_name": normalized_first_name,
        "last_name": normalized_last_name,
        "name": full_name,
        "email": email,
    }


def _ensure_contact_record(
    *,
    contacts: dict[str, dict],
    contact_key: str,
    name: str,
    contact_type: str,
    source_priority: int,
    review_items: list[dict],
    seen_conflicts: set[tuple[str, str, str]],
    title: str = "",
    first_name: str = "",
    last_name: str = "",
    organization_key: str = "",
    email: str = "",
):
    record = _get_or_create_entity(contacts, key=contact_key)
    _merge_scalar(
        record=record,
        field_name="name",
        value=name,
        source_priority=source_priority,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        entity_type="contact",
        key=contact_key,
    )
    _merge_scalar(
        record=record,
        field_name="contact_type",
        value=contact_type,
        source_priority=source_priority,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        entity_type="contact",
        key=contact_key,
    )
    _merge_scalar(
        record=record,
        field_name="title",
        value=title,
        source_priority=source_priority,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        entity_type="contact",
        key=contact_key,
    )
    _merge_scalar(
        record=record,
        field_name="first_name",
        value=first_name,
        source_priority=source_priority,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        entity_type="contact",
        key=contact_key,
    )
    _merge_scalar(
        record=record,
        field_name="last_name",
        value=last_name,
        source_priority=source_priority,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        entity_type="contact",
        key=contact_key,
    )
    _merge_scalar(
        record=record,
        field_name="organization_key",
        value=organization_key,
        source_priority=source_priority,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        entity_type="contact",
        key=contact_key,
    )
    _merge_scalar(
        record=record,
        field_name="email",
        value=email,
        source_priority=source_priority,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        entity_type="contact",
        key=contact_key,
    )
    return record


def _ensure_org_contact(
    *,
    contacts: dict[str, dict],
    organization_name: str,
    source_priority: int,
    review_items: list[dict],
    seen_conflicts: set[tuple[str, str, str]],
) -> str:
    contact_key = _org_contact_key(organization_name)
    _ensure_contact_record(
        contacts=contacts,
        contact_key=contact_key,
        name=organization_name,
        contact_type=ContactType.ORGANIZATION,
        source_priority=source_priority,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
    )
    return contact_key


def _ensure_person_contact(
    *,
    contacts: dict[str, dict],
    source_priority: int,
    review_items: list[dict],
    seen_conflicts: set[tuple[str, str, str]],
    organization_contact_key: str = "",
    title: str = "",
    first_name: str = "",
    last_name: str = "",
    fallback_name: str,
) -> tuple[str, str]:
    payload = _person_payload(
        title=title,
        first_name=first_name,
        last_name=last_name,
        fallback_name=fallback_name,
    )
    contact_key = _person_contact_key(
        payload["name"],
        organization_contact_key=organization_contact_key,
    )
    _ensure_contact_record(
        contacts=contacts,
        contact_key=contact_key,
        name=payload["name"],
        contact_type=ContactType.PERSON,
        source_priority=source_priority,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        title=payload["title"],
        first_name=payload["first_name"],
        last_name=payload["last_name"],
        organization_key=organization_contact_key,
        email=payload["email"],
    )
    return contact_key, _normalize_key(payload["name"])


def _register_generic_capability_contact(
    *,
    contacts: dict[str, dict],
    registry: dict[str, dict],
    name: str,
    capability: str,
    source_priority: int,
    review_items: list[dict],
    seen_conflicts: set[tuple[str, str, str]],
    default_contact_type: str,
):
    plain_key = _normalize_key(name)
    if default_contact_type == ContactType.PERSON or _looks_like_person(name):
        tokens = name.split(maxsplit=1)
        first_name = tokens[0] if len(tokens) == 2 else ""
        last_name = tokens[1] if len(tokens) == 2 else name
        contact_key, _person_name_key = _ensure_person_contact(
            contacts=contacts,
            source_priority=source_priority,
            review_items=review_items,
            seen_conflicts=seen_conflicts,
            title="",
            first_name=first_name,
            last_name=last_name,
            fallback_name=name,
        )
        contact_type = ContactType.PERSON
    else:
        contact_key = _ensure_org_contact(
            contacts=contacts,
            organization_name=name,
            source_priority=source_priority,
            review_items=review_items,
            seen_conflicts=seen_conflicts,
        )
        contact_type = ContactType.ORGANIZATION

    record = _get_or_create_entity(registry, key=plain_key)
    _merge_scalar(
        record=record,
        field_name="name",
        value=name,
        source_priority=source_priority,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        entity_type=capability,
        key=plain_key,
    )
    _merge_scalar(
        record=record,
        field_name="contact_key",
        value=contact_key,
        source_priority=source_priority,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        entity_type=capability,
        key=plain_key,
    )
    _merge_scalar(
        record=record,
        field_name="contact_type",
        value=contact_type,
        source_priority=source_priority,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        entity_type=capability,
        key=plain_key,
    )


def _apply_destination_correspondent_overrides(
    *,
    destinations: dict[str, dict],
    correspondents: dict[str, dict],
):
    for destination in destinations.values():
        iata_code = _normalize_iata(destination.get("iata_code"))
        override_name = CANONICAL_DESTINATION_CORRESPONDENT_OVERRIDES.get(iata_code, "")
        if not override_name:
            continue
        override_key = _normalize_key(override_name)
        destination["correspondent_key"] = override_key
        correspondent = _get_or_create_entity(correspondents, key=override_key)
        correspondent.setdefault("name", override_name)
        if destination.get("country"):
            correspondent.setdefault("country", destination["country"])


def _ensure_destination_required_fields(
    *,
    destinations: dict[str, dict],
    correspondents: dict[str, dict],
    review_items: list[dict],
    seen_conflicts: set[tuple[str, str, str]],
):
    fallback_key = _normalize_key(CANONICAL_FALLBACK_CORRESPONDENT_NAME)
    for destination in destinations.values():
        display_name = destination.get("city") or destination.get("iata_code") or destination["key"]
        if "country" not in destination:
            destination["country"] = ""
            _note_missing(
                review_items=review_items,
                seen_conflicts=seen_conflicts,
                entity_type="destination",
                key=destination["key"],
                display_name=display_name,
                field_name="country",
            )
        if destination.get("correspondent_key"):
            continue
        destination["correspondent_key"] = fallback_key
        fallback = _get_or_create_entity(correspondents, key=fallback_key)
        fallback.setdefault("name", CANONICAL_FALLBACK_CORRESPONDENT_NAME)
        _note_missing(
            review_items=review_items,
            seen_conflicts=seen_conflicts,
            entity_type="destination",
            key=destination["key"],
            display_name=display_name,
            field_name="correspondent_key",
        )


def _prune_unreferenced_correspondents(
    *,
    destinations: dict[str, dict],
    correspondents: dict[str, dict],
) -> dict[str, dict]:
    referenced_keys = {
        destination.get("correspondent_key", "")
        for destination in destinations.values()
        if destination.get("correspondent_key")
    }
    return {key: record for key, record in correspondents.items() if key in referenced_keys}


def _correspondent_placeholder_org_name(destination: dict) -> str:
    label = destination.get("iata_code") or destination.get("city") or "destination"
    return f"Correspondant {label}"


def _finalize_destination_correspondents(
    *,
    contacts: dict[str, dict],
    recipients: dict[str, dict],
    correspondents: dict[str, dict],
    destinations: dict[str, dict],
    source_priority: int,
    review_items: list[dict],
    seen_conflicts: set[tuple[str, str, str]],
):
    recipients_by_destination: dict[str, list[dict]] = {}
    for recipient in recipients.values():
        destination_iata = recipient.get("destination_iata", "")
        if destination_iata:
            recipients_by_destination.setdefault(destination_iata, []).append(recipient)

    for destination in destinations.values():
        correspondent_key = destination["correspondent_key"]
        correspondent = _get_or_create_entity(correspondents, key=correspondent_key)
        correspondent_name = correspondent.get("name") or CANONICAL_FALLBACK_CORRESPONDENT_NAME
        correspondent_org = next(
            (
                recipient
                for recipient in recipients_by_destination.get(destination["iata_code"], [])
                if recipient.get("_default_contact_plain_key") == correspondent_key
            ),
            None,
        )

        if correspondent_org is None:
            placeholder_org_name = _correspondent_placeholder_org_name(destination)
            placeholder_key = _normalize_key(placeholder_org_name)
            org_contact_key = _ensure_org_contact(
                contacts=contacts,
                organization_name=placeholder_org_name,
                source_priority=source_priority,
                review_items=review_items,
                seen_conflicts=seen_conflicts,
            )
            default_contact_key, default_contact_plain_key = _ensure_person_contact(
                contacts=contacts,
                source_priority=source_priority,
                review_items=review_items,
                seen_conflicts=seen_conflicts,
                organization_contact_key=org_contact_key,
                fallback_name=correspondent_name,
            )
            correspondent_org = _get_or_create_entity(recipients, key=placeholder_key)
            correspondent_org.setdefault("name", placeholder_org_name)
            correspondent_org.setdefault("status", "")
            correspondent_org.setdefault("contact_key", org_contact_key)
            correspondent_org.setdefault("destination_iata", destination["iata_code"])
            correspondent_org.setdefault("default_contact_key", default_contact_key)
            correspondent_org.setdefault("_default_contact_plain_key", default_contact_plain_key)
            recipients_by_destination.setdefault(destination["iata_code"], []).append(
                correspondent_org
            )

        correspondent_org["is_correspondent"] = True
        destination["correspondent_org_contact_key"] = correspondent_org["contact_key"]

        if correspondent_org.get("_default_contact_plain_key") == correspondent_key:
            correspondent_contact_key = correspondent_org["default_contact_key"]
        else:
            correspondent_contact_key, _correspondent_plain_key = _ensure_person_contact(
                contacts=contacts,
                source_priority=source_priority,
                review_items=review_items,
                seen_conflicts=seen_conflicts,
                organization_contact_key=correspondent_org["contact_key"],
                fallback_name=correspondent_name,
            )

        correspondent["contact_key"] = correspondent_contact_key
        correspondent["organization_contact_key"] = correspondent_org["contact_key"]
        destination["correspondent_contact_key"] = correspondent_contact_key


def render_review_report(review_items: list[dict]) -> str:
    if not review_items:
        return "# Contact rebuild review\n\nNo review items.\n"
    lines = ["# Contact rebuild review", ""]
    for item in review_items:
        lines.append(
            "- {entity_type} `{display_name}`: {reason} ({existing!r} -> {incoming!r})".format(
                **item
            )
        )
    lines.append("")
    return "\n".join(lines)


def build_be_contact_dataset(path: str | Path) -> BeContactDataset:
    contacts: dict[str, dict] = {}
    donors: dict[str, dict] = {}
    transporters: dict[str, dict] = {}
    volunteers: dict[str, dict] = {}
    shippers: dict[str, dict] = {}
    recipients: dict[str, dict] = {}
    correspondents: dict[str, dict] = {}
    destinations: dict[str, dict] = {}
    shipment_links: set[tuple[str, str, str]] = set()
    review_items: list[dict] = []
    seen_conflicts: set[tuple[str, str, str]] = set()
    workbook = load_workbook(Path(path), read_only=True, data_only=True)

    try:
        source_sheets = _selected_sheet_names(workbook)
        for source_priority, sheet_name in enumerate(source_sheets):
            sheet = workbook[sheet_name]
            for row, header_map in _iter_sheet_rows(sheet):
                donor_name = _normalize_display_name(_get_cell(row, header_map, "BE_DONATEUR"))
                transporter_name = _normalize_display_name(
                    _get_cell(row, header_map, "BE_TRANSPORTEUR")
                )
                volunteer_name = _normalize_display_name(
                    _get_cell(row, header_map, "BE_MISE_A_BORD_RESPONSABLE")
                )
                shipper_name = _normalize_display_name(
                    _get_cell(row, header_map, "ASSOCIATION_NOM")
                )
                shipper_country = _normalize_display_name(
                    _get_cell(row, header_map, "ASSOCIATION_PAYS")
                )
                recipient_name = _normalize_display_name(
                    _get_cell(row, header_map, "DESTINATAIRE_STRUCTURE")
                )
                recipient_status = _normalize_display_name(
                    _get_cell(row, header_map, "DESTINATAIRE_STATUT")
                )
                destination_city = _normalize_display_name(
                    _get_cell(row, header_map, "BE_DESTINATION")
                )
                destination_iata = _normalize_iata(_get_cell(row, header_map, "BE_CODE_IATA"))
                correspondent_name = _normalize_display_name(
                    " ".join(
                        part
                        for part in (
                            _get_cell(row, header_map, "CORRESPONDANT_PRENOM"),
                            _get_cell(row, header_map, "CORRESPONDANT_NOM"),
                        )
                        if part
                    )
                )
                correspondent_country = _normalize_display_name(
                    _get_cell(row, header_map, "CORRESPONDANT_PAYS")
                )

                shipper_key = _normalize_key(shipper_name)
                recipient_key = _normalize_key(recipient_name)
                correspondent_key = _normalize_key(correspondent_name)
                destination_key = destination_iata or _normalize_key(
                    f"{destination_city}|{correspondent_country}"
                )

                shipper_contact_key = ""
                if shipper_key:
                    shipper_contact_key = _ensure_org_contact(
                        contacts=contacts,
                        organization_name=shipper_name,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                    )
                    shipper = _get_or_create_entity(shippers, key=shipper_key)
                    _merge_scalar(
                        record=shipper,
                        field_name="name",
                        value=shipper_name,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="shipper",
                        key=shipper_key,
                    )
                    _merge_scalar(
                        record=shipper,
                        field_name="country",
                        value=shipper_country,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="shipper",
                        key=shipper_key,
                    )
                    _merge_scalar(
                        record=shipper,
                        field_name="contact_key",
                        value=shipper_contact_key,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="shipper",
                        key=shipper_key,
                    )
                    shipper_title = _get_cell(row, header_map, "ASSOCIATION_PRESIDENT_TITRE")
                    shipper_first_name = _get_cell(row, header_map, "ASSOCIATION_PRESIDENT_PRENOM")
                    shipper_last_name = _get_cell(row, header_map, "ASSOCIATION_PRESIDENT_NOM")
                    has_shipper_contact_data = any(
                        (shipper_title, shipper_first_name, shipper_last_name)
                    )
                    if has_shipper_contact_data or not shipper.get("default_contact_key"):
                        shipper_default_contact_key, shipper_default_plain_key = (
                            _ensure_person_contact(
                                contacts=contacts,
                                source_priority=source_priority,
                                review_items=review_items,
                                seen_conflicts=seen_conflicts,
                                organization_contact_key=shipper_contact_key,
                                title=shipper_title,
                                first_name=shipper_first_name,
                                last_name=shipper_last_name,
                                fallback_name=f"Referent {shipper_name}",
                            )
                        )
                        _merge_scalar(
                            record=shipper,
                            field_name="default_contact_key",
                            value=shipper_default_contact_key,
                            source_priority=source_priority,
                            review_items=review_items,
                            seen_conflicts=seen_conflicts,
                            entity_type="shipper",
                            key=shipper_key,
                        )
                        shipper["_default_contact_plain_key"] = shipper_default_plain_key

                if donor_name:
                    _register_generic_capability_contact(
                        contacts=contacts,
                        registry=donors,
                        name=donor_name,
                        capability=ContactCapabilityType.DONOR,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        default_contact_type=ContactType.ORGANIZATION,
                    )

                if transporter_name:
                    if (
                        _normalize_key(transporter_name) == EXPEDITEUR_TRANSPORTER_KEY
                        and shipper_key
                    ):
                        transporter = _get_or_create_entity(transporters, key=shipper_key)
                        transporter.setdefault("name", shipper_name)
                        transporter.setdefault("contact_key", shipper_contact_key)
                        transporter.setdefault("contact_type", ContactType.ORGANIZATION)
                    else:
                        _register_generic_capability_contact(
                            contacts=contacts,
                            registry=transporters,
                            name=transporter_name,
                            capability=ContactCapabilityType.TRANSPORTER,
                            source_priority=source_priority,
                            review_items=review_items,
                            seen_conflicts=seen_conflicts,
                            default_contact_type=ContactType.ORGANIZATION,
                        )

                if volunteer_name:
                    _register_generic_capability_contact(
                        contacts=contacts,
                        registry=volunteers,
                        name=volunteer_name,
                        capability=ContactCapabilityType.VOLUNTEER,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        default_contact_type=ContactType.PERSON,
                    )

                if recipient_key:
                    recipient_contact_key = _ensure_org_contact(
                        contacts=contacts,
                        organization_name=recipient_name,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                    )
                    recipient = _get_or_create_entity(recipients, key=recipient_key)
                    _merge_scalar(
                        record=recipient,
                        field_name="name",
                        value=recipient_name,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="recipient",
                        key=recipient_key,
                    )
                    _merge_scalar(
                        record=recipient,
                        field_name="status",
                        value=recipient_status,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="recipient",
                        key=recipient_key,
                    )
                    _merge_scalar(
                        record=recipient,
                        field_name="contact_key",
                        value=recipient_contact_key,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="recipient",
                        key=recipient_key,
                    )
                    if destination_iata:
                        _merge_scalar(
                            record=recipient,
                            field_name="destination_iata",
                            value=destination_iata,
                            source_priority=source_priority,
                            review_items=review_items,
                            seen_conflicts=seen_conflicts,
                            entity_type="recipient",
                            key=recipient_key,
                        )
                    recipient_title = _get_cell(
                        row,
                        header_map,
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_TITRE",
                    )
                    recipient_first_name = _get_cell(
                        row,
                        header_map,
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_PRENOM",
                    )
                    recipient_last_name = _get_cell(
                        row,
                        header_map,
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_NOM",
                    )
                    has_recipient_contact_data = any(
                        (recipient_title, recipient_first_name, recipient_last_name)
                    )
                    if has_recipient_contact_data or not recipient.get("default_contact_key"):
                        default_contact_key, default_plain_key = _ensure_person_contact(
                            contacts=contacts,
                            source_priority=source_priority,
                            review_items=review_items,
                            seen_conflicts=seen_conflicts,
                            organization_contact_key=recipient_contact_key,
                            title=recipient_title,
                            first_name=recipient_first_name,
                            last_name=recipient_last_name,
                            fallback_name=f"Referent {recipient_name}",
                        )
                        _merge_scalar(
                            record=recipient,
                            field_name="default_contact_key",
                            value=default_contact_key,
                            source_priority=source_priority,
                            review_items=review_items,
                            seen_conflicts=seen_conflicts,
                            entity_type="recipient",
                            key=recipient_key,
                        )
                        recipient["_default_contact_plain_key"] = default_plain_key

                if correspondent_key:
                    correspondent = _get_or_create_entity(correspondents, key=correspondent_key)
                    _merge_scalar(
                        record=correspondent,
                        field_name="name",
                        value=correspondent_name,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="correspondent",
                        key=correspondent_key,
                    )
                    _merge_scalar(
                        record=correspondent,
                        field_name="country",
                        value=correspondent_country,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="correspondent",
                        key=correspondent_key,
                    )

                if destination_key and destination_city:
                    destination = destinations.setdefault(
                        destination_key,
                        {
                            "key": destination_key,
                            "_field_priorities": {},
                        },
                    )
                    _merge_scalar(
                        record=destination,
                        field_name="city",
                        value=destination_city,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="destination",
                        key=destination_key,
                    )
                    _merge_scalar(
                        record=destination,
                        field_name="country",
                        value=correspondent_country,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="destination",
                        key=destination_key,
                    )
                    _merge_scalar(
                        record=destination,
                        field_name="iata_code",
                        value=destination_iata,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="destination",
                        key=destination_key,
                    )
                    _merge_scalar(
                        record=destination,
                        field_name="correspondent_key",
                        value=correspondent_key,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="destination",
                        key=destination_key,
                    )

                if shipper_key and recipient_key and destination_iata:
                    shipment_links.add((shipper_key, recipient_key, destination_iata))
    finally:
        workbook.close()

    _apply_destination_correspondent_overrides(
        destinations=destinations,
        correspondents=correspondents,
    )
    _ensure_destination_required_fields(
        destinations=destinations,
        correspondents=correspondents,
        review_items=review_items,
        seen_conflicts=seen_conflicts,
    )
    _finalize_destination_correspondents(
        contacts=contacts,
        recipients=recipients,
        correspondents=correspondents,
        destinations=destinations,
        source_priority=len(source_sheets),
        review_items=review_items,
        seen_conflicts=seen_conflicts,
    )
    correspondents = _prune_unreferenced_correspondents(
        destinations=destinations,
        correspondents=correspondents,
    )

    contacts_list = sorted(
        (_clean_record(item) for item in contacts.values()),
        key=lambda item: (item["contact_type"], item["name"], item["key"]),
    )
    recipients_list = sorted(
        (_clean_record(item) for item in recipients.values()),
        key=lambda item: item["name"],
    )
    shipment_links_list = []
    recipients_by_key = {recipient["key"]: recipient for recipient in recipients_list}
    for shipper_key, recipient_key, destination_iata in sorted(shipment_links):
        recipient = recipients_by_key[recipient_key]
        shipment_links_list.append(
            {
                "shipper_key": shipper_key,
                "recipient_key": recipient_key,
                "destination_iata": destination_iata,
                "default_recipient_contact_key": recipient["default_contact_key"],
                "authorized_recipient_contact_keys": [recipient["default_contact_key"]],
            }
        )

    return BeContactDataset(
        contacts=contacts_list,
        donors=sorted(
            (_clean_record(item) for item in donors.values()), key=lambda item: item["name"]
        ),
        transporters=sorted(
            (_clean_record(item) for item in transporters.values()),
            key=lambda item: item["name"],
        ),
        volunteers=sorted(
            (_clean_record(item) for item in volunteers.values()),
            key=lambda item: item["name"],
        ),
        shippers=sorted(
            (_clean_record(item) for item in shippers.values()),
            key=lambda item: item["name"],
        ),
        recipients=recipients_list,
        correspondents=sorted(
            (_clean_record(item) for item in correspondents.values()),
            key=lambda item: item["name"],
        ),
        destinations=sorted(
            (_clean_record(item) for item in destinations.values()),
            key=lambda item: item["iata_code"],
        ),
        shipment_links=shipment_links_list,
        review_items=review_items,
        source_sheets=source_sheets,
    )
