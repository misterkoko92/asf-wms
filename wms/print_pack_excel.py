from copy import copy
from datetime import date, datetime
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import coordinate_from_string


class PrintPackMappingError(ValueError):
    """Raised when an Excel mapping cannot be applied safely."""


def _mapping_get(mapping, key, default=None):
    if isinstance(mapping, dict):
        return mapping.get(key, default)
    return getattr(mapping, key, default)


def _resolve_source_value(payload, source_key):
    current = payload
    for segment in (source_key or "").split("."):
        segment = segment.strip()
        if not segment:
            continue
        if isinstance(current, dict):
            if segment not in current:
                return None
            current = current[segment]
            continue
        if hasattr(current, segment):
            current = getattr(current, segment)
            continue
        return None
    return current


def _split_repeating_source_key(source_key):
    key = (source_key or "").strip()
    if "[]" not in key:
        return None, None
    before, after = key.split("[]", 1)
    return before.rstrip("."), after.lstrip(".")


def _iter_repeating_values(payload, source_key):
    list_key, item_key = _split_repeating_source_key(source_key)
    if not list_key:
        return []
    rows = _resolve_source_value(payload, list_key)
    if not isinstance(rows, (list, tuple)):
        return []
    if not item_key:
        return list(rows)
    return [_resolve_source_value(row, item_key) for row in rows]


def _resolve_target_cell(worksheet, cell_ref):
    cell = worksheet[cell_ref]
    if isinstance(cell, MergedCell):
        for merged_range in worksheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return worksheet[merged_range.start_cell.coordinate]
    return cell


def _is_missing(value):
    return value is None or (isinstance(value, str) and value.strip() == "")


def _apply_transform(value, transform):
    transform_key = (transform or "").strip().lower()
    if _is_missing(value):
        return ""
    if transform_key == "upper":
        return str(value).upper()
    if transform_key == "date_fr":
        if isinstance(value, (date, datetime)):
            return value.strftime("%d/%m/%Y")
        return str(value)
    return value


def _write_cell_value(target_cell, value):
    target_cell.value = value
    alignment = copy(target_cell.alignment)
    alignment.wrap_text = True
    target_cell.alignment = alignment


def fill_workbook_cells(workbook, mappings, payload):
    for mapping in mappings:
        worksheet_name = _mapping_get(mapping, "worksheet_name", "")
        cell_ref = _mapping_get(mapping, "cell_ref", "")
        source_key = _mapping_get(mapping, "source_key", "")
        transform = _mapping_get(mapping, "transform", "")
        required = bool(_mapping_get(mapping, "required", False))

        if worksheet_name not in workbook.sheetnames:
            raise PrintPackMappingError(f"Unknown worksheet: {worksheet_name}")
        worksheet = workbook[worksheet_name]

        if "[]" in source_key:
            column, base_row = coordinate_from_string(cell_ref)
            values = _iter_repeating_values(payload, source_key)
            if required and not values:
                raise PrintPackMappingError(
                    f"Missing required mapping value for {worksheet_name}!{cell_ref} ({source_key})"
                )
            for idx, raw_value in enumerate(values):
                if required and _is_missing(raw_value):
                    raise PrintPackMappingError(
                        f"Missing required mapping value for {worksheet_name}!{column}{base_row + idx} ({source_key})"
                    )
                target_cell = _resolve_target_cell(
                    worksheet,
                    f"{column}{base_row + idx}",
                )
                _write_cell_value(
                    target_cell,
                    _apply_transform(raw_value, transform),
                )
            continue

        value = _resolve_source_value(payload, source_key)
        if required and _is_missing(value):
            raise PrintPackMappingError(
                f"Missing required mapping value for {worksheet_name}!{cell_ref} ({source_key})"
            )

        target_cell = _resolve_target_cell(worksheet, cell_ref)
        _write_cell_value(
            target_cell,
            _apply_transform(value, transform),
        )
    return workbook


def _display_length(value):
    if value is None:
        return 0
    text = str(value)
    if not text:
        return 0
    lines = text.splitlines() or [text]
    return max(len(line) for line in lines)


def autosize_workbook_columns(workbook, *, min_width=8, max_width=80, padding=2):
    safe_min = max(1, int(min_width or 1))
    safe_max = max(safe_min, int(max_width or safe_min))
    safe_padding = max(0, int(padding or 0))

    for worksheet in workbook.worksheets:
        widths = {}
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                length = _display_length(cell.value)
                if length <= 0:
                    continue
                col_letter = cell.column_letter
                widths[col_letter] = max(widths.get(col_letter, 0), length)

        for col_letter, content_len in widths.items():
            computed_width = content_len + safe_padding
            bounded_width = min(safe_max, max(safe_min, computed_width))
            worksheet.column_dimensions[col_letter].width = bounded_width
    return workbook
