from datetime import date
from types import SimpleNamespace

from django.test import SimpleTestCase
from openpyxl import Workbook

from wms.print_pack_excel import (
    PrintPackMappingError,
    autosize_workbook_columns,
    fill_workbook_cells,
)


class PrintPackExcelTests(SimpleTestCase):
    def _mapping(self, **overrides):
        payload = {
            "worksheet_name": "Main",
            "cell_ref": "D5",
            "source_key": "shipment.recipient.full_name",
            "transform": "",
            "required": False,
        }
        payload.update(overrides)
        return SimpleNamespace(**payload)

    def test_fill_workbook_cells_writes_transformed_values(self):
        workbook = Workbook()
        workbook.active.title = "Main"
        mappings = [self._mapping(transform="upper")]
        payload = {"shipment": {"recipient": {"full_name": "Doe John"}}}

        fill_workbook_cells(workbook, mappings, payload)

        self.assertEqual(workbook["Main"]["D5"].value, "DOE JOHN")

    def test_fill_workbook_cells_formats_dates(self):
        workbook = Workbook()
        workbook.active.title = "Main"
        mappings = [self._mapping(source_key="shipment.requested_delivery_date", transform="date_fr")]
        payload = {"shipment": {"requested_delivery_date": date(2026, 3, 1)}}

        fill_workbook_cells(workbook, mappings, payload)

        self.assertEqual(workbook["Main"]["D5"].value, "01/03/2026")

    def test_fill_workbook_cells_raises_when_required_value_is_missing(self):
        workbook = Workbook()
        workbook.active.title = "Main"
        mappings = [self._mapping(required=True)]
        payload = {"shipment": {}}

        with self.assertRaises(PrintPackMappingError):
            fill_workbook_cells(workbook, mappings, payload)

    def test_fill_workbook_cells_writes_repeating_rows(self):
        workbook = Workbook()
        workbook.active.title = "Main"
        mappings = [
            self._mapping(
                cell_ref="A14",
                source_key="carton.items[].product_name",
            ),
            self._mapping(
                cell_ref="C14",
                source_key="carton.items[].quantity",
            ),
        ]
        payload = {
            "carton": {
                "items": [
                    {"product_name": "Gants", "quantity": 4},
                    {"product_name": "Masques", "quantity": 10},
                ]
            }
        }

        fill_workbook_cells(workbook, mappings, payload)

        self.assertEqual(workbook["Main"]["A14"].value, "Gants")
        self.assertEqual(workbook["Main"]["A15"].value, "Masques")
        self.assertEqual(workbook["Main"]["C14"].value, 4)
        self.assertEqual(workbook["Main"]["C15"].value, 10)

    def test_fill_workbook_cells_raises_when_repeating_value_is_required_and_missing(self):
        workbook = Workbook()
        workbook.active.title = "Main"
        mappings = [
            self._mapping(
                cell_ref="A14",
                source_key="carton.items[].product_name",
                required=True,
            )
        ]
        payload = {"carton": {"items": []}}

        with self.assertRaises(PrintPackMappingError):
            fill_workbook_cells(workbook, mappings, payload)

    def test_fill_workbook_cells_writes_to_merged_range_anchor(self):
        workbook = Workbook()
        workbook.active.title = "Main"
        sheet = workbook["Main"]
        sheet.merge_cells("B7:H7")
        mappings = [self._mapping(cell_ref="C7", source_key="shipment.destination_city")]
        payload = {"shipment": {"destination_city": "ABIDJAN"}}

        fill_workbook_cells(workbook, mappings, payload)

        self.assertEqual(sheet["B7"].value, "ABIDJAN")

    def test_fill_workbook_cells_supports_dict_mapping_and_object_payload(self):
        workbook = Workbook()
        workbook.active.title = "Main"
        mappings = [
            {
                "worksheet_name": "Main",
                "cell_ref": "A1",
                "source_key": "shipment.recipient.full_name",
                "transform": "",
                "required": True,
            }
        ]
        payload = {
            "shipment": SimpleNamespace(
                recipient=SimpleNamespace(full_name="John Doe"),
            )
        }

        fill_workbook_cells(workbook, mappings, payload)

        self.assertEqual(workbook["Main"]["A1"].value, "John Doe")

    def test_fill_workbook_cells_raises_when_worksheet_is_unknown(self):
        workbook = Workbook()
        workbook.active.title = "Main"
        mappings = [self._mapping(worksheet_name="Missing")]

        with self.assertRaises(PrintPackMappingError):
            fill_workbook_cells(workbook, mappings, payload={})

    def test_fill_workbook_cells_raises_on_required_empty_repeating_value(self):
        workbook = Workbook()
        workbook.active.title = "Main"
        mappings = [
            self._mapping(
                cell_ref="A14",
                source_key="carton.items[].product_name",
                required=True,
            )
        ]
        payload = {"carton": {"items": [{"product_name": "   "}]}}

        with self.assertRaises(PrintPackMappingError):
            fill_workbook_cells(workbook, mappings, payload)

    def test_autosize_workbook_columns_adjusts_widths_from_cell_values(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Main"
        sheet["A1"] = "SHORT"
        sheet["B1"] = "A very long value for autosize"

        autosize_workbook_columns(workbook)

        self.assertGreaterEqual(sheet.column_dimensions["A"].width, 8)
        self.assertGreater(sheet.column_dimensions["B"].width, sheet.column_dimensions["A"].width)

    def test_autosize_workbook_columns_respects_max_width_bound(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Main"
        sheet["C1"] = "x" * 500

        autosize_workbook_columns(workbook, min_width=8, max_width=40, padding=2)

        self.assertEqual(sheet.column_dimensions["C"].width, 40)
