from openpyxl import Workbook
from django.test import SimpleTestCase

from wms.print_pack_mapping_catalog import ALLOWED_SOURCE_KEYS, is_allowed_source_key
from wms.print_pack_workbook import (
    build_column_choices,
    normalize_cell_ref,
    worksheet_row_choices,
)


class PrintPackWorkbookTests(SimpleTestCase):
    def test_build_column_choices_includes_excel_bounds(self):
        columns = build_column_choices()
        self.assertEqual(columns[0], "A")
        self.assertIn("Z", columns)
        self.assertIn("AA", columns)
        self.assertIn("AZ", columns)
        self.assertIn("BA", columns)
        self.assertEqual(columns[-1], "XFD")

    def test_worksheet_row_choices_bounds(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Feuil1"
        worksheet["A1"] = "header"
        worksheet["A12"] = "footer"
        rows = worksheet_row_choices(worksheet)
        self.assertEqual(rows[0], 1)
        self.assertEqual(rows[-1], 12)
        workbook.close()

    def test_normalize_cell_ref_returns_anchor_for_merged_cell(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Feuil1"
        worksheet.merge_cells("B5:B7")

        normalized, merged_range = normalize_cell_ref(worksheet, "B6")
        self.assertEqual(normalized, "B5")
        self.assertEqual(merged_range, "B5:B7")
        workbook.close()

    def test_normalize_cell_ref_returns_input_when_not_merged(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Feuil1"
        normalized, merged_range = normalize_cell_ref(worksheet, "C12")
        self.assertEqual(normalized, "C12")
        self.assertEqual(merged_range, "")
        workbook.close()

    def test_allowed_source_keys_is_closed_catalog(self):
        self.assertIn("shipment.reference", ALLOWED_SOURCE_KEYS)
        self.assertIn("carton.code", ALLOWED_SOURCE_KEYS)
        self.assertTrue(is_allowed_source_key("shipment.reference"))
        self.assertFalse(is_allowed_source_key("shipment.unknown_field"))
