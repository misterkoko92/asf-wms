import json
from io import BytesIO
from unittest import mock

from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.http import HttpResponse
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook

from contacts.models import Contact, ContactType
from wms.models import (
    Carton,
    Destination,
    PrintCellMapping,
    PrintPack,
    PrintPackDocument,
    Product,
    Shipment,
)


class PrintTemplateViewsTests(TestCase):
    def setUp(self):
        self.superuser = get_user_model().objects.create_superuser(
            username="print-admin",
            email="print-admin@example.com",
            password="pass1234",
        )
        self.client.force_login(self.superuser)
        self.pack = PrintPack.objects.create(code="TZ", name="Template Zone")
        self.pack_document = PrintPackDocument.objects.create(
            pack=self.pack,
            doc_type="shipment_note",
            variant="shipment",
            sequence=1,
            enabled=True,
        )
        self.pack_document.xlsx_template_file.save(
            "TZ__shipment_note__shipment.xlsx",
            ContentFile(self._build_workbook_bytes()),
            save=True,
        )
        PrintCellMapping.objects.create(
            pack_document=self.pack_document,
            worksheet_name="Feuil1",
            cell_ref="A24",
            source_key="shipment.shipper.title",
            transform="",
            required=True,
            sequence=1,
        )

    def _list_url(self):
        return reverse("scan:scan_print_templates")

    def _edit_url(self, doc_type="shipment_note"):
        return reverse("scan:scan_print_template_edit", args=[doc_type])

    def _preview_url(self):
        return reverse("scan:scan_print_template_preview")

    def _render_stub(self, _request, template_name, context):
        response = HttpResponse(template_name)
        response.context_data = context
        return response

    def _build_workbook_bytes(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Feuil1"
        worksheet["A1"] = "Header"
        worksheet["A60"] = "Footer"
        worksheet["D1"] = "D"
        worksheet["D60"] = "D"
        worksheet.merge_cells("B5:B7")
        other = workbook.create_sheet("Annexe")
        other["A1"] = "Annexe"
        output = BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def test_scan_print_templates_list_displays_print_pack_documents(self):
        response = self.client.get(self._list_url())
        self.assertEqual(response.status_code, 200)
        template_entry = next(
            item
            for item in response.context["templates"]
            if item["route_key"] == str(self.pack_document.id)
        )
        self.assertEqual(template_entry["pack_code"], "TZ")
        self.assertEqual(template_entry["variant"], "shipment")
        self.assertEqual(template_entry["mapping_count"], 1)
        self.assertTrue(template_entry["has_template_file"])
        self.assertEqual(template_entry["route_key"], str(self.pack_document.id))

    def test_scan_print_template_edit_404_for_unknown_doc_type(self):
        response = self.client.get(self._edit_url("unknown-template"))
        self.assertEqual(response.status_code, 404)

    def test_scan_print_template_edit_get_exposes_editor_context(self):
        response = self.client.get(self._edit_url(str(self.pack_document.id)))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["pack_document"].id, self.pack_document.id)
        self.assertEqual(response.context["worksheet_names"], ["Feuil1", "Annexe"])
        self.assertEqual(len(response.context["mapping_rows"]), 1)
        self.assertIn("shipment.reference", response.context["source_keys"])

    def test_scan_print_template_edit_save_updates_mappings_and_versions(self):
        response = self.client.post(
            self._edit_url(str(self.pack_document.id)),
            {
                "action": "save",
                "mapping_worksheet": ["Feuil1", "Feuil1"],
                "mapping_column": ["A", "B"],
                "mapping_row": ["11", "6"],
                "mapping_source_key": ["carton.code", "shipment.shipper.structure_name"],
                "mapping_transform": ["", "upper"],
                "mapping_required": ["1", "0"],
                "mapping_sequence": ["1", "2"],
                "change_note": "batch update",
            },
        )
        self.assertEqual(response.status_code, 302)
        self.pack_document.refresh_from_db()
        mappings = list(self.pack_document.cell_mappings.order_by("sequence", "id"))
        self.assertEqual(len(mappings), 2)
        self.assertEqual(mappings[0].cell_ref, "A11")
        self.assertEqual(mappings[0].source_key, "carton.code")
        self.assertEqual(mappings[1].cell_ref, "B5")
        self.assertEqual(mappings[1].source_key, "shipment.shipper.structure_name")
        version = self.pack_document.versions.get()
        self.assertEqual(version.version, 1)
        self.assertEqual(version.change_note, "batch update")
        self.assertEqual(version.created_by, self.superuser)

    def test_scan_print_template_edit_save_rejects_invalid_source_key(self):
        response = self.client.post(
            self._edit_url(str(self.pack_document.id)),
            {
                "action": "save",
                "mapping_worksheet": ["Feuil1"],
                "mapping_column": ["A"],
                "mapping_row": ["11"],
                "mapping_source_key": ["shipment.invalid"],
                "mapping_transform": [""],
                "mapping_required": ["1"],
                "mapping_sequence": ["1"],
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.pack_document.cell_mappings.count(), 1)
        self.assertEqual(self.pack_document.versions.count(), 0)

    def test_scan_print_template_edit_restore_requires_version_id(self):
        response = self.client.post(
            self._edit_url(str(self.pack_document.id)),
            {"action": "restore"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.pack_document.versions.count(), 0)

    def test_scan_print_template_edit_restore_applies_selected_version(self):
        first_save = self.client.post(
            self._edit_url(str(self.pack_document.id)),
            {
                "action": "save",
                "mapping_worksheet": ["Feuil1"],
                "mapping_column": ["A"],
                "mapping_row": ["11"],
                "mapping_source_key": ["carton.code"],
                "mapping_transform": [""],
                "mapping_required": ["1"],
                "mapping_sequence": ["1"],
            },
        )
        self.assertEqual(first_save.status_code, 302)
        v1 = self.pack_document.versions.get(version=1)
        second_save = self.client.post(
            self._edit_url(str(self.pack_document.id)),
            {
                "action": "save",
                "mapping_worksheet": ["Feuil1"],
                "mapping_column": ["A"],
                "mapping_row": ["11"],
                "mapping_source_key": ["carton.position"],
                "mapping_transform": [""],
                "mapping_required": ["1"],
                "mapping_sequence": ["1"],
            },
        )
        self.assertEqual(second_save.status_code, 302)
        response = self.client.post(
            self._edit_url(str(self.pack_document.id)),
            {"action": "restore", "version_id": str(v1.id)},
        )
        self.assertEqual(response.status_code, 302)
        self.pack_document.refresh_from_db()
        restored_mapping = self.pack_document.cell_mappings.get(
            worksheet_name="Feuil1",
            cell_ref="A11",
        )
        self.assertEqual(restored_mapping.source_key, "carton.code")
        versions = list(self.pack_document.versions.order_by("version"))
        self.assertEqual([item.version for item in versions], [1, 2, 3])
        self.assertEqual(versions[-1].change_type, "restore")

    def test_scan_print_template_preview_rejects_unknown_doc_type(self):
        response = self.client.post(
            self._preview_url(),
            {"doc_type": "unknown", "layout_json": "{}"},
        )
        self.assertEqual(response.status_code, 404)

    def test_scan_print_template_preview_rejects_invalid_json(self):
        response = self.client.post(
            self._preview_url(),
            {"doc_type": "shipment_note", "layout_json": "{"},
        )
        self.assertEqual(response.status_code, 400)

    def test_scan_print_template_preview_shipment_label_without_shipment(self):
        with mock.patch(
            "wms.views_print_templates.build_sample_label_context",
            return_value={"destination_city": "Paris"},
        ):
            with mock.patch(
                "wms.views_print_templates.render_layout_from_layout",
                return_value=[{"type": "label_city"}],
            ):
                with mock.patch(
                    "wms.views_print_templates.render",
                    side_effect=self._render_stub,
                ) as render_mock:
                    response = self.client.post(
                        self._preview_url(),
                        {"doc_type": "shipment_label", "layout_json": "{}"},
                    )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode(), "print/dynamic_labels.html")
        render_kwargs = render_mock.call_args.args[2]
        self.assertEqual(len(render_kwargs["labels"]), 1)

    def test_scan_print_template_preview_shipment_label_with_shipment_cartons(self):
        shipment = Shipment.objects.create(
            shipper_name="Sender",
            recipient_name="Recipient",
            destination_address="1 Rue Test",
            destination_country="France",
        )
        Carton.objects.create(code="C-001", shipment=shipment)
        Carton.objects.create(code="C-002", shipment=shipment)

        with mock.patch(
            "wms.views_print_templates.build_label_context",
            return_value={"destination_city": "Paris"},
        ) as build_label_context_mock:
            with mock.patch(
                "wms.views_print_templates.render_layout_from_layout",
                return_value=[{"type": "label_city"}],
            ):
                with mock.patch(
                    "wms.views_print_templates.render",
                    side_effect=self._render_stub,
                ) as render_mock:
                    response = self.client.post(
                        self._preview_url(),
                        {
                            "doc_type": "shipment_label",
                            "layout_json": "{}",
                            "shipment_id": str(shipment.id),
                        },
                    )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode(), "print/dynamic_labels.html")
        self.assertEqual(build_label_context_mock.call_count, 2)
        render_kwargs = render_mock.call_args.args[2]
        self.assertEqual(len(render_kwargs["labels"]), 2)

    def test_scan_print_template_preview_shipment_label_with_empty_shipment_uses_sample(self):
        shipment = Shipment.objects.create(
            shipper_name="Sender Empty",
            recipient_name="Recipient Empty",
            destination_address="1 Rue Test",
            destination_country="France",
        )

        with mock.patch(
            "wms.views_print_templates.build_sample_label_context",
            return_value={"destination_city": "Sample"},
        ) as sample_context_mock:
            with mock.patch(
                "wms.views_print_templates.render_layout_from_layout",
                return_value=[{"type": "label_city"}],
            ):
                with mock.patch(
                    "wms.views_print_templates.render",
                    side_effect=self._render_stub,
                ) as render_mock:
                    response = self.client.post(
                        self._preview_url(),
                        {
                            "doc_type": "shipment_label",
                            "layout_json": "{}",
                            "shipment_id": str(shipment.id),
                        },
                    )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode(), "print/dynamic_labels.html")
        sample_context_mock.assert_called_once()
        self.assertEqual(len(render_mock.call_args.args[2]["labels"]), 1)

    def test_scan_print_template_preview_product_label_with_product(self):
        product = Product.objects.create(name="Produit Preview")
        with mock.patch(
            "wms.views_print_templates.build_product_label_context",
            return_value={"name": product.name},
        ) as product_context_mock:
            with mock.patch(
                "wms.views_print_templates.build_label_pages",
                return_value=([{"rows": []}], {"page_margin": "0"}),
            ) as build_pages_mock:
                with mock.patch(
                    "wms.views_print_templates.render",
                    side_effect=self._render_stub,
                ):
                    response = self.client.post(
                        self._preview_url(),
                        {
                            "doc_type": "product_label",
                            "product_id": str(product.id),
                            "layout_json": "{}",
                        },
                    )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode(), "print/product_labels.html")
        product_context_mock.assert_called_once()
        self.assertEqual(product_context_mock.call_args.args[0].id, product.id)
        build_pages_kwargs = build_pages_mock.call_args.kwargs
        self.assertEqual(build_pages_kwargs["block_type"], "product_label")
        self.assertEqual(build_pages_kwargs["labels_per_page"], 4)

    def test_scan_print_template_preview_product_label_without_product_uses_generic_context(self):
        with mock.patch(
            "wms.views_print_templates.build_preview_context",
            return_value={"name": "fallback"},
        ) as preview_context_mock:
            with mock.patch(
                "wms.views_print_templates.build_label_pages",
                return_value=([{"rows": []}], {"page_margin": "0"}),
            ):
                with mock.patch(
                    "wms.views_print_templates.render",
                    side_effect=self._render_stub,
                ):
                    response = self.client.post(
                        self._preview_url(),
                        {
                            "doc_type": "product_label",
                            "layout_json": "{}",
                            "product_id": "not-a-number",
                        },
                    )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode(), "print/product_labels.html")
        preview_context_mock.assert_called_once_with("product_label")

    def test_scan_print_template_preview_product_qr_with_product(self):
        product = Product.objects.create(name="Produit QR")
        product.qr_code_image = ""
        product.save(update_fields=["qr_code_image"])

        with mock.patch(
            "wms.models.Product.generate_qr_code",
            autospec=True,
        ) as generate_qr_mock:
            with mock.patch(
                "wms.views_print_templates.build_preview_context",
                return_value={"sku": product.sku},
            ) as preview_context_mock:
                with mock.patch(
                    "wms.views_print_templates.extract_block_style",
                    return_value={"page_rows": "2", "page_columns": "3"},
                ):
                    with mock.patch(
                        "wms.views_print_templates.build_label_pages",
                        return_value=([{"rows": []}], {"page_margin": "0"}),
                    ) as build_pages_mock:
                        with mock.patch(
                            "wms.views_print_templates.render",
                            side_effect=self._render_stub,
                        ):
                            response = self.client.post(
                                self._preview_url(),
                                {
                                    "doc_type": "product_qr",
                                    "product_id": str(product.id),
                                    "layout_json": "{}",
                                },
                            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode(), "print/product_qr_labels.html")
        generate_qr_mock.assert_called_once()
        preview_context_mock.assert_called_once()
        build_pages_kwargs = build_pages_mock.call_args.kwargs
        self.assertEqual(build_pages_kwargs["block_type"], "product_qr_label")
        self.assertEqual(build_pages_kwargs["labels_per_page"], 6)

    def test_scan_print_template_preview_product_qr_without_product_and_invalid_grid_uses_defaults(
        self,
    ):
        with mock.patch(
            "wms.views_print_templates.build_preview_context",
            return_value={"sku": "fallback"},
        ) as preview_context_mock:
            with mock.patch(
                "wms.views_print_templates.extract_block_style",
                return_value={"page_rows": "x", "page_columns": "y"},
            ):
                with mock.patch(
                    "wms.views_print_templates.build_label_pages",
                    return_value=([{"rows": []}], {"page_margin": "0"}),
                ) as build_pages_mock:
                    with mock.patch(
                        "wms.views_print_templates.render",
                        side_effect=self._render_stub,
                    ):
                        response = self.client.post(
                            self._preview_url(),
                            {
                                "doc_type": "product_qr",
                                "layout_json": "{}",
                                "product_id": "",
                            },
                        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode(), "print/product_qr_labels.html")
        preview_context_mock.assert_called_once_with("product_qr")
        build_pages_kwargs = build_pages_mock.call_args.kwargs
        self.assertEqual(build_pages_kwargs["labels_per_page"], 15)

    def test_scan_print_template_preview_standard_document(self):
        with mock.patch(
            "wms.views_print_templates.build_preview_context",
            return_value={"shipment_ref": "SHP-1"},
        ) as preview_context_mock:
            with mock.patch(
                "wms.views_print_templates.render_layout_from_layout",
                return_value=[{"type": "text"}],
            ) as render_layout_mock:
                with mock.patch(
                    "wms.views_print_templates.render",
                    side_effect=self._render_stub,
                ) as render_mock:
                    response = self.client.post(
                        self._preview_url(),
                        {"doc_type": "shipment_note", "layout_json": "{}"},
                    )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.decode(), "print/dynamic_document.html")
        preview_context_mock.assert_called_once()
        render_layout_mock.assert_called_once()
        self.assertEqual(render_mock.call_args.args[2]["blocks"], [{"type": "text"}])
