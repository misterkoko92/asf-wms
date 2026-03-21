from __future__ import annotations

import tempfile
from io import StringIO
from pathlib import Path

from django.core.management import call_command
from django.test import TestCase
from openpyxl import Workbook

from contacts.capabilities import ContactCapabilityType
from contacts.models import Contact, ContactType
from wms.contact_rebuild import (
    BeContactDataset,
    apply_be_contact_dataset,
    build_be_contact_dataset,
    render_review_report,
)
from wms.models import (
    Destination,
    ShipmentAuthorizedRecipientContact,
    ShipmentRecipientContact,
    ShipmentRecipientOrganization,
    ShipmentShipper,
    ShipmentShipperRecipientLink,
    ShipmentValidationStatus,
)


class BeWorkbookMixin:
    HEADER = [
        "BE_DONATEUR",
        "BE_TRANSPORTEUR",
        "BE_MISE_A_BORD_RESPONSABLE",
        "ASSOCIATION_NOM",
        "ASSOCIATION_PAYS",
        "ASSOCIATION_PRESIDENT_TITRE",
        "ASSOCIATION_PRESIDENT_PRENOM",
        "ASSOCIATION_PRESIDENT_NOM",
        "DESTINATAIRE_STRUCTURE",
        "DESTINATAIRE_STATUT",
        "DESTINATAIRE_STRUCTURE_REPRESENTANT_TITRE",
        "DESTINATAIRE_STRUCTURE_REPRESENTANT_PRENOM",
        "DESTINATAIRE_STRUCTURE_REPRESENTANT_NOM",
        "CORRESPONDANT_PRENOM",
        "CORRESPONDANT_NOM",
        "CORRESPONDANT_PAYS",
        "BE_DESTINATION",
        "BE_CODE_IATA",
    ]

    def _build_be_workbook(self, rows: list[dict]) -> Path:
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "be_source.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(self.HEADER)
        for row in rows:
            sheet.append([row.get(header, "") for header in self.HEADER])
        workbook.save(path)
        return path

    def _build_be_multisheet_workbook(self, sheets: dict[str, list[dict]]) -> Path:
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        path = Path(temp_dir.name) / "be_source_multisheet.xlsx"

        workbook = Workbook()
        for index, (sheet_name, rows) in enumerate(sheets.items()):
            sheet = workbook.active if index == 0 else workbook.create_sheet()
            sheet.title = sheet_name
            sheet.append(self.HEADER)
            for row in rows:
                sheet.append([row.get(header, "") for header in self.HEADER])
        workbook.save(path)
        return path


class RebuildContactsFromBeXlsxNormalizationTests(BeWorkbookMixin, TestCase):
    def test_build_be_contact_dataset_groups_entities_and_surfaces_no_ambiguity(self):
        workbook_path = self._build_be_workbook(
            [
                {
                    "BE_DONATEUR": "Homeperf Clamart",
                    "BE_TRANSPORTEUR": "LEGENDRE",
                    "BE_MISE_A_BORD_RESPONSABLE": "STOURM",
                    "ASSOCIATION_NOM": "AVIATION SANS FRONTIERES",
                    "ASSOCIATION_PAYS": "France",
                    "ASSOCIATION_PRESIDENT_PRENOM": "Anne",
                    "ASSOCIATION_PRESIDENT_NOM": "Dupond",
                    "DESTINATAIRE_STRUCTURE": "Recipient A",
                    "DESTINATAIRE_STATUT": "actif",
                    "DESTINATAIRE_STRUCTURE_REPRESENTANT_PRENOM": "Leontine",
                    "DESTINATAIRE_STRUCTURE_REPRESENTANT_NOM": "Rahazania",
                    "CORRESPONDANT_PRENOM": "Leontine",
                    "CORRESPONDANT_NOM": "Rahazania",
                    "CORRESPONDANT_PAYS": "Madagascar",
                    "BE_DESTINATION": "ANTANANARIVO",
                    "BE_CODE_IATA": "TNR",
                },
                {
                    "ASSOCIATION_NOM": "Aviation Sans Frontieres",
                    "ASSOCIATION_PAYS": "France",
                    "DESTINATAIRE_STRUCTURE": "Recipient A",
                    "DESTINATAIRE_STATUT": "actif",
                    "CORRESPONDANT_PRENOM": "Leontine",
                    "CORRESPONDANT_NOM": "Rahazania",
                    "CORRESPONDANT_PAYS": "Madagascar",
                    "BE_DESTINATION": "ANTANANARIVO",
                    "BE_CODE_IATA": "TNR",
                },
            ]
        )

        dataset = build_be_contact_dataset(workbook_path)

        self.assertGreaterEqual(len(dataset.contacts), 5)
        self.assertEqual(len(dataset.donors), 1)
        self.assertEqual(len(dataset.transporters), 1)
        self.assertEqual(len(dataset.volunteers), 1)
        self.assertEqual(len(dataset.shippers), 1)
        self.assertEqual(len(dataset.recipients), 1)
        self.assertEqual(len(dataset.correspondents), 1)
        self.assertEqual(len(dataset.destinations), 1)
        self.assertTrue(dataset.recipients[0]["is_correspondent"])
        self.assertEqual(
            dataset.shipment_links,
            [
                {
                    "recipient_key": "recipient a",
                    "shipper_key": "aviation sans frontieres",
                    "destination_iata": "TNR",
                    "default_recipient_contact_key": dataset.recipients[0]["default_contact_key"],
                    "authorized_recipient_contact_keys": [
                        dataset.recipients[0]["default_contact_key"]
                    ],
                }
            ],
        )
        self.assertEqual(dataset.review_items, [])

    def test_build_be_contact_dataset_reports_conflicting_shipper_data(self):
        workbook_path = self._build_be_workbook(
            [
                {
                    "ASSOCIATION_NOM": "Association Test",
                    "ASSOCIATION_PAYS": "France",
                    "DESTINATAIRE_STRUCTURE": "Recipient A",
                    "DESTINATAIRE_STATUT": "actif",
                    "CORRESPONDANT_PRENOM": "Marie",
                    "CORRESPONDANT_NOM": "Durand",
                    "CORRESPONDANT_PAYS": "Senegal",
                    "BE_DESTINATION": "DAKAR",
                    "BE_CODE_IATA": "DKR",
                },
                {
                    "ASSOCIATION_NOM": "association test",
                    "ASSOCIATION_PAYS": "Belgique",
                    "DESTINATAIRE_STRUCTURE": "Recipient B",
                    "DESTINATAIRE_STATUT": "actif",
                    "CORRESPONDANT_PRENOM": "Marie",
                    "CORRESPONDANT_NOM": "Durand",
                    "CORRESPONDANT_PAYS": "Senegal",
                    "BE_DESTINATION": "DAKAR",
                    "BE_CODE_IATA": "DKR",
                },
            ]
        )

        dataset = build_be_contact_dataset(workbook_path)

        self.assertEqual(len(dataset.review_items), 1)
        self.assertIn("conflicting", dataset.review_items[0]["reason"])
        report = render_review_report(dataset.review_items)
        self.assertIn("Association Test", report)
        self.assertIn("conflicting shipper country", report)

    def test_build_be_contact_dataset_prefers_latest_sheet_values(self):
        workbook_path = self._build_be_multisheet_workbook(
            {
                "2024": [
                    {
                        "ASSOCIATION_NOM": "AVIATION SANS FRONTIERES",
                        "ASSOCIATION_PAYS": "France",
                        "DESTINATAIRE_STRUCTURE": "Recipient BGF",
                        "DESTINATAIRE_STATUT": "historique",
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_PRENOM": "Garba",
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_NOM": "TAHA ABOUBAKAR",
                        "CORRESPONDANT_PRENOM": "Garba",
                        "CORRESPONDANT_NOM": "TAHA ABOUBAKAR",
                        "CORRESPONDANT_PAYS": "Centrafrique",
                        "BE_DESTINATION": "BANGUI",
                        "BE_CODE_IATA": "BGF",
                    }
                ],
                "2025": [
                    {
                        "ASSOCIATION_NOM": "AVIATION SANS FRONTIERES",
                        "ASSOCIATION_PAYS": "France",
                        "DESTINATAIRE_STRUCTURE": "Recipient BGF",
                        "DESTINATAIRE_STATUT": "actif",
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_PRENOM": "Christian",
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_NOM": "LIMBIO",
                        "CORRESPONDANT_PRENOM": "Christian",
                        "CORRESPONDANT_NOM": "LIMBIO",
                        "CORRESPONDANT_PAYS": "Centrafrique",
                        "BE_DESTINATION": "BANGUI",
                        "BE_CODE_IATA": "BGF",
                    }
                ],
                "2026": [
                    {
                        "ASSOCIATION_NOM": "AVIATION SANS FRONTIERES",
                        "ASSOCIATION_PAYS": "France",
                        "DESTINATAIRE_STRUCTURE": "Recipient BGF",
                        "DESTINATAIRE_STATUT": "prioritaire",
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_PRENOM": "Christian",
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_NOM": "LIMBIO",
                        "CORRESPONDANT_PRENOM": "Christian",
                        "CORRESPONDANT_NOM": "LIMBIO",
                        "CORRESPONDANT_PAYS": "Centrafrique",
                        "BE_DESTINATION": "BANGUI",
                        "BE_CODE_IATA": "BGF",
                    }
                ],
            }
        )

        dataset = build_be_contact_dataset(workbook_path)

        self.assertEqual(dataset.source_sheets, ["2024", "2025", "2026"])
        self.assertEqual(
            dataset.recipients,
            [
                {
                    "key": "recipient bgf",
                    "name": "Recipient BGF",
                    "status": "prioritaire",
                    "contact_key": dataset.recipients[0]["contact_key"],
                    "destination_iata": "BGF",
                    "default_contact_key": dataset.recipients[0]["default_contact_key"],
                    "is_correspondent": True,
                }
            ],
        )
        self.assertEqual(
            dataset.correspondents,
            [
                {
                    "key": "christian limbio",
                    "name": "Christian LIMBIO",
                    "country": "Centrafrique",
                    "contact_key": dataset.correspondents[0]["contact_key"],
                    "organization_contact_key": dataset.correspondents[0][
                        "organization_contact_key"
                    ],
                }
            ],
        )
        self.assertEqual(
            dataset.destinations,
            [
                {
                    "key": "BGF",
                    "city": "BANGUI",
                    "country": "Centrafrique",
                    "iata_code": "BGF",
                    "correspondent_key": "christian limbio",
                    "correspondent_contact_key": dataset.destinations[0][
                        "correspondent_contact_key"
                    ],
                    "correspondent_org_contact_key": dataset.destinations[0][
                        "correspondent_org_contact_key"
                    ],
                }
            ],
        )
        self.assertEqual(dataset.review_items, [])

    def test_build_be_contact_dataset_applies_explicit_correspondent_overrides(self):
        workbook_path = self._build_be_multisheet_workbook(
            {
                "2024": [
                    {
                        "ASSOCIATION_NOM": "AVIATION SANS FRONTIERES",
                        "ASSOCIATION_PAYS": "France",
                        "DESTINATAIRE_STRUCTURE": "Recipient BEY",
                        "DESTINATAIRE_STATUT": "actif",
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_PRENOM": "Tony",
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_NOM": "MDAWAR",
                        "CORRESPONDANT_PRENOM": "Ancien",
                        "CORRESPONDANT_NOM": "Correspondant",
                        "CORRESPONDANT_PAYS": "Liban",
                        "BE_DESTINATION": "BEYROUTH",
                        "BE_CODE_IATA": "BEY",
                    }
                ],
                "2025": [
                    {
                        "ASSOCIATION_NOM": "AVIATION SANS FRONTIERES",
                        "ASSOCIATION_PAYS": "France",
                        "DESTINATAIRE_STRUCTURE": "Recipient BGF",
                        "DESTINATAIRE_STATUT": "actif",
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_PRENOM": "Christian",
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_NOM": "LIMBIO",
                        "CORRESPONDANT_PRENOM": "Garba",
                        "CORRESPONDANT_NOM": "TAHA ABOUBAKAR",
                        "CORRESPONDANT_PAYS": "Centrafrique",
                        "BE_DESTINATION": "BANGUI",
                        "BE_CODE_IATA": "BGF",
                    }
                ],
                "2026": [
                    {
                        "ASSOCIATION_NOM": "AVIATION SANS FRONTIERES",
                        "ASSOCIATION_PAYS": "France",
                        "DESTINATAIRE_STRUCTURE": "Recipient NDJ",
                        "DESTINATAIRE_STATUT": "actif",
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_PRENOM": "Geovanie",
                        "DESTINATAIRE_STRUCTURE_REPRESENTANT_NOM": "Kamtar NDANGMBAYE",
                        "CORRESPONDANT_PRENOM": "Rocher",
                        "CORRESPONDANT_NOM": "DJEGUERBE MBAIOUNDADE",
                        "CORRESPONDANT_PAYS": "Tchad",
                        "BE_DESTINATION": "N'DJAMENA",
                        "BE_CODE_IATA": "NDJ",
                    }
                ],
            }
        )

        dataset = build_be_contact_dataset(workbook_path)
        destinations_by_iata = {
            destination["iata_code"]: destination for destination in dataset.destinations
        }
        correspondents_by_key = {
            correspondent["key"]: correspondent for correspondent in dataset.correspondents
        }

        self.assertEqual(destinations_by_iata["BEY"]["correspondent_key"], "tony mdawar")
        self.assertEqual(
            correspondents_by_key["tony mdawar"]["name"],
            "Tony MDAWAR",
        )
        self.assertEqual(
            destinations_by_iata["BGF"]["correspondent_key"],
            "christian limbio",
        )
        self.assertEqual(
            correspondents_by_key["christian limbio"]["name"],
            "Christian LIMBIO",
        )
        self.assertEqual(
            destinations_by_iata["NDJ"]["correspondent_key"],
            "geovanie kamtar ndangmbaye",
        )
        self.assertEqual(
            correspondents_by_key["geovanie kamtar ndangmbaye"]["name"],
            "Geovanie Kamtar NDANGMBAYE",
        )
        self.assertEqual(dataset.review_items, [])

    def test_build_be_contact_dataset_assigns_fallback_for_destination_without_correspondent(self):
        workbook_path = self._build_be_workbook(
            [
                {
                    "ASSOCIATION_NOM": "AVIATION SANS FRONTIERES",
                    "ASSOCIATION_PAYS": "France",
                    "BE_DESTINATION": "HANOI",
                    "BE_CODE_IATA": "HAN",
                }
            ]
        )

        dataset = build_be_contact_dataset(workbook_path)

        self.assertEqual(
            dataset.destinations,
            [
                {
                    "key": "HAN",
                    "city": "HANOI",
                    "country": "",
                    "iata_code": "HAN",
                    "correspondent_key": "correspondant non renseigne",
                    "correspondent_contact_key": dataset.destinations[0][
                        "correspondent_contact_key"
                    ],
                    "correspondent_org_contact_key": dataset.destinations[0][
                        "correspondent_org_contact_key"
                    ],
                }
            ],
        )
        self.assertEqual(
            dataset.correspondents,
            [
                {
                    "key": "correspondant non renseigne",
                    "name": "Correspondant non renseigne",
                    "contact_key": dataset.correspondents[0]["contact_key"],
                    "organization_contact_key": dataset.correspondents[0][
                        "organization_contact_key"
                    ],
                }
            ],
        )
        self.assertEqual(
            {item["reason"] for item in dataset.review_items},
            {
                "missing destination correspondent_key",
                "missing destination country",
            },
        )


class RebuildContactsFromBeXlsxPersistenceTests(TestCase):
    def test_apply_be_contact_dataset_creates_contacts_capabilities_and_shipment_registry(self):
        dataset = BeContactDataset(
            contacts=[
                {
                    "key": "org:donor a",
                    "name": "Donor A",
                    "contact_type": ContactType.ORGANIZATION,
                },
                {
                    "key": "org:aviation sans frontieres",
                    "name": "AVIATION SANS FRONTIERES",
                    "contact_type": ContactType.ORGANIZATION,
                },
                {
                    "key": "person:org:aviation sans frontieres:anne dupond",
                    "name": "Anne Dupond",
                    "contact_type": ContactType.PERSON,
                    "first_name": "Anne",
                    "last_name": "Dupond",
                    "organization_key": "org:aviation sans frontieres",
                },
                {
                    "key": "org:recipient a",
                    "name": "Recipient A",
                    "contact_type": ContactType.ORGANIZATION,
                },
                {
                    "key": "person:org:recipient a:leontine rahazania",
                    "name": "Leontine Rahazania",
                    "contact_type": ContactType.PERSON,
                    "first_name": "Leontine",
                    "last_name": "Rahazania",
                    "organization_key": "org:recipient a",
                },
                {
                    "key": "org:legendre",
                    "name": "LEGENDRE",
                    "contact_type": ContactType.ORGANIZATION,
                },
                {
                    "key": "person:stourm",
                    "name": "STOURM",
                    "contact_type": ContactType.PERSON,
                    "last_name": "STOURM",
                },
            ],
            donors=[
                {
                    "key": "donor a",
                    "name": "Donor A",
                    "contact_key": "org:donor a",
                    "contact_type": ContactType.ORGANIZATION,
                }
            ],
            transporters=[
                {
                    "key": "legendre",
                    "name": "LEGENDRE",
                    "contact_key": "org:legendre",
                    "contact_type": ContactType.ORGANIZATION,
                }
            ],
            volunteers=[
                {
                    "key": "stourm",
                    "name": "STOURM",
                    "contact_key": "person:stourm",
                    "contact_type": ContactType.PERSON,
                }
            ],
            shippers=[
                {
                    "key": "aviation sans frontieres",
                    "name": "AVIATION SANS FRONTIERES",
                    "country": "France",
                    "contact_key": "org:aviation sans frontieres",
                    "default_contact_key": "person:org:aviation sans frontieres:anne dupond",
                }
            ],
            recipients=[
                {
                    "key": "recipient a",
                    "name": "Recipient A",
                    "status": "actif",
                    "contact_key": "org:recipient a",
                    "destination_iata": "TNR",
                    "default_contact_key": "person:org:recipient a:leontine rahazania",
                    "is_correspondent": True,
                }
            ],
            correspondents=[
                {
                    "key": "leontine rahazania",
                    "name": "Leontine Rahazania",
                    "country": "Madagascar",
                    "contact_key": "person:org:recipient a:leontine rahazania",
                    "organization_contact_key": "org:recipient a",
                }
            ],
            destinations=[
                {
                    "key": "TNR",
                    "city": "ANTANANARIVO",
                    "country": "Madagascar",
                    "iata_code": "TNR",
                    "correspondent_key": "leontine rahazania",
                    "correspondent_contact_key": "person:org:recipient a:leontine rahazania",
                    "correspondent_org_contact_key": "org:recipient a",
                }
            ],
            shipment_links=[
                {
                    "recipient_key": "recipient a",
                    "shipper_key": "aviation sans frontieres",
                    "destination_iata": "TNR",
                    "default_recipient_contact_key": "person:org:recipient a:leontine rahazania",
                    "authorized_recipient_contact_keys": [
                        "person:org:recipient a:leontine rahazania"
                    ],
                }
            ],
            review_items=[],
        )

        apply_be_contact_dataset(dataset)

        donor = Contact.objects.get(name="Donor A")
        shipper = Contact.objects.get(name="AVIATION SANS FRONTIERES")
        recipient = Contact.objects.get(name="Recipient A")
        correspondent = Contact.objects.get(name="Leontine Rahazania")
        transporter = Contact.objects.get(name="LEGENDRE")
        volunteer = Contact.objects.get(name="STOURM")
        destination = Destination.objects.get(iata_code="TNR")
        shipment_shipper = ShipmentShipper.objects.get(organization=shipper)
        shipment_recipient = ShipmentRecipientOrganization.objects.get(organization=recipient)
        shipment_recipient_contact = ShipmentRecipientContact.objects.get(
            recipient_organization=shipment_recipient,
            contact=correspondent,
        )
        shipment_link = ShipmentShipperRecipientLink.objects.get(
            shipper=shipment_shipper,
            recipient_organization=shipment_recipient,
        )

        self.assertTrue(
            donor.capabilities.filter(
                capability=ContactCapabilityType.DONOR,
                is_active=True,
            ).exists()
        )
        self.assertTrue(
            transporter.capabilities.filter(
                capability=ContactCapabilityType.TRANSPORTER,
                is_active=True,
            ).exists()
        )
        self.assertTrue(
            volunteer.capabilities.filter(
                capability=ContactCapabilityType.VOLUNTEER,
                is_active=True,
            ).exists()
        )
        self.assertEqual(shipper.contact_type, ContactType.ORGANIZATION)
        self.assertEqual(recipient.contact_type, ContactType.ORGANIZATION)
        self.assertEqual(correspondent.contact_type, ContactType.PERSON)
        self.assertEqual(destination.correspondent_contact_id, correspondent.id)
        self.assertEqual(shipment_shipper.default_contact.name, "Anne Dupond")
        self.assertTrue(shipment_recipient.is_correspondent)
        self.assertEqual(shipment_recipient.destination_id, destination.id)
        self.assertTrue(
            ShipmentAuthorizedRecipientContact.objects.filter(
                link=shipment_link,
                recipient_contact=shipment_recipient_contact,
                is_default=True,
                is_active=True,
            ).exists()
        )
        self.assertEqual(ShipmentShipper.objects.count(), 1)
        self.assertEqual(ShipmentShipperRecipientLink.objects.count(), 1)

    def test_apply_be_contact_dataset_keeps_correspondent_contact_separate_when_not_default(self):
        dataset = BeContactDataset(
            contacts=[
                {
                    "key": "org:aviation sans frontieres",
                    "name": "AVIATION SANS FRONTIERES",
                    "contact_type": ContactType.ORGANIZATION,
                },
                {
                    "key": "person:org:aviation sans frontieres:anne dupond",
                    "name": "Anne Dupond",
                    "contact_type": ContactType.PERSON,
                    "first_name": "Anne",
                    "last_name": "Dupond",
                    "organization_key": "org:aviation sans frontieres",
                },
                {
                    "key": "org:recipient a",
                    "name": "Recipient A",
                    "contact_type": ContactType.ORGANIZATION,
                },
                {
                    "key": "person:org:recipient a:recipient rep",
                    "name": "Recipient Rep",
                    "contact_type": ContactType.PERSON,
                    "first_name": "Recipient",
                    "last_name": "Rep",
                    "organization_key": "org:recipient a",
                },
                {
                    "key": "person:org:recipient a:leontine rahazania",
                    "name": "Leontine Rahazania",
                    "contact_type": ContactType.PERSON,
                    "first_name": "Leontine",
                    "last_name": "Rahazania",
                    "organization_key": "org:recipient a",
                },
            ],
            shippers=[
                {
                    "key": "aviation sans frontieres",
                    "name": "AVIATION SANS FRONTIERES",
                    "country": "France",
                    "contact_key": "org:aviation sans frontieres",
                    "default_contact_key": "person:org:aviation sans frontieres:anne dupond",
                },
            ],
            recipients=[
                {
                    "key": "recipient a",
                    "name": "Recipient A",
                    "status": "actif",
                    "contact_key": "org:recipient a",
                    "destination_iata": "TNR",
                    "default_contact_key": "person:org:recipient a:recipient rep",
                    "is_correspondent": True,
                }
            ],
            correspondents=[
                {
                    "key": "leontine rahazania",
                    "name": "Leontine Rahazania",
                    "country": "Madagascar",
                    "contact_key": "person:org:recipient a:leontine rahazania",
                    "organization_contact_key": "org:recipient a",
                }
            ],
            destinations=[
                {
                    "key": "TNR",
                    "city": "ANTANANARIVO",
                    "country": "Madagascar",
                    "iata_code": "TNR",
                    "correspondent_key": "leontine rahazania",
                    "correspondent_contact_key": "person:org:recipient a:leontine rahazania",
                    "correspondent_org_contact_key": "org:recipient a",
                },
            ],
            shipment_links=[
                {
                    "recipient_key": "recipient a",
                    "shipper_key": "aviation sans frontieres",
                    "destination_iata": "TNR",
                    "default_recipient_contact_key": "person:org:recipient a:recipient rep",
                    "authorized_recipient_contact_keys": ["person:org:recipient a:recipient rep"],
                },
            ],
            review_items=[],
        )

        apply_be_contact_dataset(dataset)

        shipment_recipient = ShipmentRecipientOrganization.objects.get(
            organization__name="Recipient A"
        )
        self.assertEqual(
            ShipmentRecipientContact.objects.filter(
                recipient_organization=shipment_recipient
            ).count(),
            2,
        )
        self.assertTrue(
            ShipmentRecipientContact.objects.filter(
                recipient_organization=shipment_recipient,
                contact__name="Leontine Rahazania",
            ).exists()
        )


class RebuildContactsFromBeXlsxCommandTests(BeWorkbookMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.workbook_path = self._build_be_workbook(
            [
                {
                    "BE_DONATEUR": "Donor A",
                    "ASSOCIATION_NOM": "AVIATION SANS FRONTIERES",
                    "ASSOCIATION_PAYS": "France",
                    "DESTINATAIRE_STRUCTURE": "Recipient A",
                    "DESTINATAIRE_STATUT": "actif",
                    "CORRESPONDANT_PRENOM": "Leontine",
                    "CORRESPONDANT_NOM": "Rahazania",
                    "CORRESPONDANT_PAYS": "Madagascar",
                    "BE_DESTINATION": "ANTANANARIVO",
                    "BE_CODE_IATA": "TNR",
                }
            ]
        )
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        self.report_path = Path(temp_dir.name) / "review.md"

    def test_rebuild_contacts_from_be_xlsx_dry_run_reports_actions_without_database_writes(self):
        stdout = StringIO()

        call_command(
            "rebuild_contacts_from_be_xlsx",
            "--source",
            str(self.workbook_path),
            "--dry-run",
            "--report-path",
            str(self.report_path),
            stdout=stdout,
        )

        output = stdout.getvalue()
        self.assertIn("Rebuild contacts from BE workbook [DRY RUN]", output)
        self.assertIn("Contacts:", output)
        self.assertIn("Shippers: 1", output)
        self.assertIn("Shipment links: 1", output)
        self.assertFalse(Contact.objects.exists())
        self.assertTrue(self.report_path.exists())

    def test_rebuild_contacts_from_be_xlsx_apply_writes_review_report_and_runtime_data(self):
        stdout = StringIO()

        call_command(
            "rebuild_contacts_from_be_xlsx",
            "--source",
            str(self.workbook_path),
            "--apply",
            "--report-path",
            str(self.report_path),
            stdout=stdout,
        )

        output = stdout.getvalue()
        self.assertIn("Rebuild contacts from BE workbook [APPLY]", output)
        self.assertTrue(Contact.objects.exists())
        self.assertTrue(self.report_path.exists())
        self.assertTrue(Destination.objects.filter(iata_code="TNR").exists())
        self.assertTrue(
            ShipmentShipper.objects.filter(organization__name="AVIATION SANS FRONTIERES").exists()
        )
        self.assertTrue(
            ShipmentShipperRecipientLink.objects.filter(
                shipper__organization__name="AVIATION SANS FRONTIERES",
                recipient_organization__organization__name="Recipient A",
            ).exists()
        )
        self.assertEqual(ShipmentShipper.objects.count(), 1)
        self.assertEqual(ShipmentShipperRecipientLink.objects.count(), 1)

    def test_rebuild_contacts_from_be_xlsx_apply_handles_destination_without_correspondent(self):
        workbook_path = self._build_be_workbook(
            [
                {
                    "ASSOCIATION_NOM": "AVIATION SANS FRONTIERES",
                    "ASSOCIATION_PAYS": "France",
                    "BE_DESTINATION": "HANOI",
                    "BE_CODE_IATA": "HAN",
                }
            ]
        )
        stdout = StringIO()

        call_command(
            "rebuild_contacts_from_be_xlsx",
            "--source",
            str(workbook_path),
            "--apply",
            "--report-path",
            str(self.report_path),
            stdout=stdout,
        )

        self.assertIn("Rebuild contacts from BE workbook [APPLY]", stdout.getvalue())
        self.assertTrue(Destination.objects.filter(iata_code="HAN", city="HANOI").exists())
        self.assertTrue(Contact.objects.filter(name="Correspondant non renseigne").exists())
        self.assertIn("missing destination correspondent_key", self.report_path.read_text())

    def test_rebuild_contacts_from_be_xlsx_reports_selected_source_sheets(self):
        workbook_path = self._build_be_multisheet_workbook(
            {
                "2024": [
                    {
                        "ASSOCIATION_NOM": "AVIATION SANS FRONTIERES",
                        "ASSOCIATION_PAYS": "France",
                        "DESTINATAIRE_STRUCTURE": "Recipient A",
                        "DESTINATAIRE_STATUT": "actif",
                        "CORRESPONDANT_PRENOM": "Leontine",
                        "CORRESPONDANT_NOM": "Rahazania",
                        "CORRESPONDANT_PAYS": "Madagascar",
                        "BE_DESTINATION": "ANTANANARIVO",
                        "BE_CODE_IATA": "TNR",
                    }
                ],
                "2025": [],
                "2026": [],
            }
        )
        stdout = StringIO()

        call_command(
            "rebuild_contacts_from_be_xlsx",
            "--source",
            str(workbook_path),
            "--dry-run",
            "--report-path",
            str(self.report_path),
            stdout=stdout,
        )

        output = stdout.getvalue()
        self.assertIn("Source sheets: 2024, 2025, 2026", output)


class RebuildContactsCanonicalCommandTests(BeWorkbookMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.workbook_path = self._build_be_workbook(
            [
                {
                    "BE_DONATEUR": "Donor A",
                    "ASSOCIATION_NOM": "AVIATION SANS FRONTIERES",
                    "ASSOCIATION_PAYS": "France",
                    "DESTINATAIRE_STRUCTURE": "Recipient A",
                    "DESTINATAIRE_STATUT": "actif",
                    "CORRESPONDANT_PRENOM": "Leontine",
                    "CORRESPONDANT_NOM": "Rahazania",
                    "CORRESPONDANT_PAYS": "Madagascar",
                    "BE_DESTINATION": "ANTANANARIVO",
                    "BE_CODE_IATA": "TNR",
                }
            ]
        )
        temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(temp_dir.cleanup)
        self.report_path = Path(temp_dir.name) / "canonical-review.md"

        self.existing_shipper = Contact.objects.create(
            name="Legacy Shipper",
            contact_type=ContactType.ORGANIZATION,
        )
        self.existing_recipient = Contact.objects.create(
            name="Legacy Recipient",
            contact_type=ContactType.ORGANIZATION,
        )
        self.existing_correspondent = Contact.objects.create(
            name="Legacy Correspondent",
            contact_type=ContactType.PERSON,
            first_name="Legacy",
            last_name="Correspondent",
            organization=self.existing_recipient,
        )
        self.existing_destination = Destination.objects.create(
            city="Legacy City",
            iata_code="OLD",
            country="France",
            correspondent_contact=self.existing_correspondent,
            is_active=True,
        )
        self.existing_shipper_contact = Contact.objects.create(
            name="Legacy Shipper Referent",
            contact_type=ContactType.PERSON,
            first_name="Legacy",
            last_name="Shipper",
            organization=self.existing_shipper,
            is_active=True,
        )
        self.existing_shipper_model = ShipmentShipper.objects.create(
            organization=self.existing_shipper,
            default_contact=self.existing_shipper_contact,
            validation_status=ShipmentValidationStatus.VALIDATED,
            is_active=True,
        )
        self.existing_recipient_model = ShipmentRecipientOrganization.objects.create(
            organization=self.existing_recipient,
            destination=self.existing_destination,
            validation_status=ShipmentValidationStatus.VALIDATED,
            is_correspondent=True,
            is_active=True,
        )
        self.existing_recipient_contact = ShipmentRecipientContact.objects.create(
            recipient_organization=self.existing_recipient_model,
            contact=self.existing_correspondent,
            is_active=True,
        )
        self.existing_link = ShipmentShipperRecipientLink.objects.create(
            shipper=self.existing_shipper_model,
            recipient_organization=self.existing_recipient_model,
            is_active=True,
        )
        ShipmentAuthorizedRecipientContact.objects.create(
            link=self.existing_link,
            recipient_contact=self.existing_recipient_contact,
            is_default=True,
            is_active=True,
        )

    def test_canonical_command_dry_run_leaves_existing_runtime_data_untouched(self):
        stdout = StringIO()

        call_command(
            "rebuild_contacts_canonical",
            "--source",
            str(self.workbook_path),
            "--dry-run",
            "--report-path",
            str(self.report_path),
            stdout=stdout,
        )

        output = stdout.getvalue()
        self.assertIn("Canonical contact rebuild [DRY RUN]", output)
        self.assertIn("Review report:", output)
        self.assertTrue(Contact.objects.filter(pk=self.existing_shipper.pk).exists())
        self.assertTrue(Destination.objects.filter(pk=self.existing_destination.pk).exists())
        self.assertTrue(ShipmentShipper.objects.filter(pk=self.existing_shipper_model.pk).exists())
        self.assertFalse(Contact.objects.filter(name="AVIATION SANS FRONTIERES").exists())
        self.assertTrue(self.report_path.exists())

    def test_canonical_command_apply_resets_then_imports_workbook_data(self):
        stdout = StringIO()

        call_command(
            "rebuild_contacts_canonical",
            "--source",
            str(self.workbook_path),
            "--apply",
            "--report-path",
            str(self.report_path),
            stdout=stdout,
        )

        output = stdout.getvalue()
        self.assertIn("Canonical contact rebuild [APPLY]", output)
        self.assertFalse(Contact.objects.filter(pk=self.existing_shipper.pk).exists())
        self.assertFalse(Destination.objects.filter(pk=self.existing_destination.pk).exists())
        self.assertTrue(Contact.objects.filter(name="AVIATION SANS FRONTIERES").exists())
        self.assertTrue(Contact.objects.filter(name="Recipient A").exists())
        self.assertTrue(Destination.objects.filter(iata_code="TNR").exists())
        self.assertTrue(
            ShipmentShipperRecipientLink.objects.filter(
                shipper__organization__name="AVIATION SANS FRONTIERES",
                recipient_organization__organization__name="Recipient A",
                recipient_organization__destination__iata_code="TNR",
                is_active=True,
            ).exists()
        )
        self.assertTrue(
            ShipmentAuthorizedRecipientContact.objects.filter(
                link__shipper__organization__name="AVIATION SANS FRONTIERES",
                recipient_contact__contact__organization__name="Recipient A",
                is_default=True,
                is_active=True,
            ).exists()
        )
        self.assertEqual(ShipmentShipper.objects.count(), 1)
        self.assertEqual(ShipmentShipperRecipientLink.objects.count(), 1)
        self.assertTrue(self.report_path.exists())
