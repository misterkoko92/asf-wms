from __future__ import annotations

import tempfile
from datetime import timedelta
from io import StringIO
from pathlib import Path
from unittest.mock import Mock, patch

from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import SimpleTestCase, TestCase
from django.utils import timezone

from openpyxl import Workbook

from contacts.models import Contact, ContactType
from wms.management.commands import import_org_roles_template as command_module
from wms.models import (
    ContactSubscription,
    Destination,
    DestinationCorrespondentDefault,
    DestinationCorrespondentOverride,
    MigrationReviewItem,
    MigrationReviewItemStatus,
    NotificationChannel,
    OrganizationContact,
    OrganizationRole,
    OrganizationRoleAssignment,
    OrganizationRoleContact,
    RecipientBinding,
    RoleEventType,
    ShipperScope,
)


class ImportOrgRolesTemplateHelpersTests(SimpleTestCase):
    def test_normalize_and_bool_parsing(self):
        self.assertEqual(command_module._normalize(None), "")
        self.assertEqual(command_module._normalize("  ACME  "), "ACME")
        self.assertIsNone(command_module._bool_or_none(""))
        self.assertTrue(command_module._bool_or_none("true"))
        self.assertFalse(command_module._bool_or_none("FALSE"))
        with self.assertRaisesMessage(ValueError, "Valeur booleenne invalide"):
            command_module._bool_or_none("maybe")

    def test_parse_datetime_or_none_accepts_iso_and_date(self):
        aware_dt = command_module._parse_datetime_or_none("2026-03-06T10:11:12+00:00")
        self.assertIsNotNone(aware_dt)
        self.assertTrue(timezone.is_aware(aware_dt))

        aware_date = command_module._parse_datetime_or_none("2026-03-06")
        self.assertIsNotNone(aware_date)
        self.assertTrue(timezone.is_aware(aware_date))
        self.assertEqual(aware_date.hour, 0)
        self.assertEqual(aware_date.minute, 0)
        self.assertEqual(aware_date.second, 0)

        self.assertIsNone(command_module._parse_datetime_or_none(None))
        self.assertIsNone(command_module._parse_datetime_or_none(""))

    def test_parse_datetime_or_none_rejects_invalid_input(self):
        with self.assertRaises(ValueError):
            command_module._parse_datetime_or_none("2026-99-99")

    def test_parse_datetime_or_none_accepts_datetime_values_and_blank_space(self):
        naive_dt = timezone.datetime(2026, 3, 6, 10, 11, 12)
        parsed_naive_dt = command_module._parse_datetime_or_none(naive_dt)
        self.assertTrue(timezone.is_aware(parsed_naive_dt))

        aware_dt = timezone.now()
        self.assertEqual(command_module._parse_datetime_or_none(aware_dt), aware_dt)
        self.assertIsNone(command_module._parse_datetime_or_none("   "))

    def test_sheet_rows_maps_headers_and_skips_blank_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Organizations"
        sheet.append(["organization_name", "role", "role_active"])
        sheet.append(["  ACME  ", "shipper", "true"])
        sheet.append([None, None, None])
        sheet.append(["Beta", "recipient", ""])

        rows = command_module._sheet_rows(workbook=workbook, sheet_name="Organizations")

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["organization_name"], "ACME")
        self.assertEqual(rows[0]["role"], "shipper")
        self.assertEqual(rows[0]["_row_number"], "2")
        self.assertEqual(rows[1]["organization_name"], "Beta")
        self.assertEqual(rows[1]["_row_number"], "4")

    def test_sheet_rows_raises_when_sheet_missing(self):
        workbook = Workbook()
        with self.assertRaisesMessage(CommandError, "Onglet absent: Organizations"):
            command_module._sheet_rows(workbook=workbook, sheet_name="Organizations")

    def test_resolve_path_handles_relative_and_absolute_values(self):
        absolute = command_module._resolve_path("/tmp/example.xlsx")
        self.assertTrue(absolute.is_absolute())
        self.assertEqual(str(absolute), "/tmp/example.xlsx")

        relative = command_module._resolve_path("docs/import/test.xlsx")
        self.assertTrue(relative.is_absolute())
        self.assertIn("/docs/import/test.xlsx", str(relative))


class ImportOrgRolesTemplateCommandTests(TestCase):
    @staticmethod
    def _build_empty_template(path: Path) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        sheet_headers = {
            "Organizations": ["organization_name", "role", "role_active"],
            "ShipperScopes": [
                "organization_name",
                "all_destinations",
                "is_active",
                "valid_from",
                "valid_to",
                "destination_iata",
            ],
            "RecipientBindings": [
                "recipient_organization_name",
                "shipper_organization_name",
                "destination_iata",
                "is_active",
                "valid_from",
                "valid_to",
            ],
            "Correspondents": [
                "correspondent_organization_name",
                "destination_iata",
                "scope_type",
                "is_active",
                "shipper_organization_name",
                "recipient_organization_name",
            ],
            "OrganizationContacts": [
                "organization_name",
                "role",
                "contact_email",
                "is_primary",
                "is_active",
            ],
            "MigrationReview": [
                "row_id",
                "resolution_action",
                "resolution_note",
                "proposed_shipper_organization_name",
                "proposed_destination_iata",
            ],
        }

        for sheet_name, headers in sheet_headers.items():
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(headers)

        workbook.save(path)

    def _create_org(self, name: str) -> Contact:
        return Contact.objects.create(
            name=name,
            contact_type=ContactType.ORGANIZATION,
            is_active=True,
        )

    def _create_destination(self, iata_code: str) -> Destination:
        correspondent = self._create_org(f"Correspondent {iata_code}")
        return Destination.objects.create(
            city=f"City {iata_code}",
            iata_code=iata_code,
            country="Country",
            correspondent_contact=correspondent,
            is_active=True,
        )

    def test_command_rejects_missing_input_file(self):
        with self.assertRaisesMessage(CommandError, "Fichier introuvable"):
            call_command(
                "import_org_roles_template",
                "--input",
                "/tmp/does-not-exist-org-roles.xlsx",
            )

    def test_command_runs_dry_run_with_empty_template(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            template_path = Path(tmp.name)
            self._build_empty_template(template_path)
            out = StringIO()

            call_command(
                "import_org_roles_template",
                "--input",
                str(template_path),
                "--dry-run",
                stdout=out,
            )

        output = out.getvalue()
        self.assertIn("Import org roles template [DRY RUN]", output)
        self.assertIn("Organizations rows: 0", output)
        self.assertIn("MigrationReview rows: 0", output)

    def test_command_runs_apply_with_empty_template(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            template_path = Path(tmp.name)
            self._build_empty_template(template_path)
            out = StringIO()

            call_command(
                "import_org_roles_template",
                "--input",
                str(template_path),
                stdout=out,
            )

        output = out.getvalue()
        self.assertIn("Import org roles template [APPLY]", output)

    def test_single_org_by_name_can_create_missing_organization(self):
        stats = command_module.ImportStats()

        organization = command_module._single_org_by_name(
            "New Org",
            row_label="Organizations row 2",
            create_missing_organizations=True,
            stats=stats,
        )

        self.assertEqual(organization.name, "New Org")
        self.assertEqual(organization.contact_type, ContactType.ORGANIZATION)
        self.assertEqual(stats.organizations_created, 1)

    def test_single_org_by_name_rejects_ambiguous_match(self):
        self._create_org("ACME")
        self._create_org("acme")

        with self.assertRaisesMessage(CommandError, "organisation ambigue"):
            command_module._single_org_by_name(
                "AcMe",
                row_label="Organizations row 2",
            )

    def test_single_org_by_name_rejects_missing_organization(self):
        with self.assertRaisesMessage(CommandError, "organisation introuvable"):
            command_module._single_org_by_name(
                "No Org",
                row_label="Organizations row 2",
            )

    def test_destination_by_iata_returns_single_match(self):
        destination = self._create_destination("DKR")

        found = command_module._destination_by_iata("dkr", row_label="Row 2")

        self.assertEqual(found.pk, destination.pk)

    def test_destination_by_iata_rejects_missing_destination(self):
        with self.assertRaisesMessage(CommandError, "escale introuvable"):
            command_module._destination_by_iata("XXX", row_label="Row 2")

    def test_destination_by_iata_rejects_ambiguous_destination(self):
        with patch.object(command_module.Destination.objects, "filter") as filter_mock:
            query = Mock()
            query.order_by.return_value = query
            query.count.return_value = 2
            filter_mock.return_value = query
            with self.assertRaisesMessage(CommandError, "escale ambigue"):
                command_module._destination_by_iata("DKR", row_label="Row 2")

    def test_apply_organizations_creates_then_updates_assignment(self):
        stats = command_module.ImportStats()
        rows = [
            {
                "_row_number": "2",
                "organization_name": "Org A",
                "role": OrganizationRole.RECIPIENT,
                "role_active": "true",
            },
        ]

        command_module._apply_organizations(
            rows,
            stats,
            create_missing_organizations=True,
        )

        organization = Contact.objects.get(name="Org A")
        assignment = OrganizationRoleAssignment.objects.get(
            organization=organization,
            role=OrganizationRole.RECIPIENT,
        )
        self.assertTrue(assignment.is_active)
        self.assertEqual(stats.organizations_rows, 1)
        self.assertEqual(stats.organization_assignments_created, 1)

        update_stats = command_module.ImportStats()
        command_module._apply_organizations(
            [
                {
                    "_row_number": "3",
                    "organization_name": "Org A",
                    "role": OrganizationRole.RECIPIENT,
                    "role_active": "false",
                }
            ],
            update_stats,
            create_missing_organizations=False,
        )
        assignment.refresh_from_db()
        self.assertFalse(assignment.is_active)
        self.assertEqual(update_stats.organization_assignments_updated, 1)

    def test_apply_organizations_rejects_invalid_rows(self):
        stats = command_module.ImportStats()
        with self.assertRaisesMessage(CommandError, "organization_name requis"):
            command_module._apply_organizations(
                [
                    {
                        "_row_number": "2",
                        "organization_name": "",
                        "role": OrganizationRole.RECIPIENT,
                    }
                ],
                stats,
                create_missing_organizations=False,
            )

        with self.assertRaisesMessage(CommandError, "role invalide"):
            command_module._apply_organizations(
                [
                    {
                        "_row_number": "3",
                        "organization_name": "Org A",
                        "role": "unknown-role",
                    }
                ],
                stats,
                create_missing_organizations=True,
            )

    def test_apply_shipper_scopes_all_destinations_creates_and_updates(self):
        shipper_org = self._create_org("Scope Global")
        stats = command_module.ImportStats()

        command_module._apply_shipper_scopes(
            [
                {
                    "_row_number": "2",
                    "organization_name": shipper_org.name,
                    "all_destinations": "true",
                    "is_active": "false",
                    "valid_from": "2026-03-01",
                    "valid_to": "2026-03-10",
                    "destination_iata": "",
                }
            ],
            stats,
            create_missing_organizations=False,
        )

        scope = ShipperScope.objects.get(role_assignment__organization=shipper_org)
        self.assertTrue(scope.all_destinations)
        self.assertFalse(scope.is_active)
        self.assertEqual(stats.shipper_scopes_created, 1)

        update_stats = command_module.ImportStats()
        command_module._apply_shipper_scopes(
            [
                {
                    "_row_number": "3",
                    "organization_name": shipper_org.name,
                    "all_destinations": "true",
                    "is_active": "true",
                    "valid_from": "2026-03-05",
                    "valid_to": "",
                    "destination_iata": "",
                }
            ],
            update_stats,
            create_missing_organizations=False,
        )
        scope.refresh_from_db()
        self.assertTrue(scope.is_active)
        self.assertEqual(update_stats.shipper_scopes_updated, 1)

    def test_apply_shipper_scopes_destination_requires_iata_for_non_global_scope(self):
        shipper_org = self._create_org("Scope Missing IATA")
        stats = command_module.ImportStats()

        with self.assertRaisesMessage(CommandError, "destination_iata requis"):
            command_module._apply_shipper_scopes(
                [
                    {
                        "_row_number": "2",
                        "organization_name": shipper_org.name,
                        "all_destinations": "false",
                        "is_active": "true",
                        "valid_from": "2026-03-01",
                        "valid_to": "",
                        "destination_iata": "",
                    }
                ],
                stats,
                create_missing_organizations=False,
            )

    def test_apply_shipper_scopes_destination_creates_and_updates(self):
        shipper_org = self._create_org("Scope Destination")
        destination = self._create_destination("CMN")
        stats = command_module.ImportStats()

        command_module._apply_shipper_scopes(
            [
                {
                    "_row_number": "2",
                    "organization_name": shipper_org.name,
                    "all_destinations": "false",
                    "is_active": "true",
                    "valid_from": "2026-03-01",
                    "valid_to": "",
                    "destination_iata": destination.iata_code,
                }
            ],
            stats,
            create_missing_organizations=False,
        )
        self.assertEqual(stats.shipper_scopes_created, 1)

        update_stats = command_module.ImportStats()
        command_module._apply_shipper_scopes(
            [
                {
                    "_row_number": "3",
                    "organization_name": shipper_org.name,
                    "all_destinations": "false",
                    "is_active": "false",
                    "valid_from": "2026-03-02",
                    "valid_to": "2026-03-20",
                    "destination_iata": destination.iata_code,
                }
            ],
            update_stats,
            create_missing_organizations=False,
        )
        self.assertEqual(update_stats.shipper_scopes_updated, 1)
        scope = ShipperScope.objects.get(
            role_assignment__organization=shipper_org,
            destination=destination,
        )
        self.assertFalse(scope.is_active)

    def test_apply_recipient_bindings_rejects_missing_fields(self):
        stats = command_module.ImportStats()
        with self.assertRaisesMessage(CommandError, "sont requis"):
            command_module._apply_recipient_bindings(
                [
                    {
                        "_row_number": "2",
                        "recipient_organization_name": "",
                        "shipper_organization_name": "",
                        "destination_iata": "",
                    }
                ],
                stats,
                create_missing_organizations=False,
            )

    def test_apply_recipient_bindings_creates_and_reuses_active_binding(self):
        recipient = self._create_org("Recipient Reuse")
        shipper = self._create_org("Shipper Reuse")
        destination = self._create_destination("LFW")
        existing_binding = RecipientBinding.objects.create(
            shipper_org=shipper,
            recipient_org=recipient,
            destination=destination,
            is_active=True,
            valid_from=timezone.now() - timedelta(days=2),
        )

        stats = command_module.ImportStats()
        command_module._apply_recipient_bindings(
            [
                {
                    "_row_number": "2",
                    "recipient_organization_name": recipient.name,
                    "shipper_organization_name": shipper.name,
                    "destination_iata": destination.iata_code,
                    "is_active": "true",
                    "valid_from": "2026-03-07",
                    "valid_to": "2026-03-21",
                }
            ],
            stats,
            create_missing_organizations=False,
        )

        existing_binding.refresh_from_db()
        self.assertEqual(RecipientBinding.objects.count(), 1)
        self.assertIsNotNone(existing_binding.valid_to)
        self.assertEqual(stats.recipient_bindings_updated, 1)
        self.assertEqual(stats.recipient_bindings_created, 0)

    def test_apply_recipient_bindings_updates_exact_binding_and_disables_other_active(self):
        recipient = self._create_org("Recipient Exact")
        shipper = self._create_org("Shipper Exact")
        destination = self._create_destination("NIM")
        now = timezone.now()
        old_active = RecipientBinding.objects.create(
            shipper_org=shipper,
            recipient_org=recipient,
            destination=destination,
            is_active=True,
            valid_from=now - timedelta(days=3),
        )
        exact_binding = RecipientBinding.objects.create(
            shipper_org=shipper,
            recipient_org=recipient,
            destination=destination,
            is_active=False,
            valid_from=now - timedelta(days=1),
        )
        valid_from_value = (now - timedelta(days=1)).isoformat()

        stats = command_module.ImportStats()
        command_module._apply_recipient_bindings(
            [
                {
                    "_row_number": "2",
                    "recipient_organization_name": recipient.name,
                    "shipper_organization_name": shipper.name,
                    "destination_iata": destination.iata_code,
                    "is_active": "true",
                    "valid_from": valid_from_value,
                    "valid_to": "2026-04-01",
                }
            ],
            stats,
            create_missing_organizations=False,
        )

        old_active.refresh_from_db()
        exact_binding.refresh_from_db()
        self.assertFalse(old_active.is_active)
        self.assertTrue(exact_binding.is_active)
        self.assertEqual(stats.recipient_bindings_updated, 1)
        self.assertEqual(stats.recipient_bindings_created, 0)

    def test_apply_recipient_bindings_creates_inactive_binding(self):
        recipient = self._create_org("Recipient Inactive")
        shipper = self._create_org("Shipper Inactive")
        destination = self._create_destination("OUA")
        stats = command_module.ImportStats()

        command_module._apply_recipient_bindings(
            [
                {
                    "_row_number": "2",
                    "recipient_organization_name": recipient.name,
                    "shipper_organization_name": shipper.name,
                    "destination_iata": destination.iata_code,
                    "is_active": "false",
                    "valid_from": "2026-03-08",
                    "valid_to": "",
                }
            ],
            stats,
            create_missing_organizations=False,
        )

        binding = RecipientBinding.objects.get(
            shipper_org=shipper,
            recipient_org=recipient,
            destination=destination,
        )
        self.assertFalse(binding.is_active)
        self.assertEqual(stats.recipient_bindings_created, 1)

    def test_apply_correspondents_rejects_invalid_rows(self):
        stats = command_module.ImportStats()
        with self.assertRaisesMessage(CommandError, "correspondent_organization_name"):
            command_module._apply_correspondents(
                [
                    {
                        "_row_number": "2",
                        "correspondent_organization_name": "",
                        "destination_iata": "",
                    }
                ],
                stats,
                create_missing_organizations=False,
            )

    def test_apply_correspondents_default_creates_then_updates(self):
        correspondent_org = self._create_org("Default Corr")
        destination = self._create_destination("DKR")
        stats = command_module.ImportStats()

        command_module._apply_correspondents(
            [
                {
                    "_row_number": "2",
                    "correspondent_organization_name": correspondent_org.name,
                    "destination_iata": destination.iata_code,
                    "scope_type": "default",
                    "is_active": "true",
                }
            ],
            stats,
            create_missing_organizations=False,
        )
        self.assertEqual(stats.correspondent_defaults_created, 1)

        update_stats = command_module.ImportStats()
        command_module._apply_correspondents(
            [
                {
                    "_row_number": "3",
                    "correspondent_organization_name": correspondent_org.name,
                    "destination_iata": destination.iata_code,
                    "scope_type": "default",
                    "is_active": "false",
                }
            ],
            update_stats,
            create_missing_organizations=False,
        )

        default_scope = DestinationCorrespondentDefault.objects.get(
            destination=destination,
            correspondent_org=correspondent_org,
        )
        self.assertFalse(default_scope.is_active)
        self.assertEqual(update_stats.correspondent_defaults_updated, 1)

    def test_apply_correspondents_override_paths(self):
        correspondent_org = self._create_org("Override Corr")
        shipper_org = self._create_org("Override Shipper")
        recipient_org = self._create_org("Override Recipient")
        destination = self._create_destination("BKO")

        stats = command_module.ImportStats()
        command_module._apply_correspondents(
            [
                {
                    "_row_number": "2",
                    "correspondent_organization_name": correspondent_org.name,
                    "destination_iata": destination.iata_code,
                    "scope_type": "shipper_and_recipient_override",
                    "shipper_organization_name": shipper_org.name,
                    "recipient_organization_name": recipient_org.name,
                    "is_active": "true",
                }
            ],
            stats,
            create_missing_organizations=False,
        )
        self.assertEqual(stats.correspondent_overrides_created, 1)

        update_stats = command_module.ImportStats()
        command_module._apply_correspondents(
            [
                {
                    "_row_number": "3",
                    "correspondent_organization_name": correspondent_org.name,
                    "destination_iata": destination.iata_code,
                    "scope_type": "shipper_and_recipient_override",
                    "shipper_organization_name": shipper_org.name,
                    "recipient_organization_name": recipient_org.name,
                    "is_active": "false",
                }
            ],
            update_stats,
            create_missing_organizations=False,
        )

        override = DestinationCorrespondentOverride.objects.get(
            destination=destination,
            correspondent_org=correspondent_org,
            shipper_org=shipper_org,
            recipient_org=recipient_org,
        )
        self.assertFalse(override.is_active)
        self.assertEqual(update_stats.correspondent_overrides_updated, 1)

    def test_apply_correspondents_rejects_invalid_scope_and_missing_scope_inputs(self):
        correspondent_org = self._create_org("Corr Invalid")
        destination = self._create_destination("YAO")
        stats = command_module.ImportStats()

        with self.assertRaisesMessage(CommandError, "scope_type invalide"):
            command_module._apply_correspondents(
                [
                    {
                        "_row_number": "2",
                        "correspondent_organization_name": correspondent_org.name,
                        "destination_iata": destination.iata_code,
                        "scope_type": "invalid_scope",
                    }
                ],
                stats,
                create_missing_organizations=False,
            )

        with self.assertRaisesMessage(CommandError, "shipper_organization_name requis"):
            command_module._apply_correspondents(
                [
                    {
                        "_row_number": "3",
                        "correspondent_organization_name": correspondent_org.name,
                        "destination_iata": destination.iata_code,
                        "scope_type": "shipper_override",
                        "shipper_organization_name": "",
                    }
                ],
                stats,
                create_missing_organizations=False,
            )

        with self.assertRaisesMessage(CommandError, "recipient_organization_name requis"):
            command_module._apply_correspondents(
                [
                    {
                        "_row_number": "4",
                        "correspondent_organization_name": correspondent_org.name,
                        "destination_iata": destination.iata_code,
                        "scope_type": "recipient_override",
                        "recipient_organization_name": "",
                    }
                ],
                stats,
                create_missing_organizations=False,
            )

    def test_find_or_create_organization_contact_paths(self):
        organization = self._create_org("Org Contact")
        stats = command_module.ImportStats()

        with self.assertRaisesMessage(CommandError, "contact_email ou"):
            command_module._find_or_create_organization_contact(
                organization=organization,
                row={"is_active": "true"},
                row_label="OrganizationContacts row 2",
                stats=stats,
            )

        contact, created = command_module._find_or_create_organization_contact(
            organization=organization,
            row={
                "contact_email": "ALICE@EXAMPLE.COM",
                "contact_first_name": "Alice",
                "contact_last_name": "Old",
                "contact_phone": "",
                "is_active": "true",
            },
            row_label="OrganizationContacts row 3",
            stats=stats,
        )
        self.assertTrue(created)
        self.assertEqual(stats.organization_contacts_created, 1)

        updated_contact, updated_created = command_module._find_or_create_organization_contact(
            organization=organization,
            row={
                "contact_email": "alice@example.com",
                "contact_first_name": "Alicia",
                "contact_last_name": "Ops",
                "contact_phone": "12345",
                "is_active": "false",
            },
            row_label="OrganizationContacts row 4",
            stats=stats,
        )
        self.assertFalse(updated_created)
        self.assertEqual(updated_contact.id, contact.id)
        self.assertEqual(updated_contact.first_name, "Alicia")
        self.assertEqual(updated_contact.last_name, "Ops")
        self.assertEqual(updated_contact.email, "alice@example.com")
        self.assertEqual(updated_contact.phone, "12345")
        self.assertFalse(updated_contact.is_active)
        self.assertEqual(stats.organization_contacts_updated, 1)

    def test_apply_subscriptions_rejects_invalid_channel(self):
        organization = self._create_org("Notifier Invalid")
        assignment = OrganizationRoleAssignment.objects.create(
            organization=organization,
            role=OrganizationRole.SHIPPER,
            is_active=True,
        )
        contact = OrganizationContact.objects.create(
            organization=organization,
            first_name="Invalid",
            last_name="Channel",
            email="invalid@example.com",
            is_active=True,
        )
        role_contact = OrganizationRoleContact.objects.create(
            role_assignment=assignment,
            contact=contact,
            is_primary=False,
            is_active=True,
        )

        with self.assertRaisesMessage(CommandError, "notification_channel invalide"):
            command_module._apply_subscriptions(
                role_contact=role_contact,
                row={"notification_channel": "sms"},
                row_label="OrganizationContacts row 2",
                stats=command_module.ImportStats(),
                create_missing_organizations=False,
            )

    def test_apply_organization_contacts_rejects_invalid_row(self):
        stats = command_module.ImportStats()
        with self.assertRaisesMessage(CommandError, "organization_name/role invalides"):
            command_module._apply_organization_contacts(
                [{"_row_number": "2", "organization_name": "", "role": ""}],
                stats,
                create_missing_organizations=False,
            )

    def test_apply_organization_contacts_creates_and_updates_role_contact(self):
        organization = self._create_org("Contacts Org")
        assignment = OrganizationRoleAssignment.objects.create(
            organization=organization,
            role=OrganizationRole.SHIPPER,
            is_active=True,
        )
        existing_contact = OrganizationContact.objects.create(
            organization=organization,
            first_name="Existing",
            last_name="Primary",
            email="existing@example.com",
            is_active=True,
        )
        existing_role_contact = OrganizationRoleContact.objects.create(
            role_assignment=assignment,
            contact=existing_contact,
            is_primary=False,
            is_active=True,
        )

        create_stats = command_module.ImportStats()
        command_module._apply_organization_contacts(
            [
                {
                    "_row_number": "2",
                    "organization_name": organization.name,
                    "role": OrganizationRole.SHIPPER,
                    "contact_email": "new@example.com",
                    "contact_first_name": "New",
                    "contact_last_name": "Primary",
                    "is_primary": "false",
                    "is_active": "true",
                    "notification_channel": "email",
                    "notify_shipment_status_updated": "true",
                }
            ],
            create_stats,
            create_missing_organizations=False,
        )

        new_contact = OrganizationContact.objects.get(email="new@example.com")
        role_contact = OrganizationRoleContact.objects.get(
            role_assignment=assignment,
            contact=new_contact,
        )
        self.assertFalse(role_contact.is_primary)
        self.assertEqual(create_stats.organization_role_contacts_created, 1)
        self.assertEqual(create_stats.subscriptions_created, 1)

        promote_stats = command_module.ImportStats()
        command_module._apply_organization_contacts(
            [
                {
                    "_row_number": "3",
                    "organization_name": organization.name,
                    "role": OrganizationRole.SHIPPER,
                    "contact_email": "new@example.com",
                    "is_primary": "true",
                    "is_active": "true",
                    "notification_channel": "email",
                    "notify_shipment_status_updated": "false",
                }
            ],
            promote_stats,
            create_missing_organizations=False,
        )
        role_contact.refresh_from_db()
        existing_role_contact.refresh_from_db()
        self.assertTrue(role_contact.is_primary)
        self.assertFalse(existing_role_contact.is_primary)
        self.assertEqual(promote_stats.organization_role_contacts_updated, 1)

        update_stats = command_module.ImportStats()
        command_module._apply_organization_contacts(
            [
                {
                    "_row_number": "4",
                    "organization_name": organization.name,
                    "role": OrganizationRole.SHIPPER,
                    "contact_email": "new@example.com",
                    "is_primary": "false",
                    "is_active": "false",
                    "notification_channel": "email",
                    "notify_shipment_status_updated": "false",
                }
            ],
            update_stats,
            create_missing_organizations=False,
        )
        role_contact.refresh_from_db()
        self.assertFalse(role_contact.is_primary)
        self.assertFalse(role_contact.is_active)
        self.assertEqual(update_stats.organization_role_contacts_updated, 1)

    def test_apply_migration_review_skips_missing_action(self):
        stats = command_module.ImportStats()
        command_module._apply_migration_review(
            [{"_row_number": "2", "row_id": "", "resolution_action": ""}],
            stats,
            create_missing_organizations=False,
        )
        self.assertEqual(stats.migration_review_rows, 1)
        self.assertEqual(stats.migration_review_skipped, 1)

    def test_apply_migration_review_rejects_invalid_action(self):
        stats = command_module.ImportStats()
        with self.assertRaisesMessage(CommandError, "resolution_action invalide"):
            command_module._apply_migration_review(
                [{"_row_number": "2", "resolution_action": "bad_action", "row_id": "MR-1"}],
                stats,
                create_missing_organizations=False,
            )

    def test_apply_migration_review_rejects_invalid_row_id_formats(self):
        stats = command_module.ImportStats()
        with self.assertRaisesMessage(CommandError, "row_id attendu au format MR-<id>"):
            command_module._apply_migration_review(
                [{"_row_number": "2", "resolution_action": "resolve_binding", "row_id": "1"}],
                stats,
                create_missing_organizations=False,
            )

        with self.assertRaisesMessage(CommandError, "row_id invalide"):
            command_module._apply_migration_review(
                [{"_row_number": "3", "resolution_action": "resolve_binding", "row_id": "MR-abc"}],
                stats,
                create_missing_organizations=False,
            )

    def test_apply_migration_review_rejects_missing_item(self):
        stats = command_module.ImportStats()
        with self.assertRaisesMessage(CommandError, "item de revue introuvable"):
            command_module._apply_migration_review(
                [{"_row_number": "2", "resolution_action": "resolve_binding", "row_id": "MR-999999"}],
                stats,
                create_missing_organizations=False,
            )

    def test_apply_migration_review_skips_non_open_item(self):
        organization = self._create_org("Recipient Closed")
        review_item = MigrationReviewItem.objects.create(
            organization=organization,
            role=OrganizationRole.RECIPIENT,
            reason_code="already_closed",
            status=MigrationReviewItemStatus.RESOLVED,
        )
        stats = command_module.ImportStats()
        command_module._apply_migration_review(
            [
                {
                    "_row_number": "2",
                    "row_id": f"MR-{review_item.id}",
                    "resolution_action": "resolve_without_binding",
                }
            ],
            stats,
            create_missing_organizations=False,
        )
        self.assertEqual(stats.migration_review_skipped, 1)

    def test_apply_migration_review_resolve_binding_requires_shipper_and_destination(self):
        organization = self._create_org("Recipient Missing Inputs")
        review_item = MigrationReviewItem.objects.create(
            organization=organization,
            role=OrganizationRole.RECIPIENT,
            reason_code="missing_data",
            status=MigrationReviewItemStatus.OPEN,
        )

        with self.assertRaisesMessage(CommandError, "shipper et destination proposes requis"):
            command_module._apply_migration_review(
                [
                    {
                        "_row_number": "2",
                        "row_id": f"MR-{review_item.id}",
                        "resolution_action": "resolve_binding",
                        "proposed_shipper_organization_name": "",
                        "proposed_destination_iata": "",
                    }
                ],
                command_module.ImportStats(),
                create_missing_organizations=False,
            )

    def test_apply_migration_review_resolve_binding_rejects_unresolvable_recipient(self):
        review_item = MigrationReviewItem.objects.create(
            organization=None,
            legacy_contact=None,
            role=OrganizationRole.RECIPIENT,
            reason_code="missing_recipient",
            status=MigrationReviewItemStatus.OPEN,
        )
        shipper = self._create_org("Resolvable Shipper")
        destination = self._create_destination("MRS")

        with self.assertRaisesMessage(CommandError, "destinataire introuvable"):
            command_module._apply_migration_review(
                [
                    {
                        "_row_number": "2",
                        "row_id": f"MR-{review_item.id}",
                        "resolution_action": "resolve_binding",
                        "proposed_shipper_organization_name": shipper.name,
                        "proposed_destination_iata": destination.iata_code,
                    }
                ],
                command_module.ImportStats(),
                create_missing_organizations=False,
            )

    def test_apply_migration_review_resolve_binding_creates_scope_and_binding(self):
        recipient = self._create_org("Recipient Resolve")
        shipper = self._create_org("Shipper Resolve")
        destination = self._create_destination("TUN")
        review_item = MigrationReviewItem.objects.create(
            organization=recipient,
            role=OrganizationRole.RECIPIENT,
            reason_code="missing_binding",
            status=MigrationReviewItemStatus.OPEN,
        )
        stats = command_module.ImportStats()

        command_module._apply_migration_review(
            [
                {
                    "_row_number": "2",
                    "row_id": f"MR-{review_item.id}",
                    "resolution_action": "resolve_binding",
                    "resolution_note": "resolved by import",
                    "proposed_shipper_organization_name": shipper.name,
                    "proposed_destination_iata": destination.iata_code,
                }
            ],
            stats,
            create_missing_organizations=False,
        )

        review_item.refresh_from_db()
        self.assertEqual(review_item.status, MigrationReviewItemStatus.RESOLVED)
        self.assertTrue(
            ShipperScope.objects.filter(
                role_assignment__organization=shipper,
                destination=destination,
                is_active=True,
            ).exists()
        )
        self.assertTrue(
            RecipientBinding.objects.filter(
                shipper_org=shipper,
                recipient_org=recipient,
                destination=destination,
                is_active=True,
            ).exists()
        )
        self.assertEqual(stats.shipper_scopes_created, 1)
        self.assertEqual(stats.recipient_bindings_created, 1)
        self.assertEqual(stats.migration_review_resolved, 1)

    def test_apply_migration_review_resolve_binding_updates_inactive_scope_without_new_binding(self):
        recipient = self._create_org("Recipient Existing")
        shipper = self._create_org("Shipper Existing")
        destination = self._create_destination("RAK")
        shipper_assignment = OrganizationRoleAssignment.objects.create(
            organization=shipper,
            role=OrganizationRole.SHIPPER,
            is_active=True,
        )
        scope = ShipperScope.objects.create(
            role_assignment=shipper_assignment,
            destination=destination,
            all_destinations=False,
            is_active=False,
            valid_from=timezone.now() - timedelta(days=4),
        )
        RecipientBinding.objects.create(
            shipper_org=shipper,
            recipient_org=recipient,
            destination=destination,
            is_active=True,
            valid_from=timezone.now() - timedelta(days=3),
        )
        review_item = MigrationReviewItem.objects.create(
            organization=recipient,
            role=OrganizationRole.RECIPIENT,
            reason_code="missing_binding",
            status=MigrationReviewItemStatus.OPEN,
        )
        stats = command_module.ImportStats()

        command_module._apply_migration_review(
            [
                {
                    "_row_number": "2",
                    "row_id": f"MR-{review_item.id}",
                    "resolution_action": "resolve_binding",
                    "proposed_shipper_organization_name": shipper.name,
                    "proposed_destination_iata": destination.iata_code,
                }
            ],
            stats,
            create_missing_organizations=False,
        )

        scope.refresh_from_db()
        self.assertTrue(scope.is_active)
        self.assertEqual(stats.shipper_scopes_created, 0)
        self.assertEqual(stats.shipper_scopes_updated, 1)
        self.assertEqual(stats.recipient_bindings_created, 0)

    def test_apply_subscriptions_replaces_existing_subscriptions(self):
        organization = self._create_org("Notifier Org")
        shipper_filter = self._create_org("Shipper Filter")
        recipient_filter = self._create_org("Recipient Filter")
        destination = self._create_destination("ABJ")

        assignment = OrganizationRoleAssignment.objects.create(
            organization=organization,
            role=OrganizationRole.SHIPPER,
            is_active=False,
        )
        contact = OrganizationContact.objects.create(
            organization=organization,
            first_name="Alice",
            last_name="Ops",
            email="alice@example.com",
            is_active=True,
        )
        role_contact = OrganizationRoleContact.objects.create(
            role_assignment=assignment,
            contact=contact,
            is_primary=True,
            is_active=True,
        )
        ContactSubscription.objects.create(
            role_contact=role_contact,
            event_type=RoleEventType.SHIPMENT_DELIVERED,
            channel=NotificationChannel.EMAIL,
            is_active=True,
        )

        stats = command_module.ImportStats()
        command_module._apply_subscriptions(
            role_contact=role_contact,
            row={
                "_row_number": "2",
                "notification_channel": "email",
                "destination_iata_filter": "ABJ",
                "shipper_organization_filter": "Shipper Filter",
                "recipient_organization_filter": "Recipient Filter",
                "notify_shipment_status_updated": "true",
                "notify_shipment_delivered": "false",
                "notify_shipment_tracking_updated": "true",
                "notify_order_document_requested": "",
            },
            row_label="OrganizationContacts row 2",
            stats=stats,
            create_missing_organizations=False,
        )

        self.assertEqual(stats.subscriptions_created, 2)
        subscriptions = ContactSubscription.objects.filter(role_contact=role_contact).order_by(
            "event_type"
        )
        self.assertEqual(subscriptions.count(), 2)
        self.assertEqual(
            set(subscriptions.values_list("event_type", flat=True)),
            {
                RoleEventType.SHIPMENT_STATUS_UPDATED,
                RoleEventType.SHIPMENT_TRACKING_UPDATED,
            },
        )
        self.assertEqual(subscriptions[0].destination_id, destination.id)
        self.assertEqual(subscriptions[0].shipper_org_id, shipper_filter.id)
        self.assertEqual(subscriptions[0].recipient_org_id, recipient_filter.id)

    def test_apply_migration_review_resolve_without_binding(self):
        organization = self._create_org("Recipient Org")
        review_item = MigrationReviewItem.objects.create(
            organization=organization,
            role=OrganizationRole.RECIPIENT,
            reason_code="missing_binding",
            status=MigrationReviewItemStatus.OPEN,
        )
        stats = command_module.ImportStats()

        command_module._apply_migration_review(
            [
                {
                    "_row_number": "2",
                    "row_id": f"MR-{review_item.id}",
                    "resolution_action": "resolve_without_binding",
                    "resolution_note": "validated",
                }
            ],
            stats,
            create_missing_organizations=False,
        )

        review_item.refresh_from_db()
        self.assertEqual(review_item.status, MigrationReviewItemStatus.RESOLVED)
        self.assertEqual(review_item.resolution_note, "validated")
        self.assertIsNotNone(review_item.resolved_at)
        self.assertEqual(stats.migration_review_rows, 1)
        self.assertEqual(stats.migration_review_resolved, 1)
