from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest import mock

from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase
from openpyxl import Workbook, load_workbook

from contacts.models import Contact, ContactType
from wms import models
from wms.management.commands import export_org_roles_review_template as command_module


class ExportOrgRolesReviewTemplateHelpersTests(TestCase):
    def _create_org(self, name: str) -> Contact:
        return Contact.objects.create(
            name=name,
            contact_type=ContactType.ORGANIZATION,
            is_active=True,
        )

    def test_project_root_and_resolve_path(self):
        project_root = command_module._project_root()
        self.assertTrue((project_root / "manage.py").exists())

        with TemporaryDirectory() as tmp_dir:
            absolute = Path(tmp_dir) / "template.xlsx"
            self.assertEqual(command_module._resolve_path(str(absolute)), absolute)

        with mock.patch.object(command_module, "_project_root", return_value=Path("/tmp/repo")):
            resolved = command_module._resolve_path("docs/import/input.xlsx")
        self.assertEqual(resolved, Path("/tmp/repo/docs/import/input.xlsx").resolve())

    def test_resolve_suggested_helpers(self):
        options = [
            SimpleNamespace(id=10, name="Shipper A"),
            SimpleNamespace(id=20, name="Shipper B"),
        ]
        destinations = [
            SimpleNamespace(id=1, iata_code="DLA", city="Douala"),
            SimpleNamespace(id=2, iata_code=None, city=None),
        ]

        self.assertEqual(
            command_module._resolve_suggested_name_by_id(options=options, suggested_id=20),
            "Shipper B",
        )
        self.assertEqual(
            command_module._resolve_suggested_name_by_id(options=options, suggested_id=999),
            "",
        )
        self.assertEqual(
            command_module._resolve_suggested_name_by_id(options=options, suggested_id=None),
            "",
        )

        self.assertEqual(
            command_module._resolve_suggested_destination(
                destination_options=destinations,
                suggested_destination_id=1,
            ),
            ("DLA", "Douala"),
        )
        self.assertEqual(
            command_module._resolve_suggested_destination(
                destination_options=destinations,
                suggested_destination_id=2,
            ),
            ("", ""),
        )
        self.assertEqual(
            command_module._resolve_suggested_destination(
                destination_options=destinations,
                suggested_destination_id=99,
            ),
            ("", ""),
        )

    def test_open_review_items_queryset_filters_open_by_default(self):
        open_org = self._create_org("Recipient Open")
        resolved_org = self._create_org("Recipient Resolved")

        open_item = models.MigrationReviewItem.objects.create(
            organization=open_org,
            role=models.OrganizationRole.RECIPIENT,
            reason_code="missing_links_open",
            status=models.MigrationReviewItemStatus.OPEN,
        )
        resolved_item = models.MigrationReviewItem.objects.create(
            organization=resolved_org,
            role=models.OrganizationRole.RECIPIENT,
            reason_code="missing_links_resolved",
            status=models.MigrationReviewItemStatus.RESOLVED,
        )

        open_ids = {
            item.id for item in command_module._open_review_items_queryset(include_resolved=False)
        }
        all_ids = {
            item.id for item in command_module._open_review_items_queryset(include_resolved=True)
        }

        self.assertEqual(open_ids, {open_item.id})
        self.assertEqual(all_ids, {open_item.id, resolved_item.id})

    def test_build_export_rows_sets_default_action_only_when_shipper_and_destination_exist(self):
        first_item = SimpleNamespace(id=11, reason_code="missing_binding")
        second_item = SimpleNamespace(id=12, reason_code="missing_binding")

        review_rows = [
            {
                "recipient_org": SimpleNamespace(name="Recipient A"),
                "shipper_options": [SimpleNamespace(id=100, name="Shipper A")],
                "suggested_shipper_id": 100,
                "destination_options": [SimpleNamespace(id=55, iata_code="ABJ", city="Abidjan")],
                "suggested_destination_id": 55,
            },
            {
                "recipient_org": None,
                "shipper_options": [],
                "suggested_shipper_id": None,
                "destination_options": [],
                "suggested_destination_id": None,
            },
        ]

        with (
            mock.patch.object(
                command_module,
                "_open_review_items_queryset",
                return_value=[first_item, second_item],
            ),
            mock.patch.object(
                command_module,
                "_build_review_row",
                side_effect=review_rows,
            ),
        ):
            rows = command_module._build_export_rows(include_resolved=False)

        self.assertEqual(
            rows[0],
            [
                "MR-11",
                "Recipient A",
                "missing_binding",
                "Shipper A",
                "ABJ",
                "Abidjan",
                "resolve_binding",
                "",
            ],
        )
        self.assertEqual(
            rows[1],
            ["MR-12", "", "missing_binding", "", "", "", "", ""],
        )


class ExportOrgRolesReviewTemplateCommandTests(TestCase):
    def _create_template(self, *, path: Path, include_target_sheet: bool):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = command_module.TARGET_SHEET_NAME if include_target_sheet else "OtherSheet"
        worksheet.append(
            [
                "review_id",
                "recipient_org",
                "reason_code",
                "shipper",
                "dest_iata",
                "dest_city",
                "default_action",
                "note",
            ]
        )
        worksheet.append(["STALE", "ROW"])
        workbook.save(path)

    def test_handle_raises_on_missing_template(self):
        missing = "/tmp/does-not-exist-export-template.xlsx"
        with self.assertRaisesMessage(CommandError, "Template introuvable"):
            call_command(
                "export_org_roles_review_template",
                template=missing,
                output="/tmp/out.xlsx",
            )

    def test_handle_raises_when_target_sheet_missing(self):
        with TemporaryDirectory() as tmp_dir:
            template_path = Path(tmp_dir) / "template.xlsx"
            output_path = Path(tmp_dir) / "out.xlsx"
            self._create_template(path=template_path, include_target_sheet=False)

            with self.assertRaisesMessage(CommandError, "Onglet 'MigrationReview' absent"):
                call_command(
                    "export_org_roles_review_template",
                    template=str(template_path),
                    output=str(output_path),
                )

    def test_handle_writes_rows_and_removes_previous_data_rows(self):
        with TemporaryDirectory() as tmp_dir:
            template_path = Path(tmp_dir) / "template.xlsx"
            output_path = Path(tmp_dir) / "nested" / "filled.xlsx"
            self._create_template(path=template_path, include_target_sheet=True)
            rows = [
                [
                    "MR-1",
                    "Recipient 1",
                    "reason_a",
                    "Shipper A",
                    "DLA",
                    "Douala",
                    "resolve_binding",
                    "",
                ],
                ["MR-2", "Recipient 2", "reason_b", "", "", "", "", ""],
            ]

            with mock.patch.object(command_module, "_build_export_rows", return_value=rows):
                call_command(
                    "export_org_roles_review_template",
                    template=str(template_path),
                    output=str(output_path),
                )

            workbook = load_workbook(output_path)
            worksheet = workbook[command_module.TARGET_SHEET_NAME]
            values = list(worksheet.iter_rows(values_only=True))

            self.assertEqual(values[0][0:3], ("review_id", "recipient_org", "reason_code"))
            self.assertEqual(values[1][0:7], tuple(rows[0][0:7]))
            self.assertEqual(values[2][0:3], tuple(rows[1][0:3]))
            self.assertEqual(values[2][3:7], (None, None, None, None))
            self.assertEqual(len(values), 3)
