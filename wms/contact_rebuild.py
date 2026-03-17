from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from contacts.models import Contact, ContactTag, ContactType
from contacts.tagging import TAG_CORRESPONDENT, TAG_DONOR, TAG_RECIPIENT, TAG_SHIPPER
from wms.models import (
    Destination,
    OrganizationRole,
    OrganizationRoleAssignment,
    RecipientBinding,
    ShipperScope,
)

CANONICAL_BE_SHEET_NAMES = ("2024", "2025", "2026")
CANONICAL_DESTINATION_CORRESPONDENT_OVERRIDES = {
    "BEY": "Tony MDAWAR",
    "BGF": "Christian LIMBIO",
    "NDJ": "Geovanie Kamtar NDANGMBAYE",
}


@dataclass
class BeContactDataset:
    donors: list[dict]
    shippers: list[dict]
    recipients: list[dict]
    correspondents: list[dict]
    destinations: list[dict]
    shipper_scopes: list[dict]
    recipient_bindings: list[dict]
    review_items: list[dict]
    source_sheets: list[str] = field(default_factory=list)


def _normalize_text(value) -> str:
    return str(value or "").strip()


def _normalize_key(value) -> str:
    text = unicodedata.normalize("NFKD", _normalize_text(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _normalize_display_name(value) -> str:
    text = _normalize_text(value)
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()


def _normalize_iata(value) -> str:
    return _normalize_text(value).upper()


def _selected_sheet_names(workbook) -> list[str]:
    names = [name for name in CANONICAL_BE_SHEET_NAMES if name in workbook.sheetnames]
    if names:
        return names
    return [workbook.active.title]


def _iter_sheet_rows(sheet):
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, ())
    header_map: dict[str, list[int]] = {}
    for index, header in enumerate(headers or ()):
        key = _normalize_text(header)
        if not key:
            continue
        header_map.setdefault(key, []).append(index)
    for row in rows:
        if not any(_normalize_text(value) for value in row):
            continue
        yield row, header_map


def _get_cell(row, header_map, name: str) -> str:
    for index in header_map.get(name, []):
        value = _normalize_text(row[index])
        if value:
            return value
    return ""


def _note_conflict(
    *,
    review_items: list[dict],
    seen_conflicts: set[tuple[str, str, str]],
    entity_type: str,
    key: str,
    display_name: str,
    field_name: str,
    existing,
    incoming,
):
    marker = (entity_type, key, field_name)
    if marker in seen_conflicts:
        return
    seen_conflicts.add(marker)
    review_items.append(
        {
            "entity_type": entity_type,
            "key": key,
            "display_name": display_name or key,
            "reason": f"conflicting {entity_type} {field_name}",
            "existing": existing,
            "incoming": incoming,
        }
    )


def _merge_scalar(
    *,
    record: dict,
    field_name: str,
    value: str,
    source_priority: int,
    review_items: list[dict],
    seen_conflicts: set[tuple[str, str, str]],
    entity_type: str,
    key: str,
):
    if not value:
        return
    existing = record.get(field_name, "")
    field_priorities = record.setdefault("_field_priorities", {})
    existing_priority = field_priorities.get(field_name, -1)
    if not existing:
        record[field_name] = value
        field_priorities[field_name] = source_priority
        return
    if source_priority > existing_priority:
        record[field_name] = value
        field_priorities[field_name] = source_priority
        return
    if source_priority < existing_priority:
        return
    if _normalize_key(existing) == _normalize_key(value):
        if len(value) > len(existing):
            record[field_name] = value
        return
    _note_conflict(
        review_items=review_items,
        seen_conflicts=seen_conflicts,
        entity_type=entity_type,
        key=key,
        display_name=(record.get("name") or record.get("city") or record.get("iata_code") or key),
        field_name=field_name,
        existing=existing,
        incoming=value,
    )


def _get_or_create_entity(store: dict[str, dict], *, key: str) -> dict:
    if key not in store:
        store[key] = {
            "key": key,
            "_field_priorities": {},
        }
    return store[key]


def _clean_record(record: dict) -> dict:
    return {key: value for key, value in record.items() if not key.startswith("_")}


def _apply_destination_correspondent_overrides(
    *,
    destinations: dict[str, dict],
    correspondents: dict[str, dict],
):
    for destination in destinations.values():
        iata_code = _normalize_iata(destination.get("iata_code"))
        override_name = CANONICAL_DESTINATION_CORRESPONDENT_OVERRIDES.get(iata_code, "")
        if not override_name:
            continue
        override_key = _normalize_key(override_name)
        destination["correspondent_key"] = override_key
        correspondent = _get_or_create_entity(correspondents, key=override_key)
        _merge_scalar(
            record=correspondent,
            field_name="name",
            value=override_name,
            source_priority=len(CANONICAL_BE_SHEET_NAMES),
            review_items=[],
            seen_conflicts=set(),
            entity_type="correspondent",
            key=override_key,
        )
        if destination.get("country"):
            _merge_scalar(
                record=correspondent,
                field_name="country",
                value=destination["country"],
                source_priority=len(CANONICAL_BE_SHEET_NAMES),
                review_items=[],
                seen_conflicts=set(),
                entity_type="correspondent",
                key=override_key,
            )


def _prune_unreferenced_correspondents(
    *,
    destinations: dict[str, dict],
    correspondents: dict[str, dict],
) -> dict[str, dict]:
    referenced_keys = {
        destination.get("correspondent_key", "")
        for destination in destinations.values()
        if destination.get("correspondent_key")
    }
    return {key: record for key, record in correspondents.items() if key in referenced_keys}


def build_be_contact_dataset(path: str | Path) -> BeContactDataset:
    donors: dict[str, dict] = {}
    shippers: dict[str, dict] = {}
    recipients: dict[str, dict] = {}
    correspondents: dict[str, dict] = {}
    destinations: dict[str, dict] = {}
    shipper_scopes: set[tuple[str, str]] = set()
    recipient_bindings: set[tuple[str, str, str]] = set()
    review_items: list[dict] = []
    seen_conflicts: set[tuple[str, str, str]] = set()
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        source_sheets = _selected_sheet_names(workbook)
        for source_priority, sheet_name in enumerate(source_sheets):
            sheet = workbook[sheet_name]
            for row, header_map in _iter_sheet_rows(sheet):
                donor_name = _normalize_display_name(_get_cell(row, header_map, "BE_DONATEUR"))
                shipper_name = _normalize_display_name(
                    _get_cell(row, header_map, "ASSOCIATION_NOM")
                )
                shipper_country = _normalize_display_name(
                    _get_cell(row, header_map, "ASSOCIATION_PAYS")
                )
                recipient_name = _normalize_display_name(
                    _get_cell(row, header_map, "DESTINATAIRE_STRUCTURE")
                )
                recipient_status = _normalize_display_name(
                    _get_cell(row, header_map, "DESTINATAIRE_STATUT")
                )
                correspondent_name = _normalize_display_name(
                    " ".join(
                        part
                        for part in (
                            _get_cell(row, header_map, "CORRESPONDANT_PRENOM"),
                            _get_cell(row, header_map, "CORRESPONDANT_NOM"),
                        )
                        if part
                    )
                )
                correspondent_country = _normalize_display_name(
                    _get_cell(row, header_map, "CORRESPONDANT_PAYS")
                )
                destination_city = _normalize_display_name(
                    _get_cell(row, header_map, "BE_DESTINATION")
                )
                destination_iata = _normalize_iata(_get_cell(row, header_map, "BE_CODE_IATA"))

                donor_key = _normalize_key(donor_name)
                shipper_key = _normalize_key(shipper_name)
                recipient_key = _normalize_key(recipient_name)
                correspondent_key = _normalize_key(correspondent_name)

                if donor_key:
                    donor = _get_or_create_entity(donors, key=donor_key)
                    _merge_scalar(
                        record=donor,
                        field_name="name",
                        value=donor_name,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="donor",
                        key=donor_key,
                    )

                if shipper_key:
                    shipper = _get_or_create_entity(shippers, key=shipper_key)
                    _merge_scalar(
                        record=shipper,
                        field_name="name",
                        value=shipper_name,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="shipper",
                        key=shipper_key,
                    )
                    _merge_scalar(
                        record=shipper,
                        field_name="country",
                        value=shipper_country,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="shipper",
                        key=shipper_key,
                    )

                if recipient_key:
                    recipient = _get_or_create_entity(recipients, key=recipient_key)
                    _merge_scalar(
                        record=recipient,
                        field_name="name",
                        value=recipient_name,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="recipient",
                        key=recipient_key,
                    )
                    _merge_scalar(
                        record=recipient,
                        field_name="status",
                        value=recipient_status,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="recipient",
                        key=recipient_key,
                    )

                if correspondent_key:
                    correspondent = _get_or_create_entity(
                        correspondents,
                        key=correspondent_key,
                    )
                    _merge_scalar(
                        record=correspondent,
                        field_name="name",
                        value=correspondent_name,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="correspondent",
                        key=correspondent_key,
                    )
                    _merge_scalar(
                        record=correspondent,
                        field_name="country",
                        value=correspondent_country,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="correspondent",
                        key=correspondent_key,
                    )

                destination_key = destination_iata or _normalize_key(
                    f"{destination_city}|{correspondent_country}"
                )
                if destination_key and destination_city:
                    destination = destinations.setdefault(
                        destination_key,
                        {
                            "key": destination_key,
                            "_field_priorities": {},
                        },
                    )
                    _merge_scalar(
                        record=destination,
                        field_name="city",
                        value=destination_city,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="destination",
                        key=destination_key,
                    )
                    _merge_scalar(
                        record=destination,
                        field_name="country",
                        value=correspondent_country,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="destination",
                        key=destination_key,
                    )
                    _merge_scalar(
                        record=destination,
                        field_name="iata_code",
                        value=destination_iata,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="destination",
                        key=destination_key,
                    )
                    _merge_scalar(
                        record=destination,
                        field_name="correspondent_key",
                        value=correspondent_key,
                        source_priority=source_priority,
                        review_items=review_items,
                        seen_conflicts=seen_conflicts,
                        entity_type="destination",
                        key=destination_key,
                    )

                if shipper_key and destination_iata:
                    shipper_scopes.add((shipper_key, destination_iata))
                if recipient_key and shipper_key and destination_iata:
                    recipient_bindings.add((recipient_key, shipper_key, destination_iata))
    finally:
        workbook.close()

    _apply_destination_correspondent_overrides(
        destinations=destinations,
        correspondents=correspondents,
    )
    correspondents = _prune_unreferenced_correspondents(
        destinations=destinations,
        correspondents=correspondents,
    )

    return BeContactDataset(
        donors=sorted(
            (_clean_record(item) for item in donors.values()),
            key=lambda item: item["name"],
        ),
        shippers=sorted(
            (_clean_record(item) for item in shippers.values()),
            key=lambda item: item["name"],
        ),
        recipients=sorted(
            (_clean_record(item) for item in recipients.values()),
            key=lambda item: item["name"],
        ),
        correspondents=sorted(
            (_clean_record(item) for item in correspondents.values()),
            key=lambda item: item["name"],
        ),
        destinations=sorted(
            (_clean_record(item) for item in destinations.values()),
            key=lambda item: item["iata_code"],
        ),
        shipper_scopes=[
            {"shipper_key": shipper_key, "destination_iata": destination_iata}
            for shipper_key, destination_iata in sorted(shipper_scopes)
        ],
        recipient_bindings=[
            {
                "recipient_key": recipient_key,
                "shipper_key": shipper_key,
                "destination_iata": destination_iata,
            }
            for recipient_key, shipper_key, destination_iata in sorted(recipient_bindings)
        ],
        review_items=review_items,
        source_sheets=source_sheets,
    )


def render_review_report(review_items: list[dict]) -> str:
    if not review_items:
        return "# Contact rebuild review\n\nNo review items.\n"
    lines = ["# Contact rebuild review", ""]
    for item in review_items:
        lines.append(
            "- {entity_type} `{display_name}`: {reason} ({existing!r} -> {incoming!r})".format(
                **item
            )
        )
    lines.append("")
    return "\n".join(lines)


def _ensure_tag(name: str) -> ContactTag:
    tag, _ = ContactTag.objects.get_or_create(name=name)
    return tag


def _get_or_create_contact(*, name: str, contact_type: str) -> Contact:
    contact, _ = Contact.objects.get_or_create(
        name=name,
        contact_type=contact_type,
        defaults={"is_active": True},
    )
    if not contact.is_active:
        contact.is_active = True
        contact.save(update_fields=["is_active"])
    return contact


def _ensure_role_assignment(*, contact: Contact, role: str) -> OrganizationRoleAssignment:
    assignment, created = OrganizationRoleAssignment.objects.get_or_create(
        organization=contact,
        role=role,
        defaults={"is_active": True},
    )
    if not created and not assignment.is_active:
        assignment.is_active = True
        assignment.save(update_fields=["is_active"])
    return assignment


def apply_be_contact_dataset(dataset: BeContactDataset) -> None:
    with transaction.atomic():
        donor_tag = _ensure_tag(TAG_DONOR[0])
        shipper_tag = _ensure_tag(TAG_SHIPPER[0])
        recipient_tag = _ensure_tag(TAG_RECIPIENT[0])
        correspondent_tag = _ensure_tag(TAG_CORRESPONDENT[0])

        contacts_by_key: dict[str, Contact] = {}

        for donor in dataset.donors:
            contact = _get_or_create_contact(
                name=donor["name"],
                contact_type=ContactType.ORGANIZATION,
            )
            contact.tags.add(donor_tag)
            contacts_by_key[donor["key"]] = contact
            _ensure_role_assignment(contact=contact, role=OrganizationRole.DONOR)

        for shipper in dataset.shippers:
            contact = _get_or_create_contact(
                name=shipper["name"],
                contact_type=ContactType.ORGANIZATION,
            )
            contact.tags.add(shipper_tag)
            contacts_by_key[shipper["key"]] = contact
            _ensure_role_assignment(contact=contact, role=OrganizationRole.SHIPPER)

        for recipient in dataset.recipients:
            contact = _get_or_create_contact(
                name=recipient["name"],
                contact_type=ContactType.ORGANIZATION,
            )
            contact.tags.add(recipient_tag)
            contacts_by_key[recipient["key"]] = contact
            _ensure_role_assignment(contact=contact, role=OrganizationRole.RECIPIENT)

        for correspondent in dataset.correspondents:
            contact = _get_or_create_contact(
                name=correspondent["name"],
                contact_type=ContactType.PERSON,
            )
            contact.tags.add(correspondent_tag)
            contacts_by_key[correspondent["key"]] = contact

        destinations_by_iata: dict[str, Destination] = {}
        for destination_data in dataset.destinations:
            correspondent = contacts_by_key[destination_data["correspondent_key"]]
            destination, _ = Destination.objects.get_or_create(
                iata_code=destination_data["iata_code"],
                defaults={
                    "city": destination_data["city"],
                    "country": destination_data["country"],
                    "correspondent_contact": correspondent,
                    "is_active": True,
                },
            )
            updated_fields = []
            if destination.city != destination_data["city"]:
                destination.city = destination_data["city"]
                updated_fields.append("city")
            if destination.country != destination_data["country"]:
                destination.country = destination_data["country"]
                updated_fields.append("country")
            if destination.correspondent_contact_id != correspondent.id:
                destination.correspondent_contact = correspondent
                updated_fields.append("correspondent_contact")
            if not destination.is_active:
                destination.is_active = True
                updated_fields.append("is_active")
            if updated_fields:
                destination.save(update_fields=updated_fields)
            destinations_by_iata[destination.iata_code] = destination

        for scope_data in dataset.shipper_scopes:
            shipper = contacts_by_key[scope_data["shipper_key"]]
            assignment = _ensure_role_assignment(contact=shipper, role=OrganizationRole.SHIPPER)
            destination = destinations_by_iata[scope_data["destination_iata"]]
            ShipperScope.objects.get_or_create(
                role_assignment=assignment,
                destination=destination,
                defaults={
                    "all_destinations": False,
                    "is_active": True,
                    "valid_from": timezone.now(),
                },
            )

        for binding_data in dataset.recipient_bindings:
            shipper = contacts_by_key[binding_data["shipper_key"]]
            recipient = contacts_by_key[binding_data["recipient_key"]]
            destination = destinations_by_iata[binding_data["destination_iata"]]
            RecipientBinding.objects.get_or_create(
                shipper_org=shipper,
                recipient_org=recipient,
                destination=destination,
                is_active=True,
                defaults={"valid_from": timezone.now()},
            )
