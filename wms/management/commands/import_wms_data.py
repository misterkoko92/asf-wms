from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from contacts.models import Contact, ContactAddress, ContactTag, ContactType
from wms.models import Destination

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency at runtime
    load_workbook = None


TRUE_VALUES = {"true", "1", "yes", "y", "oui", "o", "vrai", "actif", "active"}

TAG_DONOR = "donateur"
TAG_SHIPPER = "expediteur"
TAG_RECIPIENT = "destinataire"
TAG_CORRESPONDENT = "correspondant"


def normalize_header(value: str) -> str:
    return value.strip().lower()


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
    return str(value).strip()


def normalize_text(value):
    return " ".join(clean_text(value).split())


def normalize_address(value):
    text = clean_text(value)
    if not text:
        return ""
    text = text.replace("\r", "\n")
    parts = [part.strip() for part in text.splitlines() if part.strip()]
    return ", ".join(parts)


def normalize_key(value):
    return normalize_text(value).lower()


def parse_bool(value):
    text = normalize_text(value).lower()
    if not text:
        return None
    return text in TRUE_VALUES


def extract_email(value):
    text = clean_text(value)
    if not text:
        return ""
    if "@" in text and " " not in text:
        return text
    return ""


def append_note(existing, note):
    note = normalize_text(note)
    if not note:
        return existing
    if existing:
        if note in existing:
            return existing
        return f"{existing.rstrip()}\n{note}"
    return note


def ensure_tag(name):
    tag_name = normalize_text(name).lower()
    tag, _ = ContactTag.objects.get_or_create(name=tag_name)
    return tag


def iter_excel_rows(path):
    if load_workbook is None:
        raise CommandError("openpyxl is required to import Excel files.")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration as exc:
        raise CommandError(f"Excel file is empty: {path}") from exc
    header_map = {}
    for index, header in enumerate(headers):
        key = normalize_header(str(header or ""))
        if not key:
            continue
        header_map.setdefault(key, []).append(index)
    for row in rows:
        if not any(clean_text(cell) for cell in row):
            continue
        yield row, header_map
    workbook.close()


def get_cell(row, header_map, name):
    key = normalize_header(name)
    for index in header_map.get(key, []):
        value = row[index]
        if clean_text(value):
            return value
    return None


def upsert_contact(name, tag_name, contact_type, email="", phone="", notes="", active=None):
    name = normalize_text(name)
    if not name:
        return None, False, False
    contact = Contact.objects.filter(
        name__iexact=name,
        contact_type=contact_type,
    ).first()
    created = False
    updated = False
    if not contact:
        contact = Contact(name=name, contact_type=contact_type, is_active=True)
        created = True
    if email and not contact.email:
        contact.email = email
        updated = True
    if phone and not contact.phone:
        contact.phone = phone
        updated = True
    if notes:
        merged = append_note(contact.notes, notes)
        if merged != contact.notes:
            contact.notes = merged
            updated = True
    if active is True and not contact.is_active:
        contact.is_active = True
        updated = True
    if created or updated:
        contact.save()
    tag = ensure_tag(tag_name)
    contact.tags.add(tag)
    return contact, created, updated


def upsert_address(contact, address_line1, city, postal_code, country):
    address_line1 = normalize_address(address_line1)
    city = normalize_text(city)
    postal_code = normalize_text(postal_code)
    country = normalize_text(country)
    if not address_line1:
        return False
    exists = ContactAddress.objects.filter(
        contact=contact,
        address_line1__iexact=address_line1,
        city__iexact=city,
        postal_code__iexact=postal_code,
        country__iexact=country,
    ).exists()
    if exists:
        return False
    is_default = not contact.addresses.exists()
    ContactAddress.objects.create(
        contact=contact,
        address_line1=address_line1,
        city=city,
        postal_code=postal_code,
        country=country or "France",
        is_default=is_default,
    )
    return True


def merge_association_tags():
    shipper_tag = ensure_tag(TAG_SHIPPER)
    for name in ("association", "nom association"):
        tag = ContactTag.objects.filter(name__iexact=name).first()
        if not tag:
            continue
        for contact in tag.contact_set.all():
            contact.tags.add(shipper_tag)
        tag.delete()


class Command(BaseCommand):
    help = "Import contacts and destinations from Excel files in a folder."

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            nargs="?",
            default="data",
            help="Folder containing Excel files (default: data)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate import without writing to the database",
        )

    def handle(self, *args, **options):
        base_path = Path(options["path"])
        if not base_path.exists():
            raise CommandError(f"Folder not found: {base_path}")

        files = {
            "donateur.xlsx": base_path / "donateur.xlsx",
            "expediteur.xlsx": base_path / "expediteur.xlsx",
            "destinataire.xlsx": base_path / "destinataire.xlsx",
            "correspondant.xlsx": base_path / "correspondant.xlsx",
            "destination.xlsx": base_path / "destination.xlsx",
        }

        stats = {
            "contacts_created": 0,
            "contacts_updated": 0,
            "addresses_created": 0,
            "destinations_created": 0,
            "destinations_updated": 0,
            "rows_skipped": 0,
            "warnings": 0,
        }

        with transaction.atomic():
            merge_association_tags()

            self._import_donors(files["donateur.xlsx"], stats)
            self._import_shippers(files["expediteur.xlsx"], stats)
            self._import_recipients(files["destinataire.xlsx"], stats)
            self._import_correspondents(files["correspondant.xlsx"], stats)
            self._import_destinations(files["destination.xlsx"], stats)

            if options["dry_run"]:
                transaction.set_rollback(True)
                self.stdout.write("Dry run complete; no changes were saved.")

        self.stdout.write(
            "Import termine: "
            f"{stats['contacts_created']} contacts crees, "
            f"{stats['contacts_updated']} contacts mis a jour, "
            f"{stats['addresses_created']} adresses crees, "
            f"{stats['destinations_created']} destinations crees, "
            f"{stats['destinations_updated']} destinations mises a jour, "
            f"{stats['rows_skipped']} lignes ignorees, "
            f"{stats['warnings']} avertissements."
        )

    def _import_donors(self, path, stats):
        if not path.exists():
            stats["warnings"] += 1
            self.stderr.write(f"Fichier manquant: {path}")
            return
        for row, header_map in iter_excel_rows(path):
            name = normalize_text(get_cell(row, header_map, "BE_DONATEUR"))
            if not name:
                stats["rows_skipped"] += 1
                continue
            contact, created, updated = upsert_contact(
                name=name,
                tag_name=TAG_DONOR,
                contact_type=ContactType.ORGANIZATION,
            )
            if created:
                stats["contacts_created"] += 1
            elif updated:
                stats["contacts_updated"] += 1

    def _import_shippers(self, path, stats):
        if not path.exists():
            stats["warnings"] += 1
            self.stderr.write(f"Fichier manquant: {path}")
            return
        for row, header_map in iter_excel_rows(path):
            name = normalize_text(get_cell(row, header_map, "ASSOCIATION_NOM"))
            if not name:
                stats["rows_skipped"] += 1
                continue
            active = parse_bool(get_cell(row, header_map, "ASSOCIATION_ACTIVE"))
            title = normalize_text(get_cell(row, header_map, "ASSOCIATION_PRESIDENT_TITRE"))
            first = normalize_text(get_cell(row, header_map, "ASSOCIATION_PRESIDENT_PRENOM"))
            last = normalize_text(get_cell(row, header_map, "ASSOCIATION_PRESIDENT_NOM"))
            president = " ".join(part for part in [title, first, last] if part)
            notes = f"President: {president}" if president else ""
            email = extract_email(
                get_cell(row, header_map, "ASSOCIATION_EMAIL")
                or get_cell(row, header_map, "ASSOCIATION_MAIL")
                or ""
            )
            phone = normalize_text(get_cell(row, header_map, "ASSOCIATION_TEL_1"))
            phone2 = normalize_text(get_cell(row, header_map, "ASSOCIATION_TEL_2"))
            if phone2:
                notes = append_note(notes, f"Tel 2: {phone2}")

            contact, created, updated = upsert_contact(
                name=name,
                tag_name=TAG_SHIPPER,
                contact_type=ContactType.ORGANIZATION,
                email=email,
                phone=phone,
                notes=notes,
                active=active,
            )
            if created:
                stats["contacts_created"] += 1
            elif updated:
                stats["contacts_updated"] += 1
            if contact:
                address_created = upsert_address(
                    contact=contact,
                    address_line1=get_cell(row, header_map, "ASSOCIATION_ADRESSE"),
                    city=get_cell(row, header_map, "ASSOCIATION_VILLE"),
                    postal_code=get_cell(row, header_map, "ASSOCIATION_CODE_POSTAL"),
                    country=get_cell(row, header_map, "ASSOCIATION_PAYS"),
                )
                if address_created:
                    stats["addresses_created"] += 1

    def _import_recipients(self, path, stats):
        if not path.exists():
            stats["warnings"] += 1
            self.stderr.write(f"Fichier manquant: {path}")
            return
        for row, header_map in iter_excel_rows(path):
            name = normalize_text(get_cell(row, header_map, "DESTINATAIRE_STRUCTURE"))
            if not name:
                stats["rows_skipped"] += 1
                continue
            status = normalize_text(get_cell(row, header_map, "DESTINATAIRE_STATUT"))
            notes = f"Statut: {status}" if status else ""
            phone = normalize_text(get_cell(row, header_map, "DESTINATAIRE_STRUCTURE_TEL_1"))
            phone2 = normalize_text(get_cell(row, header_map, "DESTINATAIRE_STRUCTURE_TEL_2"))
            if phone2:
                notes = append_note(notes, f"Tel 2: {phone2}")

            contact, created, updated = upsert_contact(
                name=name,
                tag_name=TAG_RECIPIENT,
                contact_type=ContactType.ORGANIZATION,
                phone=phone,
                notes=notes,
                active=True,
            )
            if created:
                stats["contacts_created"] += 1
            elif updated:
                stats["contacts_updated"] += 1
            if contact:
                address_created = upsert_address(
                    contact=contact,
                    address_line1=get_cell(row, header_map, "DESTINATAIRE_STRUCTURE_ADRESSE"),
                    city=get_cell(row, header_map, "DESTINATAIRE_STRUCTURE_VILLE"),
                    postal_code=get_cell(row, header_map, "DESTINATAIRE_STRUCTURE_CODE_POSTAL"),
                    country=get_cell(row, header_map, "DESTINATAIRE_STRUCTURE_PAYS"),
                )
                if address_created:
                    stats["addresses_created"] += 1

    def _import_correspondents(self, path, stats):
        if not path.exists():
            stats["warnings"] += 1
            self.stderr.write(f"Fichier manquant: {path}")
            return
        for row, header_map in iter_excel_rows(path):
            title = normalize_text(get_cell(row, header_map, "CORRESPONDANT_TITRE"))
            first = normalize_text(get_cell(row, header_map, "CORRESPONDANT_PRENOM"))
            last = normalize_text(get_cell(row, header_map, "CORRESPONDANT_NOM"))
            name = normalize_text(" ".join(part for part in [first, last] if part))
            if not name:
                stats["rows_skipped"] += 1
                continue
            notes = f"Titre: {title}" if title else ""
            phone = normalize_text(get_cell(row, header_map, "CORRESPONDANT_TEL_1"))
            phone2 = normalize_text(get_cell(row, header_map, "CORRESPONDANT_TEL_2"))
            if phone2:
                notes = append_note(notes, f"Tel 2: {phone2}")
            contact_type = ContactType.PERSON if first or last else ContactType.ORGANIZATION

            contact, created, updated = upsert_contact(
                name=name,
                tag_name=TAG_CORRESPONDENT,
                contact_type=contact_type,
                phone=phone,
                notes=notes,
                active=True,
            )
            if created:
                stats["contacts_created"] += 1
            elif updated:
                stats["contacts_updated"] += 1
            if contact:
                address_created = upsert_address(
                    contact=contact,
                    address_line1=get_cell(row, header_map, "CORRESPONDANT_ADRESSE"),
                    city=get_cell(row, header_map, "CORRESPONDANT_VILLE"),
                    postal_code=get_cell(row, header_map, "CORRESPONDANT_CODE_POSTAL"),
                    country=get_cell(row, header_map, "CORRESPONDANT_PAYS"),
                )
                if address_created:
                    stats["addresses_created"] += 1

    def _import_destinations(self, path, stats):
        if not path.exists():
            stats["warnings"] += 1
            self.stderr.write(f"Fichier manquant: {path}")
            return
        correspondent_map = self._build_correspondent_map(stats)
        for row, header_map in iter_excel_rows(path):
            city = normalize_text(get_cell(row, header_map, "DESTINATION_VILLE"))
            country = normalize_text(get_cell(row, header_map, "DESTINATION_PAYS"))
            iata = normalize_text(get_cell(row, header_map, "DESTINATION_IATA")).upper()
            if not city or not country or not iata:
                stats["rows_skipped"] += 1
                continue
            key = (normalize_key(city), normalize_key(country))
            correspondent = correspondent_map.get(key)
            if not correspondent:
                stats["warnings"] += 1
                self.stderr.write(
                    f"Destination sans correspondant: {city}, {country} ({iata})"
                )
                continue
            destination = Destination.objects.filter(iata_code__iexact=iata).first()
            defaults = {
                "city": city,
                "country": country,
                "correspondent_contact": correspondent,
                "is_active": True,
            }
            if destination:
                updated = False
                for field, value in defaults.items():
                    if getattr(destination, field) != value:
                        setattr(destination, field, value)
                        updated = True
                if updated:
                    destination.save()
                    stats["destinations_updated"] += 1
            else:
                Destination.objects.create(iata_code=iata, **defaults)
                stats["destinations_created"] += 1

    def _build_correspondent_map(self, stats):
        mapping = {}
        contacts = (
            Contact.objects.filter(tags__name__iexact=TAG_CORRESPONDENT)
            .prefetch_related("addresses")
            .distinct()
        )
        for contact in contacts:
            for address in contact.addresses.all():
                city = normalize_key(address.city)
                country = normalize_key(address.country)
                if not city or not country:
                    continue
                key = (city, country)
                if key in mapping and mapping[key].id != contact.id:
                    stats["warnings"] += 1
                    self.stderr.write(
                        f"Correspondants multiples pour {address.city}, {address.country}."
                    )
                    continue
                mapping[key] = contact
        return mapping
