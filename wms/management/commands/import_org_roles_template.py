from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import load_workbook

from contacts.models import Contact, ContactType
from wms.admin_organization_roles_review import _resolve_recipient_organization
from wms.models import (
    ContactSubscription,
    Destination,
    DestinationCorrespondentDefault,
    DestinationCorrespondentOverride,
    MigrationReviewItem,
    MigrationReviewItemStatus,
    NotificationChannel,
    OrganizationContact,
    OrganizationRole,
    OrganizationRoleAssignment,
    OrganizationRoleContact,
    RecipientBinding,
    RoleEventType,
    ShipperScope,
)

DEFAULT_INPUT_PATH = "docs/import/organization_roles_template_filled_ok.xlsx"

ROLE_VALUES = {choice[0] for choice in OrganizationRole.choices}
CHANNEL_VALUES = {choice[0] for choice in NotificationChannel.choices}
EVENT_BY_COLUMN = {
    "notify_shipment_status_updated": RoleEventType.SHIPMENT_STATUS_UPDATED,
    "notify_shipment_delivered": RoleEventType.SHIPMENT_DELIVERED,
    "notify_shipment_tracking_updated": RoleEventType.SHIPMENT_TRACKING_UPDATED,
    "notify_order_document_requested": RoleEventType.ORDER_DOCUMENT_REQUESTED,
}


@dataclass
class ImportStats:
    organizations_rows: int = 0
    organizations_created: int = 0
    organization_assignments_created: int = 0
    organization_assignments_updated: int = 0
    shipper_scopes_rows: int = 0
    shipper_scopes_created: int = 0
    shipper_scopes_updated: int = 0
    recipient_bindings_rows: int = 0
    recipient_bindings_created: int = 0
    recipient_bindings_updated: int = 0
    correspondents_rows: int = 0
    correspondent_defaults_created: int = 0
    correspondent_defaults_updated: int = 0
    correspondent_overrides_created: int = 0
    correspondent_overrides_updated: int = 0
    organization_contacts_rows: int = 0
    organization_contacts_created: int = 0
    organization_contacts_updated: int = 0
    organization_role_contacts_created: int = 0
    organization_role_contacts_updated: int = 0
    subscriptions_created: int = 0
    migration_review_rows: int = 0
    migration_review_resolved: int = 0
    migration_review_skipped: int = 0


def _project_root() -> Path:
    return Path(__file__).resolve().parents[3]


def _resolve_path(value: str) -> Path:
    path = Path(value).expanduser()
    if path.is_absolute():
        return path
    return (_project_root() / path).resolve()


def _normalize(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _bool_or_none(value):
    raw = _normalize(value).lower()
    if raw == "":
        return None
    if raw == "true":
        return True
    if raw == "false":
        return False
    raise ValueError(f"Valeur booleenne invalide: {value!r}")


def _parse_datetime_or_none(value):
    if value is None or value == "":
        return None
    if hasattr(value, "tzinfo"):
        if timezone.is_naive(value):
            return timezone.make_aware(value, timezone.get_current_timezone())
        return value

    raw = _normalize(value)
    if not raw:
        return None
    dt = parse_datetime(raw)
    if dt is not None:
        if timezone.is_naive(dt):
            return timezone.make_aware(dt, timezone.get_current_timezone())
        return dt
    d = parse_date(raw)
    if d is not None:
        return timezone.make_aware(
            datetime.combine(d, time.min),
            timezone.get_current_timezone(),
        )
    raise ValueError(f"Date/heure invalide: {value!r}")


def _sheet_rows(*, workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        raise CommandError(f"Onglet absent: {sheet_name}")
    sheet = workbook[sheet_name]
    headers = [_normalize(cell.value) for cell in sheet[1]]
    rows = []
    for row_number, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(_normalize(value) == "" for value in row_values):
            continue
        row = {
            headers[index]: _normalize(row_values[index]) if index < len(row_values) else ""
            for index in range(len(headers))
        }
        row["_row_number"] = str(row_number)
        rows.append(row)
    return rows


def _single_org_by_name(
    name: str,
    *,
    row_label: str,
    create_missing_organizations: bool = False,
    stats: ImportStats | None = None,
) -> Contact:
    query = Contact.objects.filter(
        contact_type=ContactType.ORGANIZATION,
        name__iexact=name.strip(),
    ).order_by("-is_active", "id")
    count = query.count()
    if count == 0:
        if create_missing_organizations:
            contact = Contact.objects.create(
                contact_type=ContactType.ORGANIZATION,
                name=name.strip(),
                is_active=True,
            )
            if stats is not None:
                stats.organizations_created += 1
            return contact
        raise CommandError(f"{row_label}: organisation introuvable: {name}")
    if count > 1:
        ids = list(query.values_list("id", flat=True))
        raise CommandError(f"{row_label}: organisation ambigue '{name}' (ids={ids})")
    matched_contact = query.first()
    if matched_contact is None:
        raise CommandError(f"{row_label}: organisation introuvable: {name}")
    return matched_contact


def _destination_by_iata(iata_code: str, *, row_label: str) -> Destination:
    query = Destination.objects.filter(iata_code__iexact=iata_code.strip()).order_by("id")
    count = query.count()
    if count == 0:
        raise CommandError(f"{row_label}: escale introuvable: {iata_code}")
    if count > 1:
        raise CommandError(f"{row_label}: escale ambigue: {iata_code}")
    destination = query.first()
    if destination is None:
        raise CommandError(f"{row_label}: escale introuvable: {iata_code}")
    return destination


def _ensure_role_assignment(*, organization, role, is_active, stats: ImportStats):
    assignment, created = OrganizationRoleAssignment.objects.get_or_create(
        organization=organization,
        role=role,
        defaults={"is_active": bool(is_active) if is_active is not None else False},
    )
    if created:
        stats.organization_assignments_created += 1
    if is_active is not None and assignment.is_active != is_active:
        assignment.is_active = is_active
        assignment.save(update_fields=["is_active", "updated_at"])
        stats.organization_assignments_updated += 1
    return assignment


def _apply_organizations(
    rows,
    stats: ImportStats,
    *,
    create_missing_organizations: bool,
):
    for row in rows:
        row_label = f"Organizations row {row['_row_number']}"
        stats.organizations_rows += 1
        organization_name = row.get("organization_name", "").strip()
        role = row.get("role", "").strip().lower()
        if not organization_name:
            raise CommandError(f"{row_label}: organization_name requis")
        if role not in ROLE_VALUES:
            raise CommandError(f"{row_label}: role invalide '{role}'")
        role_active = _bool_or_none(row.get("role_active"))
        organization = _single_org_by_name(
            organization_name,
            row_label=row_label,
            create_missing_organizations=create_missing_organizations,
            stats=stats,
        )
        _ensure_role_assignment(
            organization=organization,
            role=role,
            is_active=role_active,
            stats=stats,
        )


def _apply_shipper_scopes(
    rows,
    stats: ImportStats,
    *,
    create_missing_organizations: bool,
):
    for row in rows:
        row_label = f"ShipperScopes row {row['_row_number']}"
        stats.shipper_scopes_rows += 1
        organization_name = row.get("organization_name", "").strip()
        all_destinations = _bool_or_none(row.get("all_destinations"))
        is_active = _bool_or_none(row.get("is_active"))
        valid_from = _parse_datetime_or_none(row.get("valid_from")) or timezone.now()
        valid_to = _parse_datetime_or_none(row.get("valid_to"))
        iata_code = row.get("destination_iata", "").strip()

        if not organization_name:
            raise CommandError(f"{row_label}: organization_name requis")
        organization = _single_org_by_name(
            organization_name,
            row_label=row_label,
            create_missing_organizations=create_missing_organizations,
            stats=stats,
        )
        assignment = _ensure_role_assignment(
            organization=organization,
            role=OrganizationRole.SHIPPER,
            is_active=True,
            stats=stats,
        )

        if all_destinations is True:
            destination = None
            existing = (
                ShipperScope.objects.filter(
                    role_assignment=assignment,
                    all_destinations=True,
                )
                .order_by("id")
                .first()
            )
            if existing is None:
                scope = ShipperScope.objects.create(
                    role_assignment=assignment,
                    all_destinations=True,
                    destination=None,
                    is_active=is_active if is_active is not None else True,
                    valid_from=valid_from,
                    valid_to=valid_to,
                )
                stats.shipper_scopes_created += 1
            else:
                scope = existing
                scope.is_active = is_active if is_active is not None else scope.is_active
                scope.valid_from = valid_from
                scope.valid_to = valid_to
                scope.save(update_fields=["is_active", "valid_from", "valid_to", "updated_at"])
                stats.shipper_scopes_updated += 1
        else:
            if not iata_code:
                raise CommandError(
                    f"{row_label}: destination_iata requis quand all_destinations != true"
                )
            destination = _destination_by_iata(iata_code, row_label=row_label)
            scope, created = ShipperScope.objects.get_or_create(
                role_assignment=assignment,
                destination=destination,
                defaults={
                    "all_destinations": False,
                    "is_active": is_active if is_active is not None else True,
                    "valid_from": valid_from,
                    "valid_to": valid_to,
                },
            )
            if created:
                stats.shipper_scopes_created += 1
            else:
                scope.all_destinations = False
                scope.is_active = is_active if is_active is not None else scope.is_active
                scope.valid_from = valid_from
                scope.valid_to = valid_to
                scope.save(
                    update_fields=[
                        "all_destinations",
                        "is_active",
                        "valid_from",
                        "valid_to",
                        "updated_at",
                    ]
                )
                stats.shipper_scopes_updated += 1


def _apply_recipient_bindings(
    rows,
    stats: ImportStats,
    *,
    create_missing_organizations: bool,
):
    for row in rows:
        row_label = f"RecipientBindings row {row['_row_number']}"
        stats.recipient_bindings_rows += 1
        recipient_name = row.get("recipient_organization_name", "").strip()
        shipper_name = row.get("shipper_organization_name", "").strip()
        iata_code = row.get("destination_iata", "").strip()
        is_active = _bool_or_none(row.get("is_active"))
        valid_from = _parse_datetime_or_none(row.get("valid_from")) or timezone.now()
        valid_to = _parse_datetime_or_none(row.get("valid_to"))

        if not recipient_name or not shipper_name or not iata_code:
            raise CommandError(
                f"{row_label}: recipient_organization_name, shipper_organization_name "
                "et destination_iata sont requis"
            )

        recipient_org = _single_org_by_name(
            recipient_name,
            row_label=row_label,
            create_missing_organizations=create_missing_organizations,
            stats=stats,
        )
        shipper_org = _single_org_by_name(
            shipper_name,
            row_label=row_label,
            create_missing_organizations=create_missing_organizations,
            stats=stats,
        )
        destination = _destination_by_iata(iata_code, row_label=row_label)

        _ensure_role_assignment(
            organization=shipper_org,
            role=OrganizationRole.SHIPPER,
            is_active=True,
            stats=stats,
        )
        _ensure_role_assignment(
            organization=recipient_org,
            role=OrganizationRole.RECIPIENT,
            is_active=True,
            stats=stats,
        )

        desired_active = True if is_active is None else is_active
        base_queryset = RecipientBinding.objects.filter(
            shipper_org=shipper_org,
            recipient_org=recipient_org,
            destination=destination,
        )

        # Prefer updating an existing version with the same valid_from to keep
        # idempotence and avoid duplicating historical rows.
        exact_binding = base_queryset.filter(valid_from=valid_from).order_by("-id").first()

        if exact_binding is not None:
            if desired_active:
                base_queryset.filter(is_active=True).exclude(pk=exact_binding.pk).update(
                    is_active=False,
                    updated_at=timezone.now(),
                )
            updates = []
            if exact_binding.is_active != desired_active:
                exact_binding.is_active = desired_active
                updates.append("is_active")
            if exact_binding.valid_to != valid_to:
                exact_binding.valid_to = valid_to
                updates.append("valid_to")
            if updates:
                updates.append("updated_at")
                exact_binding.save(update_fields=updates)
                stats.recipient_bindings_updated += 1
            continue

        # If another active binding already exists for the same triplet, reuse
        # it instead of creating a new active version (single-active invariant).
        if desired_active:
            active_binding = (
                base_queryset.filter(is_active=True).order_by("-valid_from", "-id").first()
            )
            if active_binding is not None:
                if active_binding.valid_to != valid_to:
                    active_binding.valid_to = valid_to
                    active_binding.save(update_fields=["valid_to", "updated_at"])
                    stats.recipient_bindings_updated += 1
                continue

        RecipientBinding.objects.create(
            shipper_org=shipper_org,
            recipient_org=recipient_org,
            destination=destination,
            valid_from=valid_from,
            is_active=desired_active,
            valid_to=valid_to,
        )
        stats.recipient_bindings_created += 1


def _apply_correspondents(
    rows,
    stats: ImportStats,
    *,
    create_missing_organizations: bool,
):
    for row in rows:
        row_label = f"Correspondents row {row['_row_number']}"
        stats.correspondents_rows += 1
        correspondent_name = row.get("correspondent_organization_name", "").strip()
        iata_code = row.get("destination_iata", "").strip()
        scope_type = row.get("scope_type", "").strip() or "default"
        is_active = _bool_or_none(row.get("is_active"))
        shipper_name = row.get("shipper_organization_name", "").strip()
        recipient_name = row.get("recipient_organization_name", "").strip()

        if not correspondent_name or not iata_code:
            raise CommandError(
                f"{row_label}: correspondent_organization_name et destination_iata requis"
            )

        destination = _destination_by_iata(iata_code, row_label=row_label)
        correspondent_org = _single_org_by_name(
            correspondent_name,
            row_label=row_label,
            create_missing_organizations=create_missing_organizations,
            stats=stats,
        )
        _ensure_role_assignment(
            organization=correspondent_org,
            role=OrganizationRole.CORRESPONDENT,
            is_active=True,
            stats=stats,
        )

        if scope_type == "default":
            item, created = DestinationCorrespondentDefault.objects.get_or_create(
                destination=destination,
                correspondent_org=correspondent_org,
                defaults={"is_active": is_active if is_active is not None else True},
            )
            if created:
                stats.correspondent_defaults_created += 1
            else:
                item.is_active = is_active if is_active is not None else item.is_active
                item.save(update_fields=["is_active", "updated_at"])
                stats.correspondent_defaults_updated += 1
            continue

        if scope_type not in {
            "shipper_override",
            "recipient_override",
            "shipper_and_recipient_override",
        }:
            raise CommandError(f"{row_label}: scope_type invalide '{scope_type}'")

        shipper_org = None
        recipient_org = None
        if scope_type in {"shipper_override", "shipper_and_recipient_override"}:
            if not shipper_name:
                raise CommandError(f"{row_label}: shipper_organization_name requis")
            shipper_org = _single_org_by_name(
                shipper_name,
                row_label=row_label,
                create_missing_organizations=create_missing_organizations,
                stats=stats,
            )
        if scope_type in {"recipient_override", "shipper_and_recipient_override"}:
            if not recipient_name:
                raise CommandError(f"{row_label}: recipient_organization_name requis")
            recipient_org = _single_org_by_name(
                recipient_name,
                row_label=row_label,
                create_missing_organizations=create_missing_organizations,
                stats=stats,
            )

        override, created = DestinationCorrespondentOverride.objects.get_or_create(
            destination=destination,
            correspondent_org=correspondent_org,
            shipper_org=shipper_org,
            recipient_org=recipient_org,
            defaults={"is_active": is_active if is_active is not None else True},
        )
        if created:
            stats.correspondent_overrides_created += 1
        else:
            override.is_active = is_active if is_active is not None else override.is_active
            override.save(update_fields=["is_active", "updated_at"])
            stats.correspondent_overrides_updated += 1


def _find_or_create_organization_contact(*, organization, row, row_label, stats):
    email = row.get("contact_email", "").strip()
    first_name = row.get("contact_first_name", "").strip()
    last_name = row.get("contact_last_name", "").strip()
    phone = row.get("contact_phone", "").strip()
    is_active = _bool_or_none(row.get("is_active"))

    queryset = OrganizationContact.objects.filter(organization=organization)
    if email:
        queryset = queryset.filter(email__iexact=email)
    elif first_name or last_name:
        if first_name:
            queryset = queryset.filter(first_name__iexact=first_name)
        if last_name:
            queryset = queryset.filter(last_name__iexact=last_name)
    else:
        raise CommandError(
            f"{row_label}: contact_email ou (contact_first_name/contact_last_name) requis"
        )

    contact = queryset.order_by("id").first()
    created = False
    if contact is None:
        contact = OrganizationContact.objects.create(
            organization=organization,
            first_name=first_name,
            last_name=last_name,
            email=email,
            phone=phone,
            is_active=is_active if is_active is not None else True,
        )
        created = True
        stats.organization_contacts_created += 1
    else:
        updates = []
        if first_name and contact.first_name != first_name:
            contact.first_name = first_name
            updates.append("first_name")
        if last_name and contact.last_name != last_name:
            contact.last_name = last_name
            updates.append("last_name")
        if email and contact.email != email:
            contact.email = email
            updates.append("email")
        if phone and contact.phone != phone:
            contact.phone = phone
            updates.append("phone")
        if is_active is not None and contact.is_active != is_active:
            contact.is_active = is_active
            updates.append("is_active")
        if updates:
            updates.append("updated_at")
            contact.save(update_fields=updates)
            stats.organization_contacts_updated += 1
    return contact, created


def _apply_subscriptions(
    *,
    role_contact,
    row,
    row_label,
    stats,
    create_missing_organizations: bool,
):
    channel = row.get("notification_channel", "").strip().lower() or NotificationChannel.EMAIL
    if channel not in CHANNEL_VALUES:
        raise CommandError(f"{row_label}: notification_channel invalide '{channel}'")

    destination = None
    shipper_org = None
    recipient_org = None
    destination_iata = row.get("destination_iata_filter", "").strip()
    shipper_filter_name = row.get("shipper_organization_filter", "").strip()
    recipient_filter_name = row.get("recipient_organization_filter", "").strip()
    if destination_iata:
        destination = _destination_by_iata(destination_iata, row_label=row_label)
    if shipper_filter_name:
        shipper_org = _single_org_by_name(
            shipper_filter_name,
            row_label=row_label,
            create_missing_organizations=create_missing_organizations,
            stats=stats,
        )
    if recipient_filter_name:
        recipient_org = _single_org_by_name(
            recipient_filter_name,
            row_label=row_label,
            create_missing_organizations=create_missing_organizations,
            stats=stats,
        )

    role_contact.subscriptions.all().delete()
    for column_name, event_type in EVENT_BY_COLUMN.items():
        should_notify = _bool_or_none(row.get(column_name))
        if should_notify is not True:
            continue
        ContactSubscription.objects.create(
            role_contact=role_contact,
            event_type=event_type,
            channel=channel,
            destination=destination,
            shipper_org=shipper_org,
            recipient_org=recipient_org,
            is_active=True,
        )
        stats.subscriptions_created += 1


def _apply_organization_contacts(
    rows,
    stats: ImportStats,
    *,
    create_missing_organizations: bool,
):
    for row in rows:
        row_label = f"OrganizationContacts row {row['_row_number']}"
        stats.organization_contacts_rows += 1
        organization_name = row.get("organization_name", "").strip()
        role = row.get("role", "").strip().lower()
        if not organization_name or role not in ROLE_VALUES:
            raise CommandError(f"{row_label}: organization_name/role invalides")

        organization = _single_org_by_name(
            organization_name,
            row_label=row_label,
            create_missing_organizations=create_missing_organizations,
            stats=stats,
        )
        assignment = _ensure_role_assignment(
            organization=organization,
            role=role,
            is_active=True,
            stats=stats,
        )
        contact, _created = _find_or_create_organization_contact(
            organization=organization,
            row=row,
            row_label=row_label,
            stats=stats,
        )

        is_primary = _bool_or_none(row.get("is_primary"))
        is_active = _bool_or_none(row.get("is_active"))
        role_contact, created = OrganizationRoleContact.objects.get_or_create(
            role_assignment=assignment,
            contact=contact,
            defaults={
                "is_primary": bool(is_primary),
                "is_active": is_active if is_active is not None else True,
            },
        )
        if created:
            stats.organization_role_contacts_created += 1
        else:
            updates = []
            if is_primary is not None and role_contact.is_primary != is_primary:
                role_contact.is_primary = is_primary
                updates.append("is_primary")
            if is_active is not None and role_contact.is_active != is_active:
                role_contact.is_active = is_active
                updates.append("is_active")
            if updates:
                updates.append("updated_at")
                role_contact.save(update_fields=updates)
                stats.organization_role_contacts_updated += 1

        if is_primary is True:
            OrganizationRoleContact.objects.filter(
                role_assignment=assignment,
                is_primary=True,
            ).exclude(pk=role_contact.pk).update(is_primary=False, updated_at=timezone.now())

        _apply_subscriptions(
            role_contact=role_contact,
            row=row,
            row_label=row_label,
            stats=stats,
            create_missing_organizations=create_missing_organizations,
        )


def _resolve_review_item(*, review_item, note):
    review_item.status = MigrationReviewItemStatus.RESOLVED
    review_item.resolved_at = timezone.now()
    review_item.resolution_note = note or ""
    review_item.save(update_fields=["status", "resolved_at", "resolution_note", "updated_at"])


def _apply_migration_review(
    rows,
    stats: ImportStats,
    *,
    create_missing_organizations: bool,
):
    for row in rows:
        row_label = f"MigrationReview row {row['_row_number']}"
        stats.migration_review_rows += 1
        action = row.get("resolution_action", "").strip()
        if not action:
            stats.migration_review_skipped += 1
            continue
        if action not in {"resolve_binding", "resolve_without_binding"}:
            raise CommandError(f"{row_label}: resolution_action invalide '{action}'")

        raw_row_id = row.get("row_id", "").strip()
        if not raw_row_id.startswith("MR-"):
            raise CommandError(f"{row_label}: row_id attendu au format MR-<id>")
        try:
            review_item_id = int(raw_row_id.split("-", 1)[1])
        except ValueError as exc:
            raise CommandError(f"{row_label}: row_id invalide '{raw_row_id}'") from exc

        review_item = MigrationReviewItem.objects.filter(pk=review_item_id).first()
        if review_item is None:
            raise CommandError(f"{row_label}: item de revue introuvable id={review_item_id}")
        if review_item.status != MigrationReviewItemStatus.OPEN:
            stats.migration_review_skipped += 1
            continue

        note = row.get("resolution_note", "").strip()
        if action == "resolve_without_binding":
            _resolve_review_item(review_item=review_item, note=note)
            stats.migration_review_resolved += 1
            continue

        shipper_name = row.get("proposed_shipper_organization_name", "").strip()
        iata_code = row.get("proposed_destination_iata", "").strip()
        if not shipper_name or not iata_code:
            raise CommandError(
                f"{row_label}: shipper et destination proposes requis pour resolve_binding"
            )

        recipient_org = _resolve_recipient_organization(review_item)
        if recipient_org is None:
            raise CommandError(f"{row_label}: destinataire introuvable depuis review item")
        shipper_org = _single_org_by_name(
            shipper_name,
            row_label=row_label,
            create_missing_organizations=create_missing_organizations,
            stats=stats,
        )
        destination = _destination_by_iata(iata_code, row_label=row_label)

        shipper_assignment = _ensure_role_assignment(
            organization=shipper_org,
            role=OrganizationRole.SHIPPER,
            is_active=True,
            stats=stats,
        )
        _ensure_role_assignment(
            organization=recipient_org,
            role=OrganizationRole.RECIPIENT,
            is_active=True,
            stats=stats,
        )

        scope, created_scope = ShipperScope.objects.get_or_create(
            role_assignment=shipper_assignment,
            destination=destination,
            defaults={
                "all_destinations": False,
                "is_active": True,
                "valid_from": timezone.now(),
            },
        )
        if created_scope:
            stats.shipper_scopes_created += 1
        else:
            if not scope.is_active:
                scope.is_active = True
                scope.save(update_fields=["is_active", "updated_at"])
                stats.shipper_scopes_updated += 1

        existing_binding = (
            RecipientBinding.objects.filter(
                shipper_org=shipper_org,
                recipient_org=recipient_org,
                destination=destination,
                is_active=True,
            )
            .order_by("-valid_from", "-id")
            .first()
        )
        if existing_binding is None:
            RecipientBinding.objects.create(
                shipper_org=shipper_org,
                recipient_org=recipient_org,
                destination=destination,
                is_active=True,
                valid_from=timezone.now(),
            )
            stats.recipient_bindings_created += 1

        _resolve_review_item(review_item=review_item, note=note)
        stats.migration_review_resolved += 1


def _print_summary(command, stats: ImportStats, *, dry_run: bool):
    mode = "DRY RUN" if dry_run else "APPLY"
    command.stdout.write(command.style.MIGRATE_HEADING(f"Import org roles template [{mode}]"))
    lines = [
        f"- Organizations rows: {stats.organizations_rows}",
        f"- Organizations created on-the-fly: {stats.organizations_created}",
        (
            "- OrganizationRoleAssignment: "
            f"created={stats.organization_assignments_created}, "
            f"updated={stats.organization_assignments_updated}"
        ),
        (
            "- ShipperScope rows: "
            f"{stats.shipper_scopes_rows} "
            f"(created={stats.shipper_scopes_created}, updated={stats.shipper_scopes_updated})"
        ),
        (
            "- RecipientBinding rows: "
            f"{stats.recipient_bindings_rows} "
            f"(created={stats.recipient_bindings_created}, updated={stats.recipient_bindings_updated})"
        ),
        (
            "- Correspondents rows: "
            f"{stats.correspondents_rows} "
            f"(defaults created={stats.correspondent_defaults_created}, "
            f"defaults updated={stats.correspondent_defaults_updated}, "
            f"overrides created={stats.correspondent_overrides_created}, "
            f"overrides updated={stats.correspondent_overrides_updated})"
        ),
        (
            "- OrganizationContacts rows: "
            f"{stats.organization_contacts_rows} "
            f"(org contacts created={stats.organization_contacts_created}, "
            f"org contacts updated={stats.organization_contacts_updated}, "
            f"role links created={stats.organization_role_contacts_created}, "
            f"role links updated={stats.organization_role_contacts_updated}, "
            f"subscriptions created={stats.subscriptions_created})"
        ),
        (
            "- MigrationReview rows: "
            f"{stats.migration_review_rows} "
            f"(resolved={stats.migration_review_resolved}, "
            f"skipped={stats.migration_review_skipped})"
        ),
    ]
    command.stdout.write("\n".join(lines))


class Command(BaseCommand):
    help = (
        "Importe le template organization roles rempli et applique les mappings "
        "sur les modeles role-based."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--input",
            default=DEFAULT_INPUT_PATH,
            help=f"Chemin du fichier xlsx (defaut: {DEFAULT_INPUT_PATH})",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Valide et calcule le resultat sans persister en base.",
        )
        parser.add_argument(
            "--create-missing-organizations",
            action="store_true",
            help=(
                "Cree automatiquement les organisations manquantes "
                "(contacts type structure actifs)."
            ),
        )

    def handle(self, *args, **options):
        input_path = _resolve_path(options["input"])
        dry_run = bool(options.get("dry_run"))
        create_missing_organizations = bool(options.get("create_missing_organizations"))
        if not input_path.exists():
            raise CommandError(f"Fichier introuvable: {input_path}")

        workbook = load_workbook(input_path, data_only=True)
        stats = ImportStats()

        def _apply():
            _apply_organizations(
                _sheet_rows(workbook=workbook, sheet_name="Organizations"),
                stats,
                create_missing_organizations=create_missing_organizations,
            )
            _apply_shipper_scopes(
                _sheet_rows(workbook=workbook, sheet_name="ShipperScopes"),
                stats,
                create_missing_organizations=create_missing_organizations,
            )
            _apply_recipient_bindings(
                _sheet_rows(workbook=workbook, sheet_name="RecipientBindings"),
                stats,
                create_missing_organizations=create_missing_organizations,
            )
            _apply_correspondents(
                _sheet_rows(workbook=workbook, sheet_name="Correspondents"),
                stats,
                create_missing_organizations=create_missing_organizations,
            )
            _apply_organization_contacts(
                _sheet_rows(workbook=workbook, sheet_name="OrganizationContacts"),
                stats,
                create_missing_organizations=create_missing_organizations,
            )
            _apply_migration_review(
                _sheet_rows(workbook=workbook, sheet_name="MigrationReview"),
                stats,
                create_missing_organizations=create_missing_organizations,
            )

        if dry_run:
            with transaction.atomic():
                _apply()
                _print_summary(self, stats, dry_run=True)
                transaction.set_rollback(True)
            return

        with transaction.atomic():
            _apply()
        _print_summary(self, stats, dry_run=False)
