from __future__ import annotations

import json
import os
import sys
import tempfile
from datetime import timedelta
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from wms.planning.reference_case_builder import build_reference_case_payload


class Command(BaseCommand):
    help = "Replay a legacy planning session and export a WMS solver reference-case fixture."

    def _extract_week_bounds(self, planning_benevoles_path: Path) -> tuple[str, str] | None:
        workbook = load_workbook(planning_benevoles_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for cell_ref in ("C1", "C2", "C3"):
                value = sheet[cell_ref].value
                if hasattr(value, "date"):
                    monday = value.date()
                    return monday.isoformat(), (monday + timedelta(days=6)).isoformat()
        return None

    def add_arguments(self, parser):
        parser.add_argument("--case-name", required=True)
        parser.add_argument("--session-dir", required=True)
        parser.add_argument("--output", required=True)
        parser.add_argument(
            "--legacy-root",
            default="/Users/EdouardGonnu/asf_scheduler/new_repo",
        )
        parser.add_argument(
            "--timeout-seconds",
            type=int,
            default=20,
        )

    def handle(self, *args, **options):
        legacy_root = Path(options["legacy_root"]).expanduser().resolve()
        session_dir = Path(options["session_dir"]).expanduser().resolve()
        output_path = Path(options["output"]).expanduser().resolve()

        if not legacy_root.exists():
            raise CommandError(f"Legacy root not found: {legacy_root}")
        if not session_dir.exists():
            raise CommandError(f"Session directory not found: {session_dir}")

        required_files = {
            "TABLEAU_DE_BORD.xlsx": session_dir / "TABLEAU_DE_BORD.xlsx",
            "PLANNING_BENEVOLES.xlsx": session_dir / "PLANNING_BENEVOLES.xlsx",
            "VOLS.xlsx": session_dir / "VOLS.xlsx",
        }
        for label, path in required_files.items():
            if not path.exists():
                raise CommandError(f"Missing legacy source {label}: {path}")

        os.environ["ASF_TMP_DIR"] = os.environ.get(
            "ASF_TMP_DIR",
            str(Path(tempfile.gettempdir()) / "asf_wms_legacy_reference_case"),
        )

        sys.path.insert(0, str(legacy_root))
        try:
            from scheduler.data_sources import ExcelDataSource, ExcelSourcePaths
            from scheduler.solver_ortools_v3 import solve_planning_ortools_simulation
        except ImportError as exc:  # pragma: no cover - environment dependent
            raise CommandError(f"Unable to import legacy scheduler modules: {exc}") from exc

        source_paths = ExcelSourcePaths(
            tableau_de_bord=required_files["TABLEAU_DE_BORD.xlsx"],
            planning_benevoles=required_files["PLANNING_BENEVOLES.xlsx"],
            vols=required_files["VOLS.xlsx"],
        )
        data_source = ExcelDataSource(source_paths)
        df_param_be = data_source.load_param_be()
        df_param_dest = data_source.load_param_dest()
        df_param_benev = data_source.load_param_benev()
        df_be = data_source.load_shipments_df(df_param_be)
        df_vols = data_source.load_vols_df(df_param_dest)
        df_benev = data_source.load_benevoles_df(df_param_benev)

        result = solve_planning_ortools_simulation(
            timeout_seconds=options["timeout_seconds"],
            data_source=data_source,
        )
        planning_df = result.get("planning_df")
        if planning_df is None:
            raise CommandError("Legacy solver did not return a planning dataframe.")

        week_bounds = self._extract_week_bounds(required_files["PLANNING_BENEVOLES.xlsx"])
        payload = build_reference_case_payload(
            case_name=options["case_name"],
            df_be=df_be,
            df_param_be=df_param_be,
            df_vols=df_vols,
            df_benev=df_benev,
            df_param_benev=df_param_benev,
            planning_df=planning_df,
            stats=result.get("statistiques", {}) or {},
            week_start=week_bounds[0] if week_bounds else None,
            week_end=week_bounds[1] if week_bounds else None,
        )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

        self.stdout.write(
            self.style.SUCCESS(
                f"Built {options['case_name']} with "
                f"{len(payload['shipments'])} shipments, "
                f"{len(payload['volunteers'])} volunteers, "
                f"{len(payload['flights'])} flights, "
                f"{len(payload['expected_assignments'])} assignments -> {output_path}"
            )
        )
