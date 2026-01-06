import csv
import re
import unicodedata
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.core.files import File
from django.db import transaction

from wms.models import Location, Product, ProductCategory, ProductTag, RackColor, Warehouse
from wms.services import StockError, receive_stock

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency at runtime
    load_workbook = None

try:
    import xlrd
except ImportError:  # pragma: no cover - optional dependency at runtime
    xlrd = None


TRUE_VALUES = {"true", "1", "yes", "y", "oui", "o", "vrai"}
FALSE_VALUES = {"false", "0", "no", "n", "non", "faux"}


def normalize_header(value: str) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def get_value(row, *keys):
    for key in keys:
        if key in row:
            return row.get(key)
    return None


def parse_str(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_decimal(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise CommandError(f"Invalid decimal value: {value}") from exc


def parse_int(value):
    decimal_value = parse_decimal(value)
    if decimal_value is None:
        return None
    return int(decimal_value.to_integral_value(rounding=ROUND_HALF_UP))


def parse_bool(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if not text:
        return None
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    raise CommandError(f"Invalid boolean value: {value}")


def resolve_photo_path(raw_value, base_dir: Path):
    if raw_value is None:
        return None
    text = str(raw_value).strip()
    if not text:
        return None
    photo_path = Path(text).expanduser()
    if not photo_path.is_absolute():
        photo_path = base_dir / photo_path
    return photo_path


def attach_photo(product, photo_path: Path, dry_run: bool) -> bool:
    if not photo_path:
        return False
    if not photo_path.exists():
        raise CommandError(f"Photo not found: {photo_path}")
    if dry_run:
        return False
    with photo_path.open("rb") as handle:
        product.photo.save(photo_path.name, File(handle), save=False)
    return True


def apply_quantity(product, quantity, location, dry_run: bool, row_number: int):
    if quantity is None:
        return
    if quantity <= 0:
        raise CommandError(f"Row {row_number}: invalid quantity value.")
    stock_location = location or product.default_location
    if stock_location is None:
        raise CommandError(
            f"Row {row_number}: location required to import quantity."
        )
    if dry_run:
        return
    try:
        receive_stock(
            user=None,
            product=product,
            quantity=quantity,
            location=stock_location,
        )
    except StockError as exc:
        raise CommandError(f"Row {row_number}: {exc}") from exc


def build_category_path(parts):
    parent = None
    for name in parts:
        if not name:
            continue
        category, _ = ProductCategory.objects.get_or_create(name=name, parent=parent)
        parent = category
    return parent


def build_tags(raw_value):
    if not raw_value:
        return [], False
    tags = []
    for token in raw_value.split("|"):
        name = token.strip()
        if not name:
            continue
        tag, _ = ProductTag.objects.get_or_create(name=name)
        tags.append(tag)
    return tags, True


def get_or_create_location(warehouse_name, zone, aisle, shelf, row_number, stderr):
    if not any([warehouse_name, zone, aisle, shelf]):
        return None
    if not all([warehouse_name, zone, aisle, shelf]):
        stderr.write(
            "Row {}: incomplete location (warehouse/zone/aisle/shelf or "
            "entrepot/rack/etagere/bac).".format(row_number)
        )
        return None
    warehouse, _ = Warehouse.objects.get_or_create(name=warehouse_name)
    location, _ = Location.objects.get_or_create(
        warehouse=warehouse, zone=zone, aisle=aisle, shelf=shelf
    )
    return location


def compute_volume(length_cm, width_cm, height_cm):
    if length_cm is None or width_cm is None or height_cm is None:
        return None
    volume = length_cm * width_cm * height_cm
    return int(volume.to_integral_value(rounding=ROUND_HALF_UP))


def iter_csv_rows(path):
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        for row in reader:
            normalized = {normalize_header(k): v for k, v in row.items() if k}
            yield normalized


def iter_excel_rows(path):
    if path.suffix.lower() == ".xls":
        if xlrd is None:
            raise CommandError("xlrd is required to import .xls files.")
        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows == 0:
            raise CommandError("Excel file is empty.")
        headers = [normalize_header(cell.value) for cell in sheet.row(0)]
        for row_index in range(1, sheet.nrows):
            data = {}
            for col_index, header in enumerate(headers):
                if not header:
                    continue
                data[header] = sheet.cell_value(row_index, col_index)
            yield data
        return

    if load_workbook is None:
        raise CommandError("openpyxl is required to import Excel files.")
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration as exc:
        raise CommandError("Excel file is empty.") from exc
    normalized_headers = [normalize_header(str(h or "")) for h in headers]
    for row in rows:
        data = {}
        for header, value in zip(normalized_headers, row):
            if not header:
                continue
            data[header] = value
        yield data


class Command(BaseCommand):
    help = "Import products from CSV or Excel."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to CSV or XLSX file")
        parser.add_argument(
            "--update",
            action="store_true",
            help="Update existing products when SKU already exists",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate import without writing to the database",
        )
        parser.add_argument(
            "--skip-errors",
            action="store_true",
            help="Continue on row errors",
        )

    def handle(self, *args, **options):
        path = Path(options["path"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            rows = iter_excel_rows(path)
        else:
            rows = iter_csv_rows(path)

        created = 0
        updated = 0
        skipped = 0
        errors = 0
        base_dir = path.parent

        with transaction.atomic():
            for index, row in enumerate(rows, start=2):
                try:
                    sku = parse_str(get_value(row, "sku"))
                    name = parse_str(
                        get_value(row, "name", "nom", "nom_produit", "produit")
                    )
                    if not name:
                        raise CommandError("Missing required field: name")
                    if options["update"] and not sku:
                        raise CommandError("Missing required field for update: sku")

                    product = Product.objects.filter(sku=sku).first() if sku else None
                    if product and not options["update"]:
                        skipped += 1
                        continue

                    category_parts = [
                        parse_str(
                            get_value(
                                row,
                                "category_l1",
                                "categorie_l1",
                                "category_1",
                                "categorie_1",
                            )
                        ),
                        parse_str(
                            get_value(
                                row,
                                "category_l2",
                                "categorie_l2",
                                "category_2",
                                "categorie_2",
                            )
                        ),
                        parse_str(
                            get_value(
                                row,
                                "category_l3",
                                "categorie_l3",
                                "category_3",
                                "categorie_3",
                            )
                        ),
                        parse_str(
                            get_value(
                                row,
                                "category_l4",
                                "categorie_l4",
                                "category_4",
                                "categorie_4",
                            )
                        ),
                    ]
                    category = build_category_path([p for p in category_parts if p])
                    category_provided = any(category_parts)

                    tags, tags_provided = build_tags(
                        parse_str(get_value(row, "tags", "etiquettes", "etiquette"))
                    )

                    warehouse_name = parse_str(get_value(row, "warehouse", "entrepot"))
                    zone = parse_str(get_value(row, "zone", "rack"))
                    aisle = parse_str(get_value(row, "aisle", "etagere"))
                    shelf = parse_str(get_value(row, "shelf", "bac", "emplacement"))
                    default_location = get_or_create_location(
                        warehouse_name, zone, aisle, shelf, index, self.stderr
                    )
                    location_provided = default_location is not None

                    rack_color = parse_str(
                        get_value(row, "rack_color", "couleur_rack", "color_rack")
                    )
                    if rack_color and default_location is not None:
                        RackColor.objects.update_or_create(
                            warehouse=default_location.warehouse,
                            zone=default_location.zone,
                            defaults={"color": rack_color},
                        )

                    barcode = parse_str(
                        get_value(row, "barcode", "code_barre", "codebarre")
                    )
                    brand = parse_str(get_value(row, "brand", "marque"))
                    color = parse_str(get_value(row, "color", "couleur"))
                    photo_path = resolve_photo_path(
                        get_value(row, "photo", "image", "photo_path", "image_path"),
                        base_dir,
                    )
                    length_cm = parse_decimal(row.get("length_cm"))
                    width_cm = parse_decimal(row.get("width_cm"))
                    height_cm = parse_decimal(row.get("height_cm"))
                    weight_g = parse_int(row.get("weight_g"))
                    volume_cm3 = parse_int(row.get("volume_cm3"))
                    if volume_cm3 is None:
                        computed = compute_volume(length_cm, width_cm, height_cm)
                        volume_cm3 = computed

                    storage_conditions = parse_str(
                        get_value(row, "storage_conditions", "conditions_stockage")
                    )
                    perishable = parse_bool(get_value(row, "perishable", "perissable"))
                    quarantine_default = parse_bool(
                        get_value(row, "quarantine_default", "quarantaine_defaut")
                    )
                    notes = parse_str(get_value(row, "notes", "note"))
                    quantity = parse_int(
                        get_value(row, "quantity", "quantite", "stock", "qty")
                    )

                    if product is None:
                        product = Product(sku=sku or "", name=name)
                        if category is not None:
                            product.category = category
                        if barcode is not None:
                            product.barcode = barcode
                        if brand is not None:
                            product.brand = brand
                        if color is not None:
                            product.color = color
                        if location_provided:
                            product.default_location = default_location
                        if length_cm is not None:
                            product.length_cm = length_cm
                        if width_cm is not None:
                            product.width_cm = width_cm
                        if height_cm is not None:
                            product.height_cm = height_cm
                        if weight_g is not None:
                            product.weight_g = weight_g
                        if volume_cm3 is not None:
                            product.volume_cm3 = volume_cm3
                        if storage_conditions is not None:
                            product.storage_conditions = storage_conditions
                        if perishable is not None:
                            product.perishable = perishable
                        if quarantine_default is not None:
                            product.quarantine_default = quarantine_default
                        if notes is not None:
                            product.notes = notes
                        attach_photo(product, photo_path, options["dry_run"])
                        product.save()
                        if tags_provided:
                            product.tags.set(tags)
                        apply_quantity(
                            product,
                            quantity,
                            default_location if location_provided else None,
                            options["dry_run"],
                            index,
                        )
                        created += 1
                        continue

                    updates = {}
                    if name is not None:
                        updates["name"] = name
                    if category_provided:
                        updates["category"] = category
                    if barcode is not None:
                        updates["barcode"] = barcode
                    if brand is not None:
                        updates["brand"] = brand
                    if color is not None:
                        updates["color"] = color
                    if location_provided:
                        updates["default_location"] = default_location
                    if length_cm is not None:
                        updates["length_cm"] = length_cm
                    if width_cm is not None:
                        updates["width_cm"] = width_cm
                    if height_cm is not None:
                        updates["height_cm"] = height_cm
                    if weight_g is not None:
                        updates["weight_g"] = weight_g
                    if volume_cm3 is not None:
                        updates["volume_cm3"] = volume_cm3
                    if storage_conditions is not None:
                        updates["storage_conditions"] = storage_conditions
                    if perishable is not None:
                        updates["perishable"] = perishable
                    if quarantine_default is not None:
                        updates["quarantine_default"] = quarantine_default
                    if notes is not None:
                        updates["notes"] = notes

                    photo_updated = attach_photo(
                        product, photo_path, options["dry_run"]
                    )
                    if photo_updated:
                        updates["photo"] = product.photo

                    if updates:
                        for field, value in updates.items():
                            setattr(product, field, value)
                        product.save(update_fields=list(updates.keys()))
                    if tags_provided:
                        product.tags.set(tags)
                    apply_quantity(
                        product,
                        quantity,
                        default_location if location_provided else None,
                        options["dry_run"],
                        index,
                    )
                    updated += 1
                except CommandError as exc:
                    errors += 1
                    if options["skip_errors"]:
                        self.stderr.write(f"Row {index}: {exc}")
                        continue
                    raise

            if options["dry_run"]:
                transaction.set_rollback(True)

        summary = (
            f"Imported products: created={created}, updated={updated}, "
            f"skipped={skipped}, errors={errors}"
        )
        if options["dry_run"]:
            summary += " (dry-run)"
        self.stdout.write(summary)
