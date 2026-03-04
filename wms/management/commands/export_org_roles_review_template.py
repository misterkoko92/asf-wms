from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

from openpyxl import load_workbook

from wms import models
from wms.admin_organization_roles_review import _build_review_row

DEFAULT_TEMPLATE_PATH = "docs/import/organization_roles_template.xlsx"
DEFAULT_OUTPUT_PATH = "docs/import/organization_roles_template_filled.xlsx"
TARGET_SHEET_NAME = "MigrationReview"


def _project_root() -> Path:
    return Path(__file__).resolve().parents[3]


def _resolve_path(value: str) -> Path:
    path = Path(value).expanduser()
    if path.is_absolute():
        return path
    return (_project_root() / path).resolve()


def _open_review_items_queryset(include_resolved: bool):
    queryset = models.MigrationReviewItem.objects.select_related(
        "organization",
        "legacy_contact",
        "legacy_contact__organization",
    ).order_by("created_at", "id")
    if include_resolved:
        return queryset
    return queryset.filter(status=models.MigrationReviewItemStatus.OPEN)


def _resolve_suggested_name_by_id(*, options, suggested_id):
    if not suggested_id:
        return ""
    for option in options:
        if option.id == suggested_id:
            return option.name
    return ""


def _resolve_suggested_destination(*, destination_options, suggested_destination_id):
    if not suggested_destination_id:
        return "", ""
    for destination in destination_options:
        if destination.id == suggested_destination_id:
            return destination.iata_code or "", destination.city or ""
    return "", ""


def _build_export_rows(*, include_resolved: bool):
    exported_rows = []
    for review_item in _open_review_items_queryset(include_resolved=include_resolved):
        review_row = _build_review_row(review_item)
        recipient_org = review_row.get("recipient_org")
        suggested_shipper_name = _resolve_suggested_name_by_id(
            options=review_row.get("shipper_options") or [],
            suggested_id=review_row.get("suggested_shipper_id"),
        )
        proposed_destination_iata, proposed_destination_city = _resolve_suggested_destination(
            destination_options=review_row.get("destination_options") or [],
            suggested_destination_id=review_row.get("suggested_destination_id"),
        )
        default_action = (
            "resolve_binding"
            if suggested_shipper_name and proposed_destination_iata
            else ""
        )
        exported_rows.append(
            [
                f"MR-{review_item.id}",
                (recipient_org.name if recipient_org else ""),
                review_item.reason_code,
                suggested_shipper_name,
                proposed_destination_iata,
                proposed_destination_city,
                default_action,
                "",
            ]
        )
    return exported_rows


class Command(BaseCommand):
    help = (
        "Alimente le template XLSX de revue Organization Roles avec les "
        "items MigrationReviewItem."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--template",
            default=DEFAULT_TEMPLATE_PATH,
            help=f"Chemin du template source (defaut: {DEFAULT_TEMPLATE_PATH})",
        )
        parser.add_argument(
            "--output",
            default=DEFAULT_OUTPUT_PATH,
            help=f"Chemin du fichier de sortie (defaut: {DEFAULT_OUTPUT_PATH})",
        )
        parser.add_argument(
            "--include-resolved",
            action="store_true",
            help="Inclure aussi les items deja resolus.",
        )

    def handle(self, *args, **options):
        template_path = _resolve_path(options["template"])
        output_path = _resolve_path(options["output"])
        include_resolved = bool(options.get("include_resolved"))

        if not template_path.exists():
            raise CommandError(f"Template introuvable: {template_path}")

        rows = _build_export_rows(include_resolved=include_resolved)

        workbook = load_workbook(template_path)
        if TARGET_SHEET_NAME not in workbook.sheetnames:
            raise CommandError(
                f"Onglet '{TARGET_SHEET_NAME}' absent du template: {template_path}"
            )

        worksheet = workbook[TARGET_SHEET_NAME]
        if worksheet.max_row > 1:
            worksheet.delete_rows(2, worksheet.max_row - 1)

        for row in rows:
            worksheet.append(row)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

        self.stdout.write(
            self.style.SUCCESS(
                f"Template rempli: {output_path} ({len(rows)} ligne(s) exportee(s))."
            )
        )
