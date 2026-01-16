import codecs
import csv
import io
import re
import unicodedata
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency at runtime
    load_workbook = None

try:
    import xlrd
except ImportError:  # pragma: no cover - optional dependency at runtime
    xlrd = None

try:
    import pdfplumber
except ImportError:  # pragma: no cover - optional dependency at runtime
    pdfplumber = None


TRUE_VALUES = {"true", "1", "yes", "y", "oui", "o", "vrai"}
FALSE_VALUES = {"false", "0", "no", "n", "non", "faux"}


def normalize_header(value):
    text = str(value or "").strip().lower()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _guess_utf16_encoding(data):
    sample = data[:2000]
    if len(sample) < 4:
        return None
    even_zeros = sum(1 for idx in range(0, len(sample), 2) if sample[idx] == 0)
    odd_zeros = sum(1 for idx in range(1, len(sample), 2) if sample[idx] == 0)
    if even_zeros > odd_zeros * 2:
        return "utf-16-be"
    if odd_zeros > even_zeros * 2:
        return "utf-16-le"
    return None


def decode_text(data):
    if data is None:
        return ""
    if isinstance(data, str):
        return data
    if not isinstance(data, (bytes, bytearray)):
        return str(data)
    data = bytes(data)
    if not data:
        return ""

    for bom, encoding in (
        (codecs.BOM_UTF8, "utf-8-sig"),
        (codecs.BOM_UTF16_LE, "utf-16-le"),
        (codecs.BOM_UTF16_BE, "utf-16-be"),
        (codecs.BOM_UTF32_LE, "utf-32-le"),
        (codecs.BOM_UTF32_BE, "utf-32-be"),
    ):
        if data.startswith(bom):
            return data.decode(encoding)

    if b"\x00" in data[:2000]:
        guessed = _guess_utf16_encoding(data)
        if guessed:
            return data.decode(guessed)
        for encoding in ("utf-16", "utf-32"):
            try:
                return data.decode(encoding)
            except UnicodeDecodeError:
                continue

    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("latin-1")


def iter_csv_rows(data, delimiter=";"):
    text = decode_text(data)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    for row in reader:
        normalized = {normalize_header(k): v for k, v in row.items() if k}
        yield normalized


def iter_xlsx_rows(data):
    if load_workbook is None:
        raise ValueError("openpyxl is required to import Excel files.")
    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError("Excel file is empty.") from exc
    normalized_headers = [normalize_header(str(h or "")) for h in headers]
    for row in rows:
        entry = {}
        for header, value in zip(normalized_headers, row):
            if not header:
                continue
            entry[header] = value
        yield entry


def iter_xls_rows(data):
    if xlrd is None:
        raise ValueError("xlrd is required to import .xls files.")
    workbook = xlrd.open_workbook(file_contents=data)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("Excel file is empty.")
    headers = [normalize_header(cell.value) for cell in sheet.row(0)]
    for row_index in range(1, sheet.nrows):
        entry = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            entry[header] = sheet.cell_value(row_index, col_index)
        yield entry


def iter_import_rows(data, extension):
    if extension == ".csv":
        return iter_csv_rows(data)
    if extension in {".xlsx", ".xlsm"}:
        return iter_xlsx_rows(data)
    if extension == ".xls":
        return iter_xls_rows(data)
    raise ValueError("Format de fichier non supporte.")


def _sanitize_headers(headers, row_length=None):
    cleaned = [str(value or "").strip() for value in (headers or [])]
    if row_length is None:
        row_length = len(cleaned)
    if not cleaned or all(not value for value in cleaned):
        return [f"Colonne {idx + 1}" for idx in range(row_length)]
    while len(cleaned) < row_length:
        cleaned.append("")
    normalized = []
    for idx, value in enumerate(cleaned, start=1):
        normalized.append(value or f"Colonne {idx}")
    return normalized


def _coerce_cell(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
    return str(value).strip()


def _extract_csv_table(data):
    text = decode_text(data)
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        raise ValueError("CSV vide.")
    delimiter = ";"
    if "," in lines[0] and ";" not in lines[0]:
        delimiter = ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        headers = next(reader)
    except StopIteration as exc:
        raise ValueError("CSV vide.") from exc
    rows = [[_coerce_cell(cell) for cell in row] for row in reader]
    return _sanitize_headers(headers), rows


def _extract_xlsx_table(data):
    if load_workbook is None:
        raise ValueError("openpyxl est requis pour importer .xlsx/.xlsm.")
    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError("Excel vide.") from exc
    headers = _sanitize_headers(headers)
    data_rows = []
    for row in rows:
        data_rows.append([_coerce_cell(cell) for cell in row])
    workbook.close()
    return headers, data_rows


def _extract_xls_table(data):
    if xlrd is None:
        raise ValueError("xlrd est requis pour importer .xls.")
    workbook = xlrd.open_workbook(file_contents=data)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("Excel vide.")
    headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
    headers = _sanitize_headers(headers, row_length=sheet.ncols)
    rows = []
    for row_index in range(1, sheet.nrows):
        row = [sheet.cell_value(row_index, col) for col in range(sheet.ncols)]
        rows.append([_coerce_cell(cell) for cell in row])
    return headers, rows


def _extract_pdf_table(data):
    if pdfplumber is None:
        raise ValueError("pdfplumber est requis pour importer des PDF texte.")
    with pdfplumber.open(BytesIO(data)) as pdf:
        tables = []
        for page in pdf.pages:
            table = page.extract_table()
            if table and len(table) > 1:
                tables.append(table)
                continue
            text = page.extract_text() or ""
            if text:
                lines = [line for line in text.splitlines() if line.strip()]
                if len(lines) > 1:
                    split_rows = [re.split(r"\s{2,}", line.strip()) for line in lines]
                    tables.append(split_rows)
        if not tables:
            raise ValueError("PDF scanne non supporte (aucun texte detecte).")
    headers = None
    rows = []
    for table in tables:
        if not table:
            continue
        if headers is None:
            headers = table[0]
            rows.extend(table[1:])
        else:
            rows.extend(table[1:] if table[0] == headers else table)
    if headers is None:
        raise ValueError("Impossible d'extraire un tableau du PDF.")
    max_len = max(len(headers), *(len(row) for row in rows)) if rows else len(headers)
    headers = _sanitize_headers(headers, row_length=max_len)
    normalized_rows = []
    for row in rows:
        padded = list(row) + [""] * (max_len - len(row))
        normalized_rows.append([_coerce_cell(cell) for cell in padded])
    return headers, normalized_rows


def extract_tabular_data(data, extension):
    if extension == ".csv":
        return _extract_csv_table(data)
    if extension in {".xlsx", ".xlsm"}:
        return _extract_xlsx_table(data)
    if extension == ".xls":
        return _extract_xls_table(data)
    if extension == ".pdf":
        return _extract_pdf_table(data)
    raise ValueError("Format de fichier non supporte.")


def get_value(row, *keys):
    for key in keys:
        if key in row:
            return row.get(key)
    return None


def parse_str(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_decimal(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid decimal value: {value}") from exc


def parse_int(value):
    decimal_value = parse_decimal(value)
    if decimal_value is None:
        return None
    return int(decimal_value.to_integral_value(rounding=ROUND_HALF_UP))


def parse_bool(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if not text:
        return None
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    raise ValueError(f"Invalid boolean value: {value}")


def parse_tokens(value):
    if not value:
        return []
    tokens = []
    for token in re.split(r"[|,]", str(value)):
        name = token.strip()
        if name:
            tokens.append(name)
    return tokens
