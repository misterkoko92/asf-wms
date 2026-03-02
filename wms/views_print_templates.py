import json
from io import BytesIO
from os.path import basename
from pathlib import Path

from django.conf import settings
from django.contrib import messages
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Count
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from .models import (
    PrintCellMapping,
    PrintPackDocument,
    PrintPackDocumentVersion,
    Product,
    Shipment,
)
from .print_context import (
    build_label_context,
    build_preview_context,
    build_product_label_context,
    build_sample_label_context,
)
from .print_layouts import DOCUMENT_TEMPLATES
from .print_pack_mapping_catalog import ALLOWED_SOURCE_KEYS, is_allowed_source_key
from .print_pack_template_versions import (
    restore_print_pack_document_version,
    save_print_pack_document_snapshot,
)
from .print_pack_workbook import build_column_choices, normalize_cell_ref, worksheet_row_choices
from .print_renderer import render_layout_from_layout
from .print_utils import build_label_pages, extract_block_style
from .view_permissions import (
    require_superuser as _require_superuser,
    scan_staff_required,
)

TEMPLATE_PRINT_TEMPLATE_LIST = "scan/print_template_list.html"
TEMPLATE_PRINT_TEMPLATE_EDIT = "scan/print_template_edit.html"
TEMPLATE_DYNAMIC_LABELS = "print/dynamic_labels.html"
TEMPLATE_DYNAMIC_DOCUMENT = "print/dynamic_document.html"
TEMPLATE_PRODUCT_LABELS = "print/product_labels.html"
TEMPLATE_PRODUCT_QR_LABELS = "print/product_qr_labels.html"

ACTIVE_PRINT_TEMPLATES = "print_templates"
SHELL_CLASS_WIDE = "scan-shell-wide"

DOC_LABEL_MAP = dict(DOCUMENT_TEMPLATES)
VALID_TRANSFORMS = ("", "upper", "date_fr")


def _redirect_template_edit(doc_type):
    return redirect("scan:scan_print_template_edit", doc_type=doc_type)


def _resolve_pack_document(doc_type):
    token = (doc_type or "").strip()
    if not token:
        return None
    queryset = PrintPackDocument.objects.select_related("pack")
    if token.isdigit():
        return queryset.filter(pk=int(token)).first()
    return queryset.filter(doc_type=token).order_by("pack__code", "sequence", "id").first()


def _resolve_template_search_dirs():
    configured_dirs = getattr(settings, "PRINT_PACK_TEMPLATE_DIRS", None)
    if configured_dirs is None:
        configured_dirs = [
            Path(settings.BASE_DIR) / "print_templates",
            Path(settings.BASE_DIR) / "data" / "print_templates",
        ]
    if isinstance(configured_dirs, (str, Path)):
        configured_dirs = [configured_dirs]
    directories = []
    for entry in configured_dirs:
        value = str(entry or "").strip()
        if not value:
            continue
        directories.append(Path(value))
    return directories


def _build_pack_template_filename(pack_document):
    variant = (pack_document.variant or "").strip() or "default"
    return f"{pack_document.pack.code}__{pack_document.doc_type}__{variant}.xlsx"


def _resolve_filesystem_template_path(pack_document):
    filename = _build_pack_template_filename(pack_document)
    for directory in _resolve_template_search_dirs():
        candidate = directory / filename
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def _resolve_effective_template_label(pack_document):
    if pack_document.xlsx_template_file:
        return basename(pack_document.xlsx_template_file.name or ""), True
    filesystem_path = _resolve_filesystem_template_path(pack_document)
    if filesystem_path is None:
        return "", False
    return filesystem_path.name, False


def _read_effective_template_bytes(pack_document):
    if pack_document.xlsx_template_file:
        with pack_document.xlsx_template_file.open("rb") as stream:
            return stream.read(), True
    filesystem_path = _resolve_filesystem_template_path(pack_document)
    if filesystem_path is None:
        return None, False
    with filesystem_path.open("rb") as stream:
        return stream.read(), False


def _build_template_list_items():
    queryset = (
        PrintPackDocument.objects.select_related("pack")
        .annotate(mapping_count=Count("cell_mappings"))
        .prefetch_related("versions__created_by")
        .order_by("pack__code", "sequence", "id")
    )
    items = []
    for pack_document in queryset:
        versions = sorted(
            list(pack_document.versions.all()),
            key=lambda version: version.version,
            reverse=True,
        )
        latest_version = versions[0] if versions else None
        template_filename, template_is_uploaded = _resolve_effective_template_label(
            pack_document
        )
        items.append(
            {
                "route_key": str(pack_document.id),
                "pack_code": pack_document.pack.code,
                "pack_name": pack_document.pack.name,
                "doc_type": pack_document.doc_type,
                "variant": pack_document.variant,
                "sequence": pack_document.sequence,
                "mapping_count": int(getattr(pack_document, "mapping_count", 0) or 0),
                "has_template_file": bool(template_filename),
                "template_filename": template_filename,
                "template_is_uploaded": template_is_uploaded,
                "active_version": latest_version.version if latest_version else None,
                "updated_at": latest_version.created_at if latest_version else None,
                "updated_by": latest_version.created_by if latest_version else None,
            }
        )
    return items


def _split_cell_ref(cell_ref):
    try:
        column, row = coordinate_from_string((cell_ref or "").strip().upper())
    except ValueError:
        return "", ""
    return column, str(row)


def _empty_mapping_row(*, worksheet_name="", sequence=1):
    return {
        "worksheet_name": worksheet_name,
        "column": "",
        "row": "",
        "cell_ref": "",
        "source_key": "",
        "transform": "",
        "required": False,
        "sequence": sequence,
        "merged_range": "",
    }


def _load_workbook_or_none(template_bytes):
    if template_bytes is None:
        return None
    try:
        return load_workbook(BytesIO(template_bytes))
    except Exception:
        return None


def _build_workbook_meta(workbook):
    if workbook is None:
        return {
            "worksheet_names": [],
            "columns_by_worksheet": {},
            "rows_by_worksheet": {},
        }
    all_columns = build_column_choices()
    worksheet_names = list(workbook.sheetnames)
    columns_by_worksheet = {}
    rows_by_worksheet = {}
    for worksheet_name in worksheet_names:
        worksheet = workbook[worksheet_name]
        max_column = max(1, int(getattr(worksheet, "max_column", 1) or 1))
        column_count = min(len(all_columns), max_column + 10)
        columns_by_worksheet[worksheet_name] = all_columns[:column_count]
        rows_by_worksheet[worksheet_name] = [
            str(row_number) for row_number in worksheet_row_choices(worksheet)
        ]
    return {
        "worksheet_names": worksheet_names,
        "columns_by_worksheet": columns_by_worksheet,
        "rows_by_worksheet": rows_by_worksheet,
    }


def _build_mapping_rows(pack_document, workbook):
    worksheet_names = list(workbook.sheetnames) if workbook is not None else []
    default_worksheet = worksheet_names[0] if worksheet_names else ""
    rows = []
    for mapping in pack_document.cell_mappings.order_by("sequence", "id"):
        column, row = _split_cell_ref(mapping.cell_ref)
        merged_range = ""
        if workbook is not None and mapping.worksheet_name in workbook.sheetnames:
            _, merged_range = normalize_cell_ref(workbook[mapping.worksheet_name], mapping.cell_ref)
        rows.append(
            {
                "worksheet_name": mapping.worksheet_name or default_worksheet,
                "column": column,
                "row": row,
                "cell_ref": mapping.cell_ref,
                "source_key": mapping.source_key,
                "transform": mapping.transform or "",
                "required": bool(mapping.required),
                "sequence": mapping.sequence,
                "merged_range": merged_range,
            }
        )
    if not rows:
        rows.append(_empty_mapping_row(worksheet_name=default_worksheet, sequence=1))
    return rows


def _build_edit_context(pack_document):
    template_filename, template_is_uploaded = _resolve_effective_template_label(pack_document)
    template_bytes, _template_from_db = _read_effective_template_bytes(pack_document)
    workbook = _load_workbook_or_none(template_bytes)
    workbook_meta = _build_workbook_meta(workbook)
    try:
        mapping_rows = _build_mapping_rows(pack_document, workbook)
    finally:
        if workbook is not None:
            workbook.close()
    versions = list(
        pack_document.versions.select_related("created_by").order_by("-version")
    )
    return {
        "active": ACTIVE_PRINT_TEMPLATES,
        "shell_class": SHELL_CLASS_WIDE,
        "pack_document": pack_document,
        "source_keys": ALLOWED_SOURCE_KEYS,
        "transform_choices": VALID_TRANSFORMS,
        "mapping_rows": mapping_rows,
        "worksheet_names": workbook_meta["worksheet_names"],
        "columns_by_worksheet": workbook_meta["columns_by_worksheet"],
        "rows_by_worksheet": workbook_meta["rows_by_worksheet"],
        "template_filename": template_filename,
        "template_is_uploaded": template_is_uploaded,
        "versions": versions,
    }


def _parse_preview_layout(request):
    layout_json = request.POST.get("layout_json") or ""
    try:
        return (json.loads(layout_json) if layout_json else {"blocks": []}), None
    except json.JSONDecodeError:
        return None, HttpResponse(status=400)


def _load_preview_shipment(raw_shipment_id):
    if not raw_shipment_id.isdigit():
        return None
    return (
        Shipment.objects.filter(archived_at__isnull=True)
        .select_related("destination")
        .prefetch_related("carton_set")
        .filter(pk=int(raw_shipment_id))
        .first()
    )


def _load_preview_product(raw_product_id):
    if not raw_product_id.isdigit():
        return None
    return (
        Product.objects.select_related("default_location", "default_location__warehouse")
        .filter(pk=int(raw_product_id))
        .first()
    )


def _render_shipment_label_preview(request, *, layout_data, shipment):
    labels = []
    if shipment:
        cartons = list(shipment.carton_set.order_by("code")[:6])
        total = shipment.carton_set.count() or 1
        if cartons:
            for index, _carton in enumerate(cartons, start=1):
                label_context = build_label_context(shipment, position=index, total=total)
                blocks = render_layout_from_layout(layout_data, label_context)
                labels.append({"blocks": blocks})
        else:
            label_context = build_sample_label_context()
            blocks = render_layout_from_layout(layout_data, label_context)
            labels.append({"blocks": blocks})
    else:
        label_context = build_sample_label_context()
        blocks = render_layout_from_layout(layout_data, label_context)
        labels.append({"blocks": blocks})
    return render(request, TEMPLATE_DYNAMIC_LABELS, {"labels": labels})


def _render_product_label_preview(request, *, layout_data, product):
    if product:
        base_context = build_product_label_context(product)
    else:
        base_context = build_preview_context("product_label")
    contexts = [dict(base_context) for _ in range(4)]
    pages, page_style = build_label_pages(
        layout_data,
        contexts,
        block_type="product_label",
        labels_per_page=4,
    )
    return render(
        request,
        TEMPLATE_PRODUCT_LABELS,
        {"pages": pages, "page_style": page_style},
    )


def _resolve_product_qr_grid(layout_data):
    page_style = extract_block_style(layout_data, "product_qr_label")
    try:
        rows = int(page_style.get("page_rows") or 5)
        cols = int(page_style.get("page_columns") or 3)
    except (TypeError, ValueError):
        rows, cols = 5, 3
    return max(1, rows * cols)


def _render_product_qr_preview(request, *, layout_data, product):
    if product:
        if not product.qr_code_image:
            product.generate_qr_code()
            product.save(update_fields=["qr_code_image"])
        base_context = build_preview_context("product_qr", product=product)
    else:
        base_context = build_preview_context("product_qr")
    labels_per_page = _resolve_product_qr_grid(layout_data)
    contexts = [dict(base_context) for _ in range(labels_per_page)]
    pages, page_style = build_label_pages(
        layout_data,
        contexts,
        block_type="product_qr_label",
        labels_per_page=labels_per_page,
    )
    return render(
        request,
        TEMPLATE_PRODUCT_QR_LABELS,
        {"pages": pages, "page_style": page_style},
    )


def _get_post_value(values, index, default=""):
    if index < 0 or index >= len(values):
        return default
    return values[index]


def _collect_mapping_rows(request):
    worksheets = request.POST.getlist("mapping_worksheet")
    columns = request.POST.getlist("mapping_column")
    rows = request.POST.getlist("mapping_row")
    source_keys = request.POST.getlist("mapping_source_key")
    transforms = request.POST.getlist("mapping_transform")
    required_flags = request.POST.getlist("mapping_required")
    sequences = request.POST.getlist("mapping_sequence")
    row_count = max(
        len(worksheets),
        len(columns),
        len(rows),
        len(source_keys),
        len(transforms),
        len(required_flags),
        len(sequences),
    )
    parsed_rows = []
    errors = []
    for index in range(row_count):
        worksheet_name = _get_post_value(worksheets, index, "").strip()
        column = _get_post_value(columns, index, "").strip().upper()
        row_value = _get_post_value(rows, index, "").strip()
        source_key = _get_post_value(source_keys, index, "").strip()
        transform = _get_post_value(transforms, index, "").strip()
        required_raw = _get_post_value(required_flags, index, "0").strip().lower()
        sequence_raw = _get_post_value(sequences, index, "").strip()
        if not any(
            (
                worksheet_name,
                column,
                row_value,
                source_key,
                transform,
                sequence_raw,
            )
        ):
            continue
        if not (worksheet_name and column and row_value and source_key):
            errors.append(
                f"Ligne {index + 1}: worksheet, colonne, ligne et champ source sont requis."
            )
            continue
        if sequence_raw.isdigit() and int(sequence_raw) > 0:
            sequence = int(sequence_raw)
        else:
            sequence = index + 1
        parsed_rows.append(
            {
                "worksheet_name": worksheet_name,
                "column": column,
                "row": row_value,
                "source_key": source_key,
                "transform": transform,
                "required": required_raw in {"1", "true", "yes", "on"},
                "sequence": sequence,
            }
        )
    return parsed_rows, errors


def _validate_and_normalize_rows(parsed_rows, workbook):
    normalized_rows = []
    errors = []
    if workbook is None:
        return [], ["Template XLSX introuvable ou invalide."]
    worksheet_map = {worksheet_name: workbook[worksheet_name] for worksheet_name in workbook.sheetnames}
    seen_cells = set()
    for index, row_data in enumerate(parsed_rows, start=1):
        worksheet_name = row_data["worksheet_name"]
        if worksheet_name not in worksheet_map:
            errors.append(f"Ligne {index}: worksheet inconnu ({worksheet_name}).")
            continue
        worksheet = worksheet_map[worksheet_name]
        try:
            row_number = int(row_data["row"])
        except (TypeError, ValueError):
            errors.append(f"Ligne {index}: numéro de ligne invalide ({row_data['row']}).")
            continue
        if row_number <= 0 or row_number > max(1, int(worksheet.max_row or 1)):
            errors.append(
                f"Ligne {index}: numéro de ligne hors limites pour la feuille {worksheet_name}."
            )
            continue
        column = row_data["column"].upper()
        try:
            column_index_from_string(column)
        except ValueError:
            errors.append(f"Ligne {index}: colonne invalide ({column}).")
            continue
        source_key = row_data["source_key"]
        if not is_allowed_source_key(source_key):
            errors.append(f"Ligne {index}: champ source invalide ({source_key}).")
            continue
        transform = row_data["transform"]
        if transform not in VALID_TRANSFORMS:
            errors.append(f"Ligne {index}: transformation invalide ({transform}).")
            continue
        normalized_cell_ref, _merged_range = normalize_cell_ref(
            worksheet,
            f"{column}{row_number}",
        )
        dedupe_key = (worksheet_name, normalized_cell_ref)
        if dedupe_key in seen_cells:
            errors.append(
                f"Ligne {index}: cellule dupliquée ({worksheet_name}!{normalized_cell_ref})."
            )
            continue
        seen_cells.add(dedupe_key)
        normalized_rows.append(
            {
                "worksheet_name": worksheet_name,
                "cell_ref": normalized_cell_ref,
                "source_key": source_key,
                "transform": transform,
                "required": bool(row_data["required"]),
                "sequence": int(row_data["sequence"]),
            }
        )
    return normalized_rows, errors


def _persist_pack_document_save(
    *,
    pack_document,
    normalized_rows,
    change_note,
    uploaded_filename,
    uploaded_bytes,
    fallback_template_bytes,
    user,
):
    with transaction.atomic():
        should_save_template_file = uploaded_bytes is not None or (
            not pack_document.xlsx_template_file and fallback_template_bytes is not None
        )
        if uploaded_bytes is not None:
            template_filename = basename(uploaded_filename or "") or _build_pack_template_filename(
                pack_document
            )
            pack_document.xlsx_template_file.save(
                template_filename,
                ContentFile(uploaded_bytes),
                save=False,
            )
        elif not pack_document.xlsx_template_file and fallback_template_bytes is not None:
            pack_document.xlsx_template_file.save(
                _build_pack_template_filename(pack_document),
                ContentFile(fallback_template_bytes),
                save=False,
            )
        if should_save_template_file:
            pack_document.save(update_fields=["xlsx_template_file"])

        pack_document.cell_mappings.all().delete()
        create_buffer = []
        for row_data in normalized_rows:
            create_buffer.append(
                PrintCellMapping(
                    pack_document=pack_document,
                    worksheet_name=row_data["worksheet_name"],
                    cell_ref=row_data["cell_ref"],
                    source_key=row_data["source_key"],
                    transform=row_data["transform"],
                    required=row_data["required"],
                    sequence=row_data["sequence"],
                )
            )
        if create_buffer:
            PrintCellMapping.objects.bulk_create(create_buffer)

        save_print_pack_document_snapshot(
            pack_document=pack_document,
            created_by=user,
            change_type="save",
            change_note=change_note,
        )


def _handle_pack_document_save(request, pack_document):
    uploaded_file = request.FILES.get("xlsx_template_file")
    uploaded_filename = ""
    uploaded_bytes = None
    fallback_template_bytes = None
    workbook = None
    if uploaded_file is not None:
        uploaded_filename = basename(uploaded_file.name or "")
        if not uploaded_filename.lower().endswith(".xlsx"):
            messages.error(request, "Le fichier doit être au format .xlsx.")
            return
        uploaded_bytes = uploaded_file.read()
        workbook = _load_workbook_or_none(uploaded_bytes)
        if workbook is None:
            messages.error(request, "Le fichier XLSX uploadé est invalide.")
            return
    else:
        fallback_template_bytes, _from_db = _read_effective_template_bytes(pack_document)
        workbook = _load_workbook_or_none(fallback_template_bytes)
        if workbook is None:
            messages.error(
                request,
                "Aucun template XLSX disponible. Uploadez un fichier avant de mapper.",
            )
            return
    try:
        parsed_rows, parse_errors = _collect_mapping_rows(request)
        if parse_errors:
            for error in parse_errors:
                messages.error(request, error)
            return
        normalized_rows, validation_errors = _validate_and_normalize_rows(parsed_rows, workbook)
        if validation_errors:
            for error in validation_errors:
                messages.error(request, error)
            return
        _persist_pack_document_save(
            pack_document=pack_document,
            normalized_rows=normalized_rows,
            change_note=(request.POST.get("change_note") or "").strip(),
            uploaded_filename=uploaded_filename,
            uploaded_bytes=uploaded_bytes,
            fallback_template_bytes=fallback_template_bytes,
            user=request.user,
        )
    finally:
        workbook.close()
    messages.success(request, "Template XLSX et mappings enregistrés.")


def _handle_pack_document_restore(request, pack_document):
    version_id = (request.POST.get("version_id") or "").strip()
    if not version_id.isdigit():
        messages.error(request, "Version requise.")
        return
    version = get_object_or_404(
        PrintPackDocumentVersion,
        pk=int(version_id),
        pack_document=pack_document,
    )
    restore_print_pack_document_version(
        version=version,
        created_by=request.user,
        change_note=(request.POST.get("change_note") or "").strip(),
    )
    messages.success(request, f"Version v{version.version} restaurée.")


@scan_staff_required
@require_http_methods(["GET"])
def scan_print_templates(request):
    _require_superuser(request)
    return render(
        request,
        TEMPLATE_PRINT_TEMPLATE_LIST,
        {
            "active": ACTIVE_PRINT_TEMPLATES,
            "shell_class": SHELL_CLASS_WIDE,
            "templates": _build_template_list_items(),
        },
    )


@scan_staff_required
@require_http_methods(["GET", "POST"])
def scan_print_template_edit(request, doc_type):
    _require_superuser(request)
    pack_document = _resolve_pack_document(doc_type)
    if pack_document is None:
        raise Http404("Template not found")
    if request.method == "POST":
        action = (request.POST.get("action") or "save").strip().lower()
        if action == "restore":
            _handle_pack_document_restore(request, pack_document)
        elif action == "save":
            _handle_pack_document_save(request, pack_document)
        else:
            messages.error(request, f"Action inconnue: {action}")
        return _redirect_template_edit(str(pack_document.id))

    return render(
        request,
        TEMPLATE_PRINT_TEMPLATE_EDIT,
        _build_edit_context(pack_document),
    )


@scan_staff_required
@require_http_methods(["POST"])
def scan_print_template_preview(request):
    _require_superuser(request)
    doc_type = (request.POST.get("doc_type") or "").strip()
    if doc_type not in DOC_LABEL_MAP:
        raise Http404("Template not found")

    layout_data, response = _parse_preview_layout(request)
    if response:
        return response

    shipment = _load_preview_shipment(request.POST.get("shipment_id") or "")
    if doc_type == "shipment_label":
        return _render_shipment_label_preview(
            request,
            layout_data=layout_data,
            shipment=shipment,
        )

    product = _load_preview_product(request.POST.get("product_id") or "")
    if doc_type == "product_label":
        return _render_product_label_preview(
            request,
            layout_data=layout_data,
            product=product,
        )
    if doc_type == "product_qr":
        return _render_product_qr_preview(
            request,
            layout_data=layout_data,
            product=product,
        )

    context = build_preview_context(doc_type, shipment=shipment)
    blocks = render_layout_from_layout(layout_data, context)
    return render(request, TEMPLATE_DYNAMIC_DOCUMENT, {"blocks": blocks})
