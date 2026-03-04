from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_FILE = Path(__file__).resolve().parent / "organization_roles_template.xlsx"

ROLE_LIST = '"shipper,recipient,correspondent,donor,transporter"'
BOOL_LIST = '"true,false"'
CHANNEL_LIST = '"email"'
CORRESPONDENT_SCOPE_LIST = (
    '"default,shipper_override,recipient_override,shipper_and_recipient_override"'
)
REVIEW_ACTION_LIST = '"resolve_binding,resolve_without_binding"'

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(vertical="center", horizontal="left")


def _apply_sheet_format(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 24
    for col_idx in range(1, ws.max_column + 1):
        values = [
            str(ws.cell(row=row_idx, column=col_idx).value or "")
            for row_idx in range(1, ws.max_row + 1)
        ]
        max_len = max(len(value) for value in values) if values else 10
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 48)


def _write_table_sheet(*, wb, title, headers, sample_rows):
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
    for row in sample_rows:
        ws.append(row)
    _apply_sheet_format(ws)
    return ws


def _add_list_validation(ws, *, col_idx, values_formula, first_row=2, last_row=500):
    validation = DataValidation(
        type="list",
        formula1=values_formula,
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(validation)
    col_letter = get_column_letter(col_idx)
    validation.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def _build_readme_sheet(wb):
    ws = wb.active
    ws.title = "README"
    ws.append(["section", "details"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    rows = [
        ("template_version", "2026-03-04"),
        (
            "purpose",
            (
                "Template de collecte des donnees pour la migration Organization Roles "
                "(expediteur, destinataire, correspondant, donateur, transporteur)."
            ),
        ),
        (
            "important",
            (
                "Ce fichier sert a la revue et a la saisie; il n y a pas encore "
                "d import automatique officiel depuis ce template."
            ),
        ),
        ("boolean_values", "Utiliser uniquement true ou false."),
        ("date_format", "Utiliser le format YYYY-MM-DD."),
        ("destination_code", "Utiliser le code IATA de l escale (ex: BKO, DLA, NKC)."),
        (
            "sheet_Organizations",
            (
                "1 ligne = 1 role pour 1 organisation. "
                "Optionnel: informations de compte portail a creer."
            ),
        ),
        (
            "sheet_ShipperScopes",
            "1 ligne = 1 portee expediteur (organisation + escale).",
        ),
        (
            "sheet_RecipientBindings",
            "1 ligne = 1 liaison active destinataire <-> expediteur <-> escale.",
        ),
        (
            "sheet_Correspondents",
            (
                "1 ligne = 1 correspondant pour une escale (default ou override "
                "par expediteur/destinataire)."
            ),
        ),
        (
            "sheet_OrganizationContacts",
            (
                "Contacts operationnels par role. "
                "Recommande: 1 contact principal avec email par role actif."
            ),
        ),
        (
            "sheet_MigrationReview",
            (
                "Utiliser pour traiter les dossiers en file de revue "
                "(reason_code, proposition de mapping, action)."
            ),
        ),
    ]
    for row in rows:
        ws.append(row)
    _apply_sheet_format(ws)


def build_workbook():
    wb = Workbook()
    _build_readme_sheet(wb)

    organizations = _write_table_sheet(
        wb=wb,
        title="Organizations",
        headers=[
            "row_id",
            "organization_name",
            "organization_external_id",
            "role",
            "role_active",
            "create_portal_profile",
            "portal_user_email",
            "portal_username",
            "profile_must_change_password",
            "notes",
        ],
        sample_rows=[
            [
                "ORG-001",
                "Medecins du Monde",
                "MDM-001",
                "shipper",
                "true",
                "true",
                "ops@mdm.org",
                "mdm.ops",
                "true",
                "Creation compte portail expediteur",
            ]
        ],
    )
    _add_list_validation(organizations, col_idx=4, values_formula=ROLE_LIST)
    _add_list_validation(organizations, col_idx=5, values_formula=BOOL_LIST)
    _add_list_validation(organizations, col_idx=6, values_formula=BOOL_LIST)
    _add_list_validation(organizations, col_idx=9, values_formula=BOOL_LIST)

    shipper_scopes = _write_table_sheet(
        wb=wb,
        title="ShipperScopes",
        headers=[
            "row_id",
            "organization_name",
            "destination_iata",
            "destination_city",
            "all_destinations",
            "is_active",
            "valid_from",
            "valid_to",
            "notes",
        ],
        sample_rows=[
            [
                "SS-001",
                "Medecins du Monde",
                "BKO",
                "Bamako",
                "false",
                "true",
                "2026-03-04",
                "",
                "",
            ]
        ],
    )
    _add_list_validation(shipper_scopes, col_idx=5, values_formula=BOOL_LIST)
    _add_list_validation(shipper_scopes, col_idx=6, values_formula=BOOL_LIST)

    recipient_bindings = _write_table_sheet(
        wb=wb,
        title="RecipientBindings",
        headers=[
            "row_id",
            "recipient_organization_name",
            "shipper_organization_name",
            "destination_iata",
            "destination_city",
            "is_active",
            "valid_from",
            "valid_to",
            "notes",
        ],
        sample_rows=[
            [
                "RB-001",
                "Hopital Gabriel Toure",
                "Medecins du Monde",
                "BKO",
                "Bamako",
                "true",
                "2026-03-04",
                "",
                "",
            ]
        ],
    )
    _add_list_validation(recipient_bindings, col_idx=6, values_formula=BOOL_LIST)

    correspondents = _write_table_sheet(
        wb=wb,
        title="Correspondents",
        headers=[
            "row_id",
            "correspondent_organization_name",
            "destination_iata",
            "destination_city",
            "scope_type",
            "shipper_organization_name",
            "recipient_organization_name",
            "is_active",
            "notes",
        ],
        sample_rows=[
            [
                "CO-001",
                "Correspondant ASF Douala",
                "DLA",
                "Douala",
                "default",
                "",
                "",
                "true",
                "",
            ]
        ],
    )
    _add_list_validation(correspondents, col_idx=5, values_formula=CORRESPONDENT_SCOPE_LIST)
    _add_list_validation(correspondents, col_idx=8, values_formula=BOOL_LIST)

    org_contacts = _write_table_sheet(
        wb=wb,
        title="OrganizationContacts",
        headers=[
            "row_id",
            "organization_name",
            "role",
            "contact_first_name",
            "contact_last_name",
            "contact_email",
            "contact_phone",
            "is_primary",
            "is_active",
            "notification_channel",
            "notify_shipment_status_updated",
            "notify_shipment_delivered",
            "notify_shipment_tracking_updated",
            "notify_order_document_requested",
            "destination_iata_filter",
            "shipper_organization_filter",
            "recipient_organization_filter",
            "notes",
        ],
        sample_rows=[
            [
                "OC-001",
                "Medecins du Monde",
                "shipper",
                "Awa",
                "Diallo",
                "awa.diallo@mdm.org",
                "+223000000",
                "true",
                "true",
                "email",
                "true",
                "true",
                "true",
                "false",
                "BKO",
                "",
                "",
                "Contact principal operations",
            ]
        ],
    )
    _add_list_validation(org_contacts, col_idx=3, values_formula=ROLE_LIST)
    _add_list_validation(org_contacts, col_idx=8, values_formula=BOOL_LIST)
    _add_list_validation(org_contacts, col_idx=9, values_formula=BOOL_LIST)
    _add_list_validation(org_contacts, col_idx=10, values_formula=CHANNEL_LIST)
    for col_idx in (11, 12, 13, 14):
        _add_list_validation(org_contacts, col_idx=col_idx, values_formula=BOOL_LIST)

    migration_review = _write_table_sheet(
        wb=wb,
        title="MigrationReview",
        headers=[
            "row_id",
            "recipient_organization_name",
            "reason_code",
            "proposed_shipper_organization_name",
            "proposed_destination_iata",
            "proposed_destination_city",
            "resolution_action",
            "resolution_note",
        ],
        sample_rows=[
            [
                "MR-001",
                "Destinataire Z",
                "missing_destination",
                "Medecins du Monde",
                "BKO",
                "Bamako",
                "resolve_binding",
                "Associer au shipper MDM sur BKO",
            ]
        ],
    )
    _add_list_validation(migration_review, col_idx=7, values_formula=REVIEW_ACTION_LIST)

    wb.save(OUTPUT_FILE)


if __name__ == "__main__":
    build_workbook()
    print(f"Template generated: {OUTPUT_FILE}")
